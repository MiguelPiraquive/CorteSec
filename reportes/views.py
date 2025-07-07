"""
Sistema de Reportes Profesional para la Gestión Empresarial

Este módulo proporciona:
- Generación de reportes en PDF con gráficos
- Exportación a Excel con formateo profesional
- Reportes programados y automatizados
- Plantillas personalizables
- Análisis estadísticos avanzados
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from django.utils import timezone
from django.db.models import Sum, Count, Avg, Q
from datetime import datetime, timedelta
from decimal import Decimal

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.drawing.image import Image

# from reportlab.lib import colors
# from reportlab.lib.pagesizes import letter, A4
# from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
# from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
# from reportlab.lib.units import inch
# from reportlab.graphics.shapes import Drawing
# from reportlab.graphics.charts.linecharts import HorizontalLineChart
# from reportlab.graphics.charts.barcharts import VerticalBarChart
# from reportlab.graphics.charts.piecharts import Pie

from payroll.models import Empleado, Nomina
from prestamos.models import Prestamo, PagoPrestamo
from roles.models import Rol
from dashboard.models import Contractor, Project


@login_required
def lista_reportes(request):
    """
    Vista principal para listar y gestionar reportes
    """
    reportes_disponibles = [
        {
            'id': 'empleados',
            'nombre': 'Reporte de Empleados',
            'descripcion': 'Listado completo de empleados con estadísticas',
            'icono': 'fas fa-users',
            'color': 'bg-blue-500',
            'formatos': ['pdf', 'excel']
        },
        {
            'id': 'nominas',
            'nombre': 'Reporte de Nóminas',
            'descripcion': 'Análisis de pagos y nóminas por período',
            'icono': 'fas fa-money-bill-wave',
            'color': 'bg-green-500',
            'formatos': ['pdf', 'excel']
        },
        {
            'id': 'prestamos',
            'nombre': 'Reporte de Préstamos',
            'descripcion': 'Estado de préstamos y cuotas pendientes',
            'icono': 'fas fa-hand-holding-usd',
            'color': 'bg-yellow-500',
            'formatos': ['pdf', 'excel']
        },
        {
            'id': 'proyectos',
            'nombre': 'Reporte de Proyectos',
            'descripcion': 'Estado y progreso de proyectos activos',
            'icono': 'fas fa-project-diagram',
            'color': 'bg-purple-500',
            'formatos': ['pdf', 'excel']
        },
        {
            'id': 'dashboard',
            'nombre': 'Dashboard Ejecutivo',
            'descripcion': 'Resumen ejecutivo con métricas clave',
            'icono': 'fas fa-chart-line',
            'color': 'bg-indigo-500',
            'formatos': ['pdf']
        },
        {
            'id': 'auditoria',
            'nombre': 'Reporte de Auditoría',
            'descripcion': 'Registro de accesos y cambios del sistema',
            'icono': 'fas fa-shield-alt',
            'color': 'bg-red-500',
            'formatos': ['pdf', 'excel']
        }
    ]
    
    context = {
        'title': 'Centro de Reportes',
        'reportes_disponibles': reportes_disponibles,
    }
    
    return render(request, 'reportes/lista.html', context)


# ============ REPORTES PDF ============

@login_required
def reporte_empleados_pdf(request):
    """
    Genera reporte de empleados en PDF
    """
    # Funcionalidad temporalmente deshabilitada - requiere reportlab
    from django.http import HttpResponse
    return HttpResponse("Funcionalidad en desarrollo - requiere instalación de reportlab", content_type='text/plain')
    
    # TODO: Uncomment when reportlab is installed
    """
    # Obtener estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2563eb'),
        alignment=1,
        spaceAfter=30
    )
    
    # Elementos del reporte
    story = []
    
    # Título
    title = Paragraph(f"Reporte de Empleados - {timezone.now().strftime('%d/%m/%Y')}", title_style)
    story.append(title)
    
    # Resumen estadístico
    total_empleados = Empleado.objects.count()
    empleados_activos = Empleado.objects.filter(activo=True).count()
    promedio_salario = Empleado.objects.filter(activo=True).aggregate(
        promedio=Avg('salario_base')
    )['promedio'] or Decimal('0')
    
    resumen_data = [
        ['Métrica', 'Valor'],
        ['Total de Empleados', str(total_empleados)],
        ['Empleados Activos', str(empleados_activos)],
        ['Empleados Inactivos', str(total_empleados - empleados_activos)],
        ['Salario Promedio', f'${promedio_salario:,.2f}'],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[3*inch, 2*inch])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(resumen_table)
    story.append(Spacer(1, 20))
    
    # Lista de empleados
    empleados = Empleado.objects.all().order_by('nombre', 'apellido')
    
    empleados_data = [['Nombre', 'Cargo', 'Salario', 'Fecha Contratación', 'Estado']]
    
    for empleado in empleados:
        empleados_data.append([
            f"{empleado.nombre} {empleado.apellido}",
            empleado.cargo.nombre if empleado.cargo else 'Sin cargo',
            f"${empleado.salario_base:,.2f}",
            empleado.fecha_contratacion.strftime('%d/%m/%Y'),
            'Activo' if empleado.activo else 'Inactivo'
        ])
    
    empleados_table = Table(empleados_data, colWidths=[2*inch, 1.5*inch, 1*inch, 1*inch, 0.8*inch])
    empleados_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    story.append(Paragraph("Listado de Empleados", styles['Heading2']))
    story.append(Spacer(1, 12))
    story.append(empleados_table)
    
    # Construir PDF
    doc.build(story)
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response
    """


# ============ REPORTES EXCEL ============

@login_required
def reporte_empleados_excel(request):
    """
    Genera reporte de empleados en Excel con formateo profesional
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_empleados_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empleados"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título principal
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Reporte de Empleados - {timezone.now().strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(size=16, bold=True, color="2563EB")
    ws['A1'].alignment = center_alignment
    
    # Headers
    headers = ['Nombre Completo', 'Cargo', 'Salario Base', 'Fecha Contratación', 'Estado', 'Teléfono']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Datos de empleados
    empleados = Empleado.objects.all().order_by('nombre', 'apellido')
    
    for row, empleado in enumerate(empleados, 4):
        data = [
            f"{empleado.nombre} {empleado.apellido}",
            empleado.cargo.nombre if empleado.cargo else 'Sin cargo',
            empleado.salario_base,
            empleado.fecha_contratacion,
            'Activo' if empleado.activo else 'Inactivo',
            empleado.telefono or 'No especificado'
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = center_alignment
            
            # Formateo especial para diferentes columnas
            if col == 3:  # Salario
                cell.number_format = '$#,##0.00'
            elif col == 4:  # Fecha
                cell.number_format = 'DD/MM/YYYY'
            elif col == 5:  # Estado
                if value == 'Activo':
                    cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                else:
                    cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
    
    # Ajustar ancho de columnas
    column_widths = [25, 20, 15, 18, 12, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    wb.save(response)
    return response


# ============ API PARA REPORTES ============

@login_required
def api_generar_reporte(request):
    """
    API para generar reportes dinámicamente
    """
    tipo_reporte = request.POST.get('tipo')
    formato = request.POST.get('formato', 'pdf')
    
    if not tipo_reporte:
        return JsonResponse({'error': 'Tipo de reporte requerido'}, status=400)
    
    try:
        if formato == 'pdf':
            if tipo_reporte == 'empleados':
                return reporte_empleados_pdf(request)
            # Agregar más tipos según necesidad
        
        elif formato == 'excel':
            if tipo_reporte == 'empleados':
                return reporte_empleados_excel(request)
            # Agregar más tipos según necesidad
        
        return JsonResponse({'error': 'Tipo de reporte o formato no soportado'}, status=400)
        
    except Exception as e:
        return JsonResponse({'error': f'Error al generar reporte: {str(e)}'}, status=500)
