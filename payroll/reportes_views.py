"""
Sistema de Reportes Avanzados para Nómina Electrónica
Exportación a Excel, PDF, CSV con filtros y agrupaciones
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from django.http import HttpResponse
from django.db.models import Sum, Count, Avg, Q
from django.utils import timezone
from datetime import timedelta
import csv
import json

# Para Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from payroll.models import (
    NominaElectronica, Nomina,
    DevengadoNominaElectronica, DeduccionNominaElectronica
)


class ReportesViewSet(viewsets.ViewSet):
    """ViewSet para generación de reportes"""
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def nominas_excel(self, request):
        """
        Exporta nóminas a Excel
        
        GET /api/payroll/reportes/nominas_excel/
        
        Query params:
        - fecha_desde: YYYY-MM-DD
        - fecha_hasta: YYYY-MM-DD
        - estado: Estado de nómina
        - empleado_id: ID de empleado específico
        """
        if not EXCEL_AVAILABLE:
            return Response(
                {'error': 'openpyxl no está instalado'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        organization = request.user.organization
        
        # Filtros
        filtro = Q(organization=organization)
        
        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')
        estado = request.query_params.get('estado')
        empleado_id = request.query_params.get('empleado_id')
        
        if fecha_desde:
            filtro &= Q(fecha_emision__gte=fecha_desde)
        if fecha_hasta:
            filtro &= Q(fecha_emision__lte=fecha_hasta)
        if estado:
            filtro &= Q(estado=estado)
        if empleado_id:
            filtro &= Q(nomina__empleado_id=empleado_id)
        
        nominas = NominaElectronica.objects.filter(filtro).select_related(
            'nomina', 'nomina__empleado'
        ).order_by('-fecha_emision')
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Nóminas Electrónicas"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = [
            'Número', 'Empleado', 'Documento', 'Fecha Emisión',
            'Periodo Inicio', 'Periodo Fin', 'Total Devengado',
            'Total Deducciones', 'Neto a Pagar', 'Estado',
            'CUNE', 'Fecha Validación DIAN'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Datos
        for row_num, nomina in enumerate(nominas, 2):
            ws.cell(row=row_num, column=1, value=nomina.numero_documento)
            ws.cell(row=row_num, column=2, value=nomina.nomina.empleado.nombre_completo)
            ws.cell(row=row_num, column=3, value=nomina.nomina.empleado.documento)
            ws.cell(row=row_num, column=4, value=nomina.fecha_emision.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=5, value=nomina.nomina.periodo_inicio.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=6, value=nomina.nomina.periodo_fin.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=7, value=float(nomina.nomina.total_devengado))
            ws.cell(row=row_num, column=8, value=float(nomina.nomina.total_deducciones))
            ws.cell(row=row_num, column=9, value=float(nomina.nomina.neto_pagar))
            ws.cell(row=row_num, column=10, value=nomina.get_estado_display())
            ws.cell(row=row_num, column=11, value=nomina.cune)
            ws.cell(row=row_num, column=12, value=nomina.fecha_validacion_dian.strftime('%Y-%m-%d %H:%M') if nomina.fecha_validacion_dian else '')
        
        # Ajustar anchos de columna
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="nominas_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'])
    def nominas_csv(self, request):
        """
        Exporta nóminas a CSV
        
        GET /api/payroll/reportes/nominas_csv/
        """
        organization = request.user.organization
        
        # Mismo filtro que Excel
        filtro = Q(organization=organization)
        
        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')
        
        if fecha_desde:
            filtro &= Q(fecha_emision__gte=fecha_desde)
        if fecha_hasta:
            filtro &= Q(fecha_emision__lte=fecha_hasta)
        
        nominas = NominaElectronica.objects.filter(filtro).select_related(
            'nomina', 'nomina__empleado'
        ).order_by('-fecha_emision')
        
        # Crear CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="nominas_{timezone.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        
        # Encabezados
        writer.writerow([
            'Número', 'Empleado', 'Documento', 'Fecha Emisión',
            'Total Devengado', 'Total Deducciones', 'Neto a Pagar',
            'Estado', 'CUNE'
        ])
        
        # Datos
        for nomina in nominas:
            writer.writerow([
                nomina.numero_documento,
                nomina.nomina.empleado.nombre_completo,
                nomina.nomina.empleado.documento,
                nomina.fecha_emision.strftime('%Y-%m-%d'),
                float(nomina.nomina.total_devengado),
                float(nomina.nomina.total_deducciones),
                float(nomina.nomina.neto_pagar),
                nomina.get_estado_display(),
                nomina.cune
            ])
        
        return response
    
    @action(detail=False, methods=['get'])
    def reporte_mensual_excel(self, request):
        """
        Reporte mensual consolidado en Excel
        
        GET /api/payroll/reportes/reporte_mensual_excel/
        
        Query params:
        - anio: Año
        - mes: Mes
        """
        if not EXCEL_AVAILABLE:
            return Response(
                {'error': 'openpyxl no está instalado'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        organization = request.user.organization
        anio = int(request.query_params.get('anio', timezone.now().year))
        mes = int(request.query_params.get('mes', timezone.now().month))
        
        # Obtener nóminas del mes
        nominas = Nomina.objects.filter(
            organization=organization,
            periodo_inicio__year=anio,
            periodo_inicio__month=mes
        ).select_related('empleado', 'empleado__cargo')
        
        # Crear workbook
        wb = Workbook()
        
        # Hoja 1: Resumen
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        
        ws_resumen['A1'] = f"REPORTE MENSUAL DE NÓMINA - {timezone.datetime(anio, mes, 1).strftime('%B %Y').upper()}"
        ws_resumen['A1'].font = Font(bold=True, size=14)
        
        ws_resumen['A3'] = "Total Empleados:"
        ws_resumen['B3'] = nominas.count()
        
        totales = nominas.aggregate(
            devengado=Sum('ingreso_real_periodo'),
            deducciones=Sum('otras_deducciones'),
            neto=Sum('ingreso_real_periodo')
        )
        
        ws_resumen['A4'] = "Total Devengado:"
        ws_resumen['B4'] = float(totales['devengado'] or 0)
        ws_resumen['B4'].number_format = '$#,##0.00'
        
        ws_resumen['A5'] = "Total Deducciones:"
        ws_resumen['B5'] = float(totales['deducciones'] or 0)
        ws_resumen['B5'].number_format = '$#,##0.00'
        
        ws_resumen['A6'] = "Total Neto:"
        ws_resumen['B6'] = float(totales['neto'] or 0)
        ws_resumen['B6'].number_format = '$#,##0.00'
        ws_resumen['B6'].font = Font(bold=True)
        
        # Hoja 2: Detalle por empleado
        ws_detalle = wb.create_sheet("Detalle por Empleado")
        
        headers = [
            'Empleado', 'Documento', 'Cargo', 'Salario Base',
            'Total Devengado', 'Salud', 'Pensión', 'Otras Deducciones',
            'Total Deducciones', 'Neto a Pagar'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_detalle.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        for row_num, nomina in enumerate(nominas, 2):
            ws_detalle.cell(row=row_num, column=1, value=nomina.empleado.nombre_completo)
            ws_detalle.cell(row=row_num, column=2, value=nomina.empleado.documento)
            ws_detalle.cell(row=row_num, column=3, value=nomina.empleado.cargo.nombre if nomina.empleado.cargo else 'N/A')
            ws_detalle.cell(row=row_num, column=4, value=float(nomina.contrato.salario if nomina.contrato else 0))
            ws_detalle.cell(row=row_num, column=5, value=float(nomina.ingreso_real_periodo))
            ws_detalle.cell(row=row_num, column=6, value=float(nomina.deduccion_salud))
            ws_detalle.cell(row=row_num, column=7, value=float(nomina.deduccion_pension))
            ws_detalle.cell(row=row_num, column=8, value=float(nomina.otras_deducciones))
            ws_detalle.cell(row=row_num, column=9, value=float(nomina.otras_deducciones))
            ws_detalle.cell(row=row_num, column=10, value=float(nomina.ingreso_real_periodo))
            
            # Formato moneda
            for col in [4, 5, 6, 7, 8, 9, 10]:
                ws_detalle.cell(row=row_num, column=col).number_format = '$#,##0.00'
        
        # Crear respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_mensual_{anio}_{mes:02d}.xlsx"'
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'])
    def reporte_anual(self, request):
        """
        Reporte anual consolidado (JSON)
        
        GET /api/payroll/reportes/reporte_anual/
        
        Query params:
        - anio: Año (default: año actual)
        """
        organization = request.user.organization
        anio = int(request.query_params.get('anio', timezone.now().year))
        
        # Nóminas del año
        nominas = Nomina.objects.filter(
            organization=organization,
            periodo_inicio__year=anio
        )
        
        # Totales por mes
        from django.db.models.functions import TruncMonth
        por_mes = nominas.annotate(
            mes=TruncMonth('periodo_inicio')
        ).values('mes').annotate(
            empleados=Count('empleado', distinct=True),
            devengado=Sum('ingreso_real_periodo'),
            deducciones=Sum('otras_deducciones'),
            neto=Sum('ingreso_real_periodo')
        ).order_by('mes')
        
        # Estadísticas nómina electrónica
        nominas_elect = NominaElectronica.objects.filter(
            organization=organization,
            fecha_emision__year=anio
        )
        
        stats_electronica = nominas_elect.values('estado').annotate(
            cantidad=Count('id')
        )
        
        data = {
            'anio': anio,
            'totales': {
                'empleados_unicos': nominas.values('empleado').distinct().count(),
                'nominas_procesadas': nominas.count(),
                'total_devengado': float(nominas.aggregate(t=Sum('ingreso_real_periodo'))['t'] or 0),
                'total_deducciones': float(nominas.aggregate(t=Sum('otras_deducciones'))['t'] or 0),
                'total_neto': float(nominas.aggregate(t=Sum('ingreso_real_periodo'))['t'] or 0)
            },
            'por_mes': [
                {
                    'mes': item['mes'].strftime('%Y-%m'),
                    'empleados': item['empleados'],
                    'devengado': float(item['devengado']),
                    'deducciones': float(item['deducciones']),
                    'neto': float(item['neto'])
                }
                for item in por_mes
            ],
            'estadisticas_electronica': {
                'total': nominas_elect.count(),
                'por_estado': {item['estado']: item['cantidad'] for item in stats_electronica}
            }
        }
        
        return Response(data)
    
    @action(detail=False, methods=['get'])
    def certificado_ingresos_pdf(self, request):
        """
        Genera certificado de ingresos en PDF
        
        GET /api/payroll/reportes/certificado_ingresos_pdf/
        
        Query params:
        - empleado_id: ID del empleado
        - anio: Año
        """
        # TODO: Implementar con reportlab
        return Response({
            'mensaje': 'Función en desarrollo',
            'status': 'pendiente'
        })
