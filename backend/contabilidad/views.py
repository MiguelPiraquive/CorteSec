from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils.encoding import escape_uri_path
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.views.decorators.http import require_http_methods
from django.urls import reverse_lazy
from django.utils import timezone
from django.contrib.auth import get_user_model
from decimal import Decimal
from io import BytesIO
import json
from datetime import datetime, timedelta

from .models import PlanCuentas, ComprobanteContable, MovimientoContable, FlujoCaja
from .forms import (
    PlanCuentasForm, ComprobanteContableForm, MovimientoContableForm, 
    FlujoCajaForm, MovimientoFormSet, FiltroContableForm,
    BalanceGeneralForm, EstadoResultadosForm
)

User = get_user_model()


# ===================== PLAN DE CUENTAS =====================

@login_required
def plan_cuentas_list(request):
    """Lista del plan de cuentas con filtros y paginación."""
    queryset = PlanCuentas.objects.select_related('cuenta_padre').all()
    
    # Filtros
    search = request.GET.get('search', '')
    tipo_cuenta = request.GET.get('tipo_cuenta', '')
    activa = request.GET.get('activa', '')
    nivel = request.GET.get('nivel', '')
    
    if search:
        queryset = queryset.filter(
            Q(codigo__icontains=search) |
            Q(nombre__icontains=search) |
            Q(descripcion__icontains=search)
        )
    
    if tipo_cuenta:
        queryset = queryset.filter(tipo_cuenta=tipo_cuenta)
    
    if activa:
        queryset = queryset.filter(activa=activa == 'true')
    
    if nivel:
        queryset = queryset.filter(nivel=nivel)
    
    # Orden jerárquico
    queryset = queryset.order_by('codigo')
    
    # Paginación
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    cuentas = paginator.get_page(page_number)
    
    # Estadísticas
    stats = {
        'total': PlanCuentas.objects.count(),
        'activas': PlanCuentas.objects.filter(activa=True).count(),
        'inactivas': PlanCuentas.objects.filter(activa=False).count(),
        'con_movimientos': PlanCuentas.objects.filter(acepta_movimientos=True).count(),
    }
    
    context = {
        'cuentas': cuentas,
        'stats': stats,
        'search': search,
        'tipo_cuenta': tipo_cuenta,
        'activa': activa,
        'nivel': nivel,
        'tipos_cuenta': PlanCuentas.TIPO_CUENTA_CHOICES,
        'niveles_disponibles': range(1, 6),  # Hasta 5 niveles
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'contabilidad/plan_cuentas/table.html', context)
    
    return render(request, 'contabilidad/plan_cuentas/list.html', context)


@login_required
def plan_cuentas_detail(request, pk):
    """Detalle de una cuenta contable."""
    cuenta = get_object_or_404(PlanCuentas, pk=pk)
    subcuentas = cuenta.subcuentas.filter(activa=True)
    
    # Movimientos recientes
    movimientos = MovimientoContable.objects.filter(
        cuenta=cuenta
    ).select_related('comprobante').order_by('-comprobante__fecha')[:10]
    
    context = {
        'cuenta': cuenta,
        'subcuentas': subcuentas,
        'movimientos': movimientos,
        'saldo_actual': cuenta.saldo_actual,
    }
    return render(request, 'contabilidad/plan_cuentas/detail.html', context)


@login_required
def plan_cuentas_create(request):
    """Crear una nueva cuenta contable."""
    if request.method == 'POST':
        form = PlanCuentasForm(request.POST)
        if form.is_valid():
            cuenta = form.save()
            messages.success(request, f'Cuenta "{cuenta.codigo} - {cuenta.nombre}" creada exitosamente.')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Cuenta "{cuenta.codigo} - {cuenta.nombre}" creada exitosamente.',
                    'redirect_url': reverse_lazy('contabilidad:plan_cuentas_list')
                })
            
            return redirect('contabilidad:plan_cuentas_list')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
    else:
        form = PlanCuentasForm()
    
    context = {'form': form, 'object': None}
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'contabilidad/plan_cuentas/form.html', context)
    
    return render(request, 'contabilidad/plan_cuentas/form.html', context)


@login_required
def plan_cuentas_edit(request, pk):
    """Editar una cuenta contable."""
    cuenta = get_object_or_404(PlanCuentas, pk=pk)
    
    if request.method == 'POST':
        form = PlanCuentasForm(request.POST, instance=cuenta)
        if form.is_valid():
            cuenta = form.save()
            messages.success(request, f'Cuenta "{cuenta.codigo} - {cuenta.nombre}" actualizada exitosamente.')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Cuenta "{cuenta.codigo} - {cuenta.nombre}" actualizada exitosamente.',
                    'redirect_url': reverse_lazy('contabilidad:plan_cuentas_list')
                })
            
            return redirect('contabilidad:plan_cuentas_list')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
    else:
        form = PlanCuentasForm(instance=cuenta)
    
    context = {'form': form, 'object': cuenta}
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'contabilidad/plan_cuentas/form.html', context)
    
    return render(request, 'contabilidad/plan_cuentas/form.html', context)


@login_required
@require_http_methods(["DELETE"])
def plan_cuentas_delete(request, pk):
    """Eliminar una cuenta contable."""
    cuenta = get_object_or_404(PlanCuentas, pk=pk)
    
    # Verificar si tiene movimientos
    if cuenta.movimientos.exists():
        message = 'No se puede eliminar una cuenta que tiene movimientos contables.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': message
            })
        messages.error(request, message)
        return redirect('contabilidad:plan_cuentas_list')
    
    # Verificar si tiene subcuentas
    if cuenta.subcuentas.exists():
        message = 'No se puede eliminar una cuenta que tiene subcuentas.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': message
            })
        messages.error(request, message)
        return redirect('contabilidad:plan_cuentas_list')
    
    cuenta_nombre = f"{cuenta.codigo} - {cuenta.nombre}"
    cuenta.delete()
    
    message = f'Cuenta "{cuenta_nombre}" eliminada exitosamente.'
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'message': message
        })
    
    messages.success(request, message)
    return redirect('contabilidad:plan_cuentas_list')


# ===================== COMPROBANTES CONTABLES =====================

@login_required
def comprobantes_list(request):
    """Lista de comprobantes contables con filtros y paginación."""
    queryset = ComprobanteContable.objects.select_related('creado_por').all()
    
    # Filtros
    search = request.GET.get('search', '')
    tipo = request.GET.get('tipo', '')
    estado = request.GET.get('estado', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if search:
        queryset = queryset.filter(
            Q(numero__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(referencia__icontains=search)
        )
    
    if tipo:
        queryset = queryset.filter(tipo_comprobante=tipo)
    
    if estado:
        queryset = queryset.filter(estado=estado)
    
    if fecha_desde:
        queryset = queryset.filter(fecha__gte=fecha_desde)
    
    if fecha_hasta:
        queryset = queryset.filter(fecha__lte=fecha_hasta)
    
    # Orden
    queryset = queryset.order_by('-fecha', '-numero')
    
    # Paginación
    paginator = Paginator(queryset, 15)
    page_number = request.GET.get('page')
    comprobantes = paginator.get_page(page_number)
    
    # Estadísticas
    stats = {
        'total': ComprobanteContable.objects.count(),
        'confirmados': ComprobanteContable.objects.filter(estado='contabilizado').count(),
        'pendientes': ComprobanteContable.objects.filter(estado='borrador').count(),
        'valor_total': ComprobanteContable.objects.filter(estado='contabilizado').aggregate(
            total=Sum('total_debito')
        )['total'] or 0
    }
    
    context = {
        'comprobantes': comprobantes,
        'stats': stats,
        'search': search,
        'tipo': tipo,
        'estado': estado,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tipos_comprobante': ComprobanteContable.TIPO_COMPROBANTE_CHOICES,
        'estados_comprobante': ComprobanteContable.ESTADO_CHOICES,
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'contabilidad/comprobantes/table.html', context)
    
    return render(request, 'contabilidad/comprobantes/list.html', context)


@login_required
def comprobante_detail(request, pk):
    """Detalle de un comprobante contable."""
    comprobante = get_object_or_404(ComprobanteContable, pk=pk)
    movimientos = comprobante.movimientos.select_related('cuenta').all()
    
    # Calcular totales
    total_debitos = movimientos.aggregate(total=Sum('valor_debito'))['total'] or 0
    total_creditos = movimientos.aggregate(total=Sum('valor_credito'))['total'] or 0
    
    context = {
        'comprobante': comprobante,
        'movimientos': movimientos,
        'total_debitos': total_debitos,
        'total_creditos': total_creditos,
    }
    return render(request, 'contabilidad/comprobantes/detail.html', context)


@login_required
def comprobante_create(request):
    """Crear un nuevo comprobante contable."""
    if request.method == 'POST':
        form = ComprobanteContableForm(request.POST)
        movimientos_data = json.loads(request.POST.get('movimientos', '[]'))
        
        if form.is_valid() and movimientos_data:
            # Validar que el comprobante esté balanceado
            total_debitos = sum(Decimal(str(mov.get('debito', 0))) for mov in movimientos_data)
            total_creditos = sum(Decimal(str(mov.get('credito', 0))) for mov in movimientos_data)
            
            if total_debitos != total_creditos:
                return JsonResponse({
                    'success': False,
                    'message': 'El comprobante debe estar balanceado (débitos = créditos)'
                })
            
            comprobante = form.save(commit=False)
            comprobante.creado_por = request.user
            comprobante.total_debito = total_debitos
            comprobante.total_credito = total_creditos
            comprobante.save()
            
            # Crear movimientos
            for mov_data in movimientos_data:
                MovimientoContable.objects.create(
                    comprobante=comprobante,
                    cuenta_id=mov_data['cuenta_id'],
                    descripcion=mov_data.get('descripcion', comprobante.descripcion),
                    valor_debito=Decimal(str(mov_data.get('debito', 0))),
                    valor_credito=Decimal(str(mov_data.get('credito', 0)))
                )
            
            messages.success(request, f'Comprobante {comprobante.numero} creado exitosamente.')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Comprobante {comprobante.numero} creado exitosamente.'
                })
            
            return redirect('contabilidad:comprobante_detail', pk=comprobante.pk)
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
    else:
        form = ComprobanteContableForm()
    
    context = {'form': form}
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'contabilidad/comprobantes/form.html', context)
    
    return render(request, 'contabilidad/comprobantes/create.html', context)


@login_required
def comprobante_edit(request, pk):
    """Editar un comprobante contable (solo si está en borrador)."""
    comprobante = get_object_or_404(ComprobanteContable, pk=pk)
    
    if comprobante.estado != 'borrador':
        messages.error(request, 'Solo se pueden editar comprobantes en estado borrador.')
        return redirect('contabilidad:comprobante_detail', pk=pk)
    
    if request.method == 'POST':
        form = ComprobanteContableForm(request.POST, instance=comprobante)
        movimientos_data = json.loads(request.POST.get('movimientos', '[]'))
        
        if form.is_valid() and movimientos_data:
            # Validar balance
            total_debitos = sum(Decimal(str(mov.get('debito', 0))) for mov in movimientos_data)
            total_creditos = sum(Decimal(str(mov.get('credito', 0))) for mov in movimientos_data)
            
            if total_debitos != total_creditos:
                return JsonResponse({
                    'success': False,
                    'message': 'El comprobante debe estar balanceado'
                })
            
            comprobante = form.save(commit=False)
            comprobante.total_debito = total_debitos
            comprobante.total_credito = total_creditos
            comprobante.save()
            
            # Eliminar movimientos existentes y crear nuevos
            comprobante.movimientos.all().delete()
            for mov_data in movimientos_data:
                MovimientoContable.objects.create(
                    comprobante=comprobante,
                    cuenta_id=mov_data['cuenta_id'],
                    descripcion=mov_data.get('descripcion', comprobante.descripcion),
                    valor_debito=Decimal(str(mov_data.get('debito', 0))),
                    valor_credito=Decimal(str(mov_data.get('credito', 0)))
                )
            
            messages.success(request, f'Comprobante {comprobante.numero} actualizado exitosamente.')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            
            return redirect('contabilidad:comprobante_detail', pk=comprobante.pk)
    else:
        form = ComprobanteContableForm(instance=comprobante)
        movimientos_json = json.dumps([
            {
                'cuenta_id': mov.cuenta.id,
                'debito': float(mov.valor_debito),
                'credito': float(mov.valor_credito),
                'descripcion': mov.descripcion
            } for mov in comprobante.movimientos.all()
        ])
    
    context = {
        'form': form, 
        'comprobante': comprobante,
        'movimientos_json': movimientos_json
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'contabilidad/comprobantes/form.html', context)
    
    return render(request, 'contabilidad/comprobantes/edit.html', context)


@login_required
@require_http_methods(["POST"])
def comprobante_confirm(request, pk):
    """Confirmar un comprobante contable."""
    comprobante = get_object_or_404(ComprobanteContable, pk=pk)
    
    if comprobante.estado != 'borrador':
        return JsonResponse({
            'success': False,
            'message': 'Solo se pueden confirmar comprobantes en borrador'
        })
    
    # Validar que esté balanceado
    total_debitos = comprobante.movimientos.aggregate(total=Sum('valor_debito'))['total'] or 0
    total_creditos = comprobante.movimientos.aggregate(total=Sum('valor_credito'))['total'] or 0
    
    if total_debitos != total_creditos:
        return JsonResponse({
            'success': False,
            'message': 'El comprobante debe estar balanceado para confirmar'
        })
    
    comprobante.estado = 'contabilizado'
    comprobante.contabilizado_por = request.user
    comprobante.fecha_contabilizacion = timezone.now()
    comprobante.save()
    
    messages.success(request, f'Comprobante {comprobante.numero} confirmado exitosamente.')
    
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["DELETE"])
def comprobante_delete(request, pk):
    """Eliminar un comprobante contable (solo borradores)."""
    comprobante = get_object_or_404(ComprobanteContable, pk=pk)
    
    if comprobante.estado != 'borrador':
        return JsonResponse({
            'success': False,
            'message': 'Solo se pueden eliminar comprobantes en borrador'
        })
    
    numero = comprobante.numero
    comprobante.delete()
    
    messages.success(request, f'Comprobante {numero} eliminado exitosamente.')
    
    return JsonResponse({'success': True})


# ===================== MOVIMIENTOS CONTABLES =====================

@login_required
def movimientos_list(request):
    """Lista de movimientos contables con filtros."""
    queryset = MovimientoContable.objects.select_related(
        'comprobante', 'cuenta'
    ).all()
    
    # Filtros
    search = request.GET.get('search', '')
    cuenta = request.GET.get('cuenta', '')
    tipo_comprobante = request.GET.get('tipo_comprobante', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if search:
        queryset = queryset.filter(
            Q(cuenta__codigo__icontains=search) |
            Q(cuenta__nombre__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(comprobante__numero__icontains=search)
        )
    
    if cuenta:
        queryset = queryset.filter(cuenta_id=cuenta)
    
    if tipo_comprobante:
        queryset = queryset.filter(comprobante__tipo_comprobante=tipo_comprobante)
    
    if fecha_desde:
        queryset = queryset.filter(comprobante__fecha__gte=fecha_desde)
    
    if fecha_hasta:
        queryset = queryset.filter(comprobante__fecha__lte=fecha_hasta)
    
    # Orden
    queryset = queryset.order_by('-comprobante__fecha', '-id')
    
    # Paginación
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    movimientos = paginator.get_page(page_number)
    
    # Estadísticas
    stats = {
        'total': MovimientoContable.objects.count(),
        'total_debitos': MovimientoContable.objects.aggregate(total=Sum('valor_debito'))['total'] or 0,
        'total_creditos': MovimientoContable.objects.aggregate(total=Sum('valor_credito'))['total'] or 0,
    }
    stats['balance'] = abs(stats['total_debitos'] - stats['total_creditos'])
    
    context = {
        'movimientos': movimientos,
        'stats': stats,
        'search': search,
        'cuenta': cuenta,
        'tipo_comprobante': tipo_comprobante,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'cuentas': PlanCuentas.objects.filter(acepta_movimientos=True),
        'tipos_comprobante': ComprobanteContable.TIPO_COMPROBANTE_CHOICES,
    }
    
    return render(request, 'contabilidad/movimientos/list.html', context)


@login_required
def movimiento_detail(request, pk):
    """Detalle de un movimiento contable."""
    movimiento = get_object_or_404(
        MovimientoContable.objects.select_related('comprobante', 'cuenta'), 
        pk=pk
    )
    
    context = {
        'movimiento': movimiento,
    }
    return render(request, 'contabilidad/movimientos/detail.html', context)


# ===================== FLUJO DE CAJA =====================

@login_required
def flujo_caja_list(request):
    """Lista de flujo de caja con filtros y gráficos."""
    queryset = FlujoCaja.objects.all()
    
    # Filtros
    search = request.GET.get('search', '')
    tipo = request.GET.get('tipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if search:
        queryset = queryset.filter(
            Q(concepto__icontains=search) |
            Q(observaciones__icontains=search)
        )
    
    if tipo:
        queryset = queryset.filter(tipo_movimiento=tipo)
    
    if fecha_desde:
        queryset = queryset.filter(fecha__gte=fecha_desde)
    
    if fecha_hasta:
        queryset = queryset.filter(fecha__lte=fecha_hasta)
    
    # Orden
    queryset = queryset.order_by('-fecha', '-id')
    
    # Paginación
    paginator = Paginator(queryset, 15)
    page_number = request.GET.get('page')
    flujos = paginator.get_page(page_number)
    
    # Estadísticas
    current_month = timezone.now().month
    current_year = timezone.now().year
    
    stats = {
        'saldo_actual': FlujoCaja.get_saldo_actual(),
        'ingresos_mes': FlujoCaja.objects.filter(
            tipo_movimiento='ingreso',
            fecha__month=current_month,
            fecha__year=current_year
        ).aggregate(total=Sum('valor'))['total'] or 0,
        'egresos_mes': FlujoCaja.objects.filter(
            tipo_movimiento='egreso',
            fecha__month=current_month,
            fecha__year=current_year
        ).aggregate(total=Sum('valor'))['total'] or 0,
    }
    stats['flujo_neto'] = stats['ingresos_mes'] - stats['egresos_mes']
    
    # Datos para gráficos
    monthly_data = FlujoCaja.get_monthly_data()
    
    context = {
        'flujos': flujos,
        'stats': stats,
        'search': search,
        'tipo': tipo,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tipos_flujo': FlujoCaja.TIPO_MOVIMIENTO_CHOICES,
        'monthly_ingresos': [month['ingresos'] for month in monthly_data],
        'monthly_egresos': [month['egresos'] for month in monthly_data],
        'projection_labels': [f"Mes {i+1}" for i in range(3)],
        'projection_data': FlujoCaja.get_projection_data(),
    }
    
    return render(request, 'contabilidad/flujo_caja/list.html', context)


@login_required
def flujo_caja_create(request):
    """Crear un nuevo movimiento de flujo de caja."""
    if request.method == 'POST':
        form = FlujoCajaForm(request.POST)
        if form.is_valid():
            flujo = form.save(commit=False)
            flujo.creado_por = request.user
            flujo.save()
            
            messages.success(request, 'Movimiento de caja creado exitosamente.')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            
            return redirect('contabilidad:flujo_caja_list')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
    else:
        form = FlujoCajaForm()
    
    context = {'form': form}
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'contabilidad/flujo_caja/form.html', context)
    
    return render(request, 'contabilidad/flujo_caja/create.html', context)


@login_required
def flujo_caja_edit(request, pk):
    """Editar un movimiento de flujo de caja."""
    flujo = get_object_or_404(FlujoCaja, pk=pk)
    
    if request.method == 'POST':
        form = FlujoCajaForm(request.POST, instance=flujo)
        if form.is_valid():
            flujo = form.save(commit=False)
            flujo.modificado_por = request.user
            flujo.save()
            
            messages.success(request, 'Movimiento de caja actualizado exitosamente.')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            
            return redirect('contabilidad:flujo_caja_list')
    else:
        form = FlujoCajaForm(instance=flujo)
    
    context = {'form': form, 'flujo': flujo}
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'contabilidad/flujo_caja/form.html', context)
    
    return render(request, 'contabilidad/flujo_caja/edit.html', context)


@login_required
def flujo_caja_detail(request, pk):
    """Detalle de un movimiento de flujo de caja."""
    flujo = get_object_or_404(FlujoCaja, pk=pk)
    context = {'flujo': flujo}
    return render(request, 'contabilidad/flujo_caja/detail.html', context)


@login_required
@require_http_methods(["DELETE"])
def flujo_caja_delete(request, pk):
    """Eliminar un movimiento de flujo de caja."""
    flujo = get_object_or_404(FlujoCaja, pk=pk)
    descripcion = flujo.descripcion
    flujo.delete()
    
    messages.success(request, f'Movimiento "{descripcion}" eliminado exitosamente.')
    
    return JsonResponse({'success': True})


# ===================== REPORTES CONTABLES =====================

REPORT_TYPES = {
    'estado_resultados': 'Estado de Resultados',
    'balance_general': 'Balance General',
    'flujo_efectivo': 'Flujo de Efectivo',
    'libro_mayor': 'Libro Mayor',
    'balance_comprobacion': 'Balance de Comprobación',
    'auxiliares': 'Auxiliares',
}

REPORT_TYPE_IDS = {
    1: 'estado_resultados',
    2: 'balance_general',
    3: 'flujo_efectivo',
    4: 'libro_mayor',
    5: 'balance_comprobacion',
    6: 'auxiliares',
}


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except Exception:
        return None


def _export_excel(headers, rows, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = filename.replace('_', ' ').title()

    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    for col_idx, _ in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[column_letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 45)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{escape_uri_path(filename + '.xlsx')}"
    wb.save(response)
    return response


def _export_pdf(title, headers, rows, filename):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Heading2'], alignment=1, spaceAfter=12)

    elements = [Paragraph(title, title_style), Spacer(1, 8)]

    data = [headers] + rows
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#E5E7EB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
    ]))
    elements.append(table)

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{escape_uri_path(filename + '.pdf')}"
    response.write(pdf)
    return response


def _get_movimientos_periodo(fecha_desde=None, fecha_hasta=None):
    movimientos = MovimientoContable.objects.select_related('comprobante', 'cuenta').order_by('-comprobante__fecha')
    if fecha_desde:
        movimientos = movimientos.filter(comprobante__fecha__gte=fecha_desde)
    if fecha_hasta:
        movimientos = movimientos.filter(comprobante__fecha__lte=fecha_hasta)
    return movimientos


def _build_balance_comprobacion(fecha_desde=None, fecha_hasta=None):
    movimientos = _get_movimientos_periodo(fecha_desde, fecha_hasta)
    cuentas = PlanCuentas.objects.filter(acepta_movimientos=True).order_by('codigo')

    rows = []
    for cuenta in cuentas:
        totales = movimientos.filter(cuenta=cuenta).aggregate(
            total_debito=Sum('valor_debito'),
            total_credito=Sum('valor_credito')
        )
        debito = totales['total_debito'] or Decimal('0.00')
        credito = totales['total_credito'] or Decimal('0.00')
        saldo = debito - credito
        rows.append([
            cuenta.codigo,
            cuenta.nombre,
            float(debito),
            float(credito),
            float(saldo),
        ])

    headers = ['Código', 'Cuenta', 'Débito', 'Crédito', 'Saldo']
    return headers, rows


def _build_movimientos_report(fecha_desde=None, fecha_hasta=None):
    movimientos = _get_movimientos_periodo(fecha_desde, fecha_hasta)
    headers = ['Fecha', 'Comprobante', 'Cuenta', 'Descripción', 'Débito', 'Crédito']
    rows = []
    for mov in movimientos:
        rows.append([
            mov.comprobante.fecha.strftime('%Y-%m-%d') if mov.comprobante and mov.comprobante.fecha else '',
            mov.comprobante.numero if mov.comprobante else '',
            f"{mov.cuenta.codigo} - {mov.cuenta.nombre}" if mov.cuenta else '',
            mov.descripcion,
            float(mov.valor_debito),
            float(mov.valor_credito),
        ])
    return headers, rows


def _build_flujo_efectivo(fecha_desde=None, fecha_hasta=None):
    flujos = FlujoCaja.objects.order_by('-fecha')
    if fecha_desde:
        flujos = flujos.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        flujos = flujos.filter(fecha__lte=fecha_hasta)

    headers = ['Fecha', 'Tipo', 'Concepto', 'Valor', 'Observaciones']
    rows = []
    for flujo in flujos:
        rows.append([
            flujo.fecha.strftime('%Y-%m-%d') if flujo.fecha else '',
            flujo.get_tipo_movimiento_display(),
            flujo.concepto,
            float(flujo.valor),
            flujo.observaciones or '',
        ])
    return headers, rows

@login_required
def reportes_list(request):
    """Dashboard de reportes contables."""
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)

    stats = {
        'reportes_generados': ComprobanteContable.objects.count(),
        'descargas_mes': ComprobanteContable.objects.filter(fecha__gte=inicio_mes).count(),
        'reportes_programados': 0,
        'usuarios_activos': User.objects.filter(is_active=True).count(),
    }

    monthly_data = FlujoCaja.get_monthly_data()

    tipos = ComprobanteContable.TIPO_COMPROBANTE_CHOICES
    category_labels = [label for _, label in tipos]
    category_data = [
        ComprobanteContable.objects.filter(tipo_comprobante=code).count()
        for code, _ in tipos
    ]

    context = {
        'stats': stats,
        'monthly_income': [month['ingresos'] for month in monthly_data],
        'monthly_expenses': [month['egresos'] for month in monthly_data],
        'category_labels': category_labels,
        'category_data': category_data,
        'last_update': {
            'estado_resultados': timezone.now().strftime('%d/%m/%Y'),
            'balance_general': timezone.now().strftime('%d/%m/%Y'),
            'flujo_efectivo': timezone.now().strftime('%d/%m/%Y'),
            'libro_mayor': timezone.now().strftime('%d/%m/%Y'),
            'balance_comprobacion': timezone.now().strftime('%d/%m/%Y'),
            'analisis_financiero': timezone.now().strftime('%d/%m/%Y'),
        }
    }
    
    return render(request, 'contabilidad/reportes/list.html', context)


@login_required
def export_comprobantes(request):
    """Exportar comprobantes a Excel."""
    comprobantes = ComprobanteContable.objects.order_by('-fecha', '-numero')

    headers = ['Número', 'Fecha', 'Tipo', 'Estado', 'Descripción', 'Total Débito', 'Total Crédito']
    rows = []
    for comp in comprobantes:
        rows.append([
            comp.numero,
            comp.fecha.strftime('%Y-%m-%d'),
            comp.get_tipo_comprobante_display(),
            comp.get_estado_display(),
            comp.descripcion,
            float(comp.total_debito),
            float(comp.total_credito),
        ])

    return _export_excel(headers, rows, 'comprobantes')


@login_required
def export_movimientos(request):
    """Exportar movimientos a Excel."""
    movimientos = MovimientoContable.objects.select_related('comprobante', 'cuenta').order_by('-comprobante__fecha')

    headers = ['Fecha', 'Comprobante', 'Cuenta', 'Descripción', 'Débito', 'Crédito']
    rows = []
    for mov in movimientos:
        rows.append([
            mov.comprobante.fecha.strftime('%Y-%m-%d') if mov.comprobante and mov.comprobante.fecha else '',
            mov.comprobante.numero if mov.comprobante else '',
            f"{mov.cuenta.codigo} - {mov.cuenta.nombre}" if mov.cuenta else '',
            mov.descripcion,
            float(mov.valor_debito),
            float(mov.valor_credito),
        ])

    return _export_excel(headers, rows, 'movimientos')


@login_required
def export_flujo_caja(request):
    """Exportar flujo de caja a Excel."""
    flujos = FlujoCaja.objects.order_by('-fecha')

    headers = ['Fecha', 'Tipo', 'Concepto', 'Valor', 'Observaciones']
    rows = []
    for flujo in flujos:
        rows.append([
            flujo.fecha.strftime('%Y-%m-%d') if flujo.fecha else '',
            flujo.get_tipo_movimiento_display(),
            flujo.concepto,
            float(flujo.valor),
            flujo.observaciones or '',
        ])

    return _export_excel(headers, rows, 'flujo_caja')


@login_required
@require_http_methods(["POST"])
def flujo_caja_preview(request):
    """Vista previa del impacto de un movimiento de flujo de caja."""
    data = json.loads(request.body)
    tipo = data.get('tipo')
    monto = Decimal(str(data.get('monto', 0)))
    tipo_normalizado = tipo.lower() if isinstance(tipo, str) else ''
    
    saldo_actual = FlujoCaja.get_saldo_actual()
    
    if tipo_normalizado == 'ingreso':
        saldo_proyectado = saldo_actual + monto
    elif tipo_normalizado == 'egreso':
        saldo_proyectado = saldo_actual - monto
    else:
        saldo_proyectado = saldo_actual
    
    return JsonResponse({
        'saldo_actual': float(saldo_actual),
        'tipo': tipo,
        'monto': float(monto),
        'saldo_proyectado': float(saldo_proyectado),
        'show': True
    })


@login_required
def comprobante_pdf(request, pk):
    """Generar PDF de un comprobante."""
    comprobante = get_object_or_404(ComprobanteContable, pk=pk)
    movimientos = MovimientoContable.objects.filter(comprobante=comprobante).select_related('cuenta')

    headers = ['Cuenta', 'Descripción', 'Débito', 'Crédito']
    rows = []
    for mov in movimientos:
        rows.append([
            f"{mov.cuenta.codigo} - {mov.cuenta.nombre}",
            mov.descripcion,
            float(mov.valor_debito),
            float(mov.valor_credito),
        ])

    title = f"Comprobante {comprobante.numero} - {comprobante.get_tipo_comprobante_display()}"
    return _export_pdf(title, headers, rows, f"comprobante_{comprobante.numero}")


@login_required
@require_http_methods(["POST"])
def generate_report(request):
    """Generar un reporte contable."""
    data = json.loads(request.body)
    report_type = data.get('type')

    fecha_desde = _parse_date(data.get('fecha_desde'))
    fecha_hasta = _parse_date(data.get('fecha_hasta'))

    if report_type == 'balance_comprobacion':
        headers, rows = _build_balance_comprobacion(fecha_desde, fecha_hasta)
    elif report_type == 'flujo_efectivo':
        headers, rows = _build_flujo_efectivo(fecha_desde, fecha_hasta)
    else:
        headers, rows = _build_movimientos_report(fecha_desde, fecha_hasta)

    title = REPORT_TYPES.get(report_type, 'Reporte Contable')
    filename = report_type or 'reporte_contable'
    return _export_pdf(title, headers, rows, filename)


@login_required
def download_report(request, pk):
    """Descargar un reporte generado."""
    report_type = REPORT_TYPE_IDS.get(pk)
    if not report_type:
        return JsonResponse({'error': 'Reporte no encontrado'}, status=404)

    headers, rows = _build_movimientos_report()
    if report_type == 'balance_comprobacion':
        headers, rows = _build_balance_comprobacion()
    elif report_type == 'flujo_efectivo':
        headers, rows = _build_flujo_efectivo()

    title = REPORT_TYPES.get(report_type, 'Reporte Contable')
    return _export_pdf(title, headers, rows, f"reporte_{report_type}")


@login_required
def view_report(request, pk):
    """Ver un reporte en el navegador."""
    report_type = REPORT_TYPE_IDS.get(pk)
    if not report_type:
        return JsonResponse({'error': 'Reporte no encontrado'}, status=404)

    headers, rows = _build_movimientos_report()
    if report_type == 'balance_comprobacion':
        headers, rows = _build_balance_comprobacion()
    elif report_type == 'flujo_efectivo':
        headers, rows = _build_flujo_efectivo()

    title = REPORT_TYPES.get(report_type, 'Reporte Contable')
    response = _export_pdf(title, headers, rows, f"reporte_{report_type}")
    response['Content-Disposition'] = f"inline; filename*=UTF-8''{escape_uri_path('reporte_' + report_type + '.pdf')}"
    return response


@login_required
def custom_report_form(request):
    """Formulario para reportes personalizados."""
    return render(request, 'contabilidad/reportes/custom_form.html')


@login_required
def schedule_report_form(request):
    """Formulario para programar reportes."""
    return render(request, 'contabilidad/reportes/schedule_form.html')


@login_required
def recent_reports(request):
    """API para obtener reportes recientes."""
    now = timezone.now()
    reports = []
    for report_id, report_type in REPORT_TYPE_IDS.items():
        reports.append({
            'id': report_id,
            'nombre': REPORT_TYPES.get(report_type, report_type),
            'usuario': request.user.get_full_name() or request.user.username,
            'fecha_generacion': now.isoformat(),
            'periodo': now.strftime('%B %Y'),
            'estado': 'GENERADO',
            'estado_display': 'Generado'
        })
    return JsonResponse(reports, safe=False)


# ===================== API ENDPOINTS =====================

@login_required
def cuentas_api(request):
    """API para obtener cuentas contables."""
    cuentas = PlanCuentas.objects.filter(acepta_movimientos=True).values(
        'id', 'codigo', 'nombre'
    )
    return JsonResponse(list(cuentas), safe=False)


@login_required
def comprobantes_api(request):
    """API para obtener comprobantes contables."""
    comprobantes = ComprobanteContable.objects.select_related('creado_por').values(
        'id', 'numero', 'fecha', 'tipo_comprobante', 'descripcion', 'estado', 'total_debito', 'total_credito',
        'creado_por__username'
    )
    
    # Convertir a lista y agregar campos calculados
    comprobantes_list = []
    for comp in comprobantes:
        comp['tipo_display'] = dict(ComprobanteContable.TIPO_COMPROBANTE_CHOICES)[comp['tipo_comprobante']]
        comp['estado_display'] = dict(ComprobanteContable.ESTADO_CHOICES)[comp['estado']]
        comp['fecha'] = comp['fecha'].isoformat() if comp['fecha'] else None
        comprobantes_list.append(comp)
    
    return JsonResponse(comprobantes_list, safe=False)


@login_required
def movimientos_api(request):
    """API para obtener movimientos contables."""
    movimientos = MovimientoContable.objects.select_related(
        'comprobante', 'cuenta'
    ).values(
        'id', 'descripcion', 'valor_debito', 'valor_credito',
        'comprobante_id', 'comprobante__numero', 'comprobante__fecha', 'comprobante__tipo_comprobante',
        'cuenta_id', 'cuenta__codigo', 'cuenta__nombre'
    )
    
    # Convertir a lista y reorganizar datos
    movimientos_list = []
    for mov in movimientos:
        movimientos_list.append({
            'id': mov['id'],
            'descripcion': mov['descripcion'],
            'debito': float(mov['valor_debito']),
            'credito': float(mov['valor_credito']),
            'comprobante': {
                'id': mov['comprobante_id'],
                'numero': mov['comprobante__numero'],
                'fecha': mov['comprobante__fecha'].isoformat() if mov['comprobante__fecha'] else None,
                'tipo': mov['comprobante__tipo_comprobante'],
                'tipo_display': dict(ComprobanteContable.TIPO_COMPROBANTE_CHOICES)[mov['comprobante__tipo_comprobante']]
            },
            'cuenta': {
                'id': mov['cuenta_id'],
                'codigo': mov['cuenta__codigo'],
                'nombre': mov['cuenta__nombre']
            }
        })
    
    return JsonResponse(movimientos_list, safe=False)


@login_required
def flujo_caja_api(request):
    """API para obtener flujo de caja."""
    flujos = FlujoCaja.objects.values(
        'id', 'fecha', 'tipo_movimiento', 'concepto', 'valor', 'observaciones'
    )
    
    # Convertir a lista y agregar campos calculados
    flujos_list = []
    for flujo in flujos:
        flujo['tipo_display'] = dict(FlujoCaja.TIPO_MOVIMIENTO_CHOICES)[flujo['tipo_movimiento']]
        flujo['fecha'] = flujo['fecha'].isoformat() if flujo['fecha'] else None
        flujo['valor'] = float(flujo['valor'])
        flujos_list.append(flujo)
    
    return JsonResponse(flujos_list, safe=False)


# ===================== ESTADÍSTICAS Y MÉTRICAS =====================

@login_required
def comprobantes_stats(request):
    """Estadísticas de comprobantes."""
    stats = {
        'total': ComprobanteContable.objects.count(),
        'confirmados': ComprobanteContable.objects.filter(estado='contabilizado').count(),
        'pendientes': ComprobanteContable.objects.filter(estado='borrador').count(),
        'valor_total': float(ComprobanteContable.objects.filter(
            estado='contabilizado'
        ).aggregate(total=Sum('total_debito'))['total'] or 0)
    }
    return JsonResponse(stats)


@login_required
def movimientos_stats(request):
    """Estadísticas de movimientos."""
    stats = {
        'total': MovimientoContable.objects.count(),
        'total_debitos': float(MovimientoContable.objects.aggregate(
            total=Sum('valor_debito'))['total'] or 0),
        'total_creditos': float(MovimientoContable.objects.aggregate(
            total=Sum('valor_credito'))['total'] or 0)
    }
    return JsonResponse(stats)


@login_required
def flujo_caja_stats(request):
    """Estadísticas de flujo de caja."""
    current_month = timezone.now().month
    current_year = timezone.now().year
    
    stats = {
        'saldo_actual': float(FlujoCaja.get_saldo_actual()),
        'ingresos_mes': float(FlujoCaja.objects.filter(
            tipo_movimiento='ingreso',
            fecha__month=current_month,
            fecha__year=current_year
        ).aggregate(total=Sum('valor'))['total'] or 0),
        'egresos_mes': float(FlujoCaja.objects.filter(
            tipo_movimiento='egreso',
            fecha__month=current_month,
            fecha__year=current_year
        ).aggregate(total=Sum('valor'))['total'] or 0),
    }
    stats['flujo_neto'] = stats['ingresos_mes'] - stats['egresos_mes']
    
    return JsonResponse(stats)


@login_required
def saldo_actual(request):
    """Obtener saldo actual de caja."""
    return JsonResponse({
        'saldo_actual': float(FlujoCaja.get_saldo_actual())
    })


@login_required
def reportes_stats(request):
    """Estadísticas de reportes."""
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    stats = {
        'reportes_generados': ComprobanteContable.objects.count(),
        'descargas_mes': ComprobanteContable.objects.filter(fecha__gte=inicio_mes).count(),
        'reportes_programados': 0,
        'usuarios_activos': User.objects.filter(is_active=True).count(),
    }
    return JsonResponse(stats)


# ===================== UTILIDADES =====================
