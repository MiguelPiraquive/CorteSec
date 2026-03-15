from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, Sum, Count, F
from django.utils import timezone
from datetime import datetime, timedelta
from io import BytesIO

from django.http import HttpResponse
from django.utils.encoding import escape_uri_path

from .models import PlanCuentas, ComprobanteContable, MovimientoContable, FlujoCaja, CentroCosto, BalanceComprobacion
from .serializers import (
    PlanCuentasSerializer, ComprobanteContableSerializer,
    MovimientoContableSerializer, FlujoCajaSerializer,
    CentroCostoSerializer, BalanceComprobacionSerializer
)
from core.mixins import MultiTenantViewSetMixin
from .policies import ContabilidadAccessPolicy


from core.utils import get_active_project_for_request as _get_active_project_for_request


class PlanCuentasViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet para gestionar Plan de Cuentas"""
    queryset = PlanCuentas.objects.select_related('cuenta_padre').all()
    serializer_class = PlanCuentasSerializer
    permission_classes = [ContabilidadAccessPolicy]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['tipo_cuenta', 'naturaleza', 'activa', 'acepta_movimientos', 'nivel']
    search_fields = ['codigo', 'nombre', 'descripcion']
    ordering_fields = ['codigo', 'nombre', 'nivel', 'fecha_creacion']
    ordering = ['codigo']

    def get_queryset(self):
        queryset = super().get_queryset()
        # Filtros adicionales por query params
        activa = self.request.query_params.get('activa')
        if activa is not None:
            queryset = queryset.filter(activa=activa.lower() == 'true')
        
        nivel = self.request.query_params.get('nivel')
        if nivel:
            queryset = queryset.filter(nivel=nivel)
            
        return queryset

    @action(detail=False, methods=['get'])
    def jerarquia(self, request):
        """Obtener cuentas organizadas en jerarquía"""
        cuentas_raiz = self.get_queryset().filter(cuenta_padre__isnull=True)
        def construir_jerarquia(cuenta):
            serializer = self.get_serializer(cuenta)
            data = serializer.data
            subcuentas = cuenta.subcuentas.filter(activa=True)
            if subcuentas.exists():
                data['subcuentas'] = [construir_jerarquia(sub) for sub in subcuentas]
            return data
        
        jerarquia = [construir_jerarquia(cuenta) for cuenta in cuentas_raiz]
        return Response(jerarquia)

    @action(detail=True, methods=['get'])
    def saldo(self, request, pk=None):
        """Obtener saldo actual de una cuenta"""
        cuenta = self.get_object()
        saldo = cuenta.calcular_saldo()
        return Response({
            'cuenta_id': cuenta.id,
            'codigo': cuenta.codigo,
            'nombre': cuenta.nombre,
            'saldo_actual': saldo,
            'naturaleza': cuenta.naturaleza
        })

    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Obtener estadísticas del plan de cuentas"""
        queryset = self.get_queryset()
        stats = {
            'total_cuentas': queryset.count(),
            'cuentas_activas': queryset.filter(activa=True).count(),
            'cuentas_por_tipo': {
                tipo[0]: queryset.filter(tipo_cuenta=tipo[0]).count()
                for tipo in PlanCuentas.TIPOS_CUENTA
            },
            'cuentas_por_naturaleza': {
                nat[0]: queryset.filter(naturaleza=nat[0]).count()
                for nat in PlanCuentas.NATURALEZA_CHOICES
            },
            'cuentas_con_movimientos': queryset.filter(acepta_movimientos=True).count()
        }
        return Response(stats)


class ComprobanteContableViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet para gestionar Comprobantes Contables"""
    queryset = ComprobanteContable.objects.select_related(
        'creado_por', 'contabilizado_por', 'nomina_relacionada', 'prestamo_relacionado'
    ).all()
    serializer_class = ComprobanteContableSerializer
    permission_classes = [ContabilidadAccessPolicy]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['tipo_comprobante', 'estado', 'fecha']
    search_fields = ['numero', 'descripcion']
    ordering_fields = ['numero', 'fecha', 'fecha_creacion']
    ordering = ['-fecha', '-fecha_creacion']

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filtrar por proyecto activo
        project = _get_active_project_for_request(self.request)
        if project:
            queryset = queryset.filter(proyecto=project)
        
        # Filtrar por rango de fechas
        fecha_desde = self.request.query_params.get('fecha_desde')
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        
        if fecha_desde:
            queryset = queryset.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(fecha__lte=fecha_hasta)
            
        return queryset

    def perform_create(self, serializer):
        serializer.save(creado_por=self.request.user)

    @action(detail=True, methods=['post'])
    def contabilizar(self, request, pk=None):
        """Contabilizar un comprobante"""
        comprobante = self.get_object()
        
        if comprobante.estado == 'contabilizado':
            return Response(
                {'error': 'El comprobante ya está contabilizado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verificar que esté cuadrado (es @property, no método)
        if not comprobante.esta_cuadrado:
            return Response(
                {'error': 'El comprobante no está cuadrado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        comprobante.estado = 'contabilizado'
        comprobante.contabilizado_por = request.user
        comprobante.fecha_contabilizacion = timezone.now()
        comprobante.save()
        
        return Response({'message': 'Comprobante contabilizado correctamente'})

    @action(detail=True, methods=['post'])
    def anular(self, request, pk=None):
        """Anular un comprobante"""
        comprobante = self.get_object()
        
        if comprobante.estado == 'anulado':
            return Response(
                {'error': 'El comprobante ya está anulado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        comprobante.estado = 'anulado'
        comprobante.save()
        
        return Response({'message': 'Comprobante anulado correctamente'})

    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Obtener estadísticas de comprobantes"""
        queryset = self.get_queryset()
        
        # Filtros de fecha para estadísticas
        hoy = timezone.now().date()
        inicio_mes = hoy.replace(day=1)
        
        stats = {
            'total_comprobantes': queryset.count(),
            'comprobantes_mes': queryset.filter(fecha__gte=inicio_mes).count(),
            'comprobantes_por_estado': {
                estado[0]: queryset.filter(estado=estado[0]).count()
                for estado in ComprobanteContable.ESTADOS
            },
            'comprobantes_por_tipo': {
                tipo[0]: queryset.filter(tipo_comprobante=tipo[0]).count()
                for tipo in ComprobanteContable.TIPOS_COMPROBANTE
            },
            'valor_total_mes': queryset.filter(
                fecha__gte=inicio_mes,
                estado='contabilizado'
            ).aggregate(total=Sum('total_debito'))['total'] or 0
        }
        return Response(stats)


class MovimientoContableViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet para gestionar Movimientos Contables"""
    queryset = MovimientoContable.objects.select_related('comprobante', 'cuenta', 'centro_costo')
    serializer_class = MovimientoContableSerializer
    permission_classes = [ContabilidadAccessPolicy]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['comprobante', 'cuenta', 'comprobante__estado', 'centro_costo']
    search_fields = ['descripcion', 'tercero', 'cuenta__nombre', 'comprobante__numero']
    ordering_fields = ['comprobante__fecha', 'valor_debito', 'valor_credito']
    ordering = ['-comprobante__fecha']

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filtrar por proyecto activo
        project = _get_active_project_for_request(self.request)
        if project:
            queryset = queryset.filter(comprobante__proyecto=project)
        
        # Filtrar por cuenta específica
        cuenta_id = self.request.query_params.get('cuenta_id')
        if cuenta_id:
            queryset = queryset.filter(cuenta_id=cuenta_id)

        centro_costo_id = self.request.query_params.get('centro_costo_id')
        if centro_costo_id:
            queryset = queryset.filter(centro_costo_id=centro_costo_id)
            
        # Filtrar por rango de fechas
        fecha_desde = self.request.query_params.get('fecha_desde')
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        
        if fecha_desde:
            queryset = queryset.filter(comprobante__fecha__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(comprobante__fecha__lte=fecha_hasta)
            
        return queryset

    def _apply_date_filters(self, queryset, request):
        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')

        if fecha_desde:
            queryset = queryset.filter(comprobante__fecha__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(comprobante__fecha__lte=fecha_hasta)
        return queryset

    def _export_excel(self, headers, rows, filename, totals=None):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = filename.replace('_', ' ').title()

        header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        total_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
        total_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for row in rows:
            ws.append(row)

        if totals:
            ws.append(totals)
            total_row = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=total_row, column=col_idx)
                cell.fill = total_fill
                cell.font = total_font
                cell.border = thin_border

        currency_cols = [idx + 1 for idx, header in enumerate(headers) if str(header).lower() in {'debito', 'crédito', 'credito', 'saldo'}]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if cell.column in currency_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

        for col_idx, header in enumerate(headers, start=1):
            max_length = max(len(str(header)), 12)
            for cell in ws[get_column_letter(col_idx)]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 45)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{escape_uri_path(filename + '.xlsx')}"
        return response

    def _export_pdf(self, title, headers, rows, filename, organization=None, totals=None):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.units import inch

        from configuracion.models import ConfiguracionGeneral

        output = BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=letter,
            leftMargin=36,
            rightMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()
        elements = []

        logo = None
        company_name = ''
        company_nit = ''
        if organization:
            configuracion = ConfiguracionGeneral.objects.filter(organization=organization).first()
            if configuracion:
                company_name = configuracion.nombre_empresa or ''
                company_nit = configuracion.nit or ''
                if configuracion.logo:
                    try:
                        logo = Image(configuracion.logo.path, width=1.8 * inch, height=0.9 * inch)
                    except Exception:
                        logo = None

        header_style = ParagraphStyle(
            'HeaderTitle',
            parent=styles['Title'],
            fontSize=16,
            leading=20,
            textColor=colors.HexColor('#111827')
        )

        subtitle_style = ParagraphStyle(
            'HeaderSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            leading=12,
            textColor=colors.HexColor('#374151')
        )

        header_cells = []
        header_cells.append(logo if logo else '')

        header_text = Paragraph(f"{title}", header_style)
        subtitle_text = Paragraph(
            f"{company_name}<br/>NIT: {company_nit}" if company_name or company_nit else "",
            subtitle_style
        )

        header_table = Table(
            [[header_cells[0], header_text], ['', subtitle_text]],
            colWidths=[2.0 * inch, 4.6 * inch]
        )
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 16))

        table_data = [headers]
        table_data.extend([[str(value) for value in row] for row in rows])
        if totals:
            table_data.append([str(value) for value in totals])

        col_count = len(headers)
        col_widths = [doc.width / col_count for _ in range(col_count)]
        table = Table(table_data, repeatRows=1, hAlign='LEFT', colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        if totals:
            last_row = len(table_data) - 1
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, last_row), (-1, last_row), colors.HexColor('#E5E7EB')),
                ('FONTNAME', (0, last_row), (-1, last_row), 'Helvetica-Bold'),
            ]))
        elements.append(table)

        elements.append(Spacer(1, 24))

        firma_label = 'Firma autorizada'
        firma_nombre = company_name or 'Representante legal'
        firma_table = Table([
            ['', ''],
            [firma_label, firma_nombre],
        ], colWidths=[2.8 * inch, 3.6 * inch])
        firma_table.setStyle(TableStyle([
            ('LINEABOVE', (0, 1), (0, 1), 1, colors.black),
            ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 1), (1, 1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (1, 1), 9),
            ('TOPPADDING', (0, 1), (1, 1), 6),
        ]))
        elements.append(firma_table)

        doc.build(elements)
        output.seek(0)

        response = HttpResponse(output.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{escape_uri_path(filename + '.pdf')}"
        return response

    @action(detail=False, methods=['get'])
    def balance_prueba(self, request):
        """Balance de prueba con filtro por período."""
        queryset = self._apply_date_filters(self.get_queryset(), request)

        data = queryset.values(
            'cuenta_id',
            'cuenta__codigo',
            'cuenta__nombre',
            'cuenta__naturaleza'
        ).annotate(
            total_debito=Sum('valor_debito'),
            total_credito=Sum('valor_credito')
        ).order_by('cuenta__codigo')

        results = []
        for row in data:
            total_debito = row['total_debito'] or 0
            total_credito = row['total_credito'] or 0
            naturaleza = row['cuenta__naturaleza']
            if naturaleza == 'debito':
                saldo = total_debito - total_credito
            else:
                saldo = total_credito - total_debito
            results.append({
                'cuenta_id': row['cuenta_id'],
                'codigo': row['cuenta__codigo'],
                'nombre': row['cuenta__nombre'],
                'naturaleza': naturaleza,
                'total_debito': total_debito,
                'total_credito': total_credito,
                'saldo': saldo,
            })

        return Response(results)

    @action(detail=False, methods=['get'])
    def libro_mayor(self, request):
        """Libro mayor por cuenta y período."""
        cuenta_id = request.query_params.get('cuenta_id')
        if not cuenta_id:
            return Response({'error': 'Se requiere cuenta_id'}, status=status.HTTP_400_BAD_REQUEST)

        queryset = self._apply_date_filters(
            self.get_queryset().filter(cuenta_id=cuenta_id),
            request
        )

        movimientos = queryset.order_by('comprobante__fecha', 'id')
        serializer = self.get_serializer(movimientos, many=True)
        totales = movimientos.aggregate(
            total_debito=Sum('valor_debito'),
            total_credito=Sum('valor_credito')
        )

        return Response({
            'movimientos': serializer.data,
            'totales': totales,
            'saldo': (totales['total_debito'] or 0) - (totales['total_credito'] or 0)
        })

    @action(detail=False, methods=['get'])
    def auxiliares(self, request):
        """Auxiliares por cuenta/tercero con filtro por período."""
        queryset = self.get_queryset()
        cuenta_id = request.query_params.get('cuenta_id')

        if cuenta_id:
            queryset = queryset.filter(cuenta_id=cuenta_id)
        queryset = self._apply_date_filters(queryset, request)

        data = queryset.values(
            'cuenta__codigo',
            'cuenta__nombre',
            'tercero'
        ).annotate(
            total_debito=Sum('valor_debito'),
            total_credito=Sum('valor_credito')
        ).order_by('cuenta__codigo', 'tercero')

        return Response(list(data))

    @action(detail=False, methods=['get'])
    def auditoria_puc(self, request):
        """Auditoría rápida de configuración y cuentas PUC."""
        from configuracion.models import ConfiguracionGeneral

        org = request.user.organization
        configuracion = ConfiguracionGeneral.objects.filter(organization=org).first()
        faltan_config = []
        if not configuracion:
            faltan_config = [
                'cuenta_efectivo_defecto',
                'cuenta_nomina_defecto',
                'cuenta_prestamos_defecto',
                'cuenta_intereses_prestamo_defecto',
                'cuenta_mora_prestamo_defecto',
                'cuenta_otras_deducciones_defecto'
            ]
        else:
            for field in [
                'cuenta_efectivo_defecto',
                'cuenta_nomina_defecto',
                'cuenta_prestamos_defecto',
                'cuenta_intereses_prestamo_defecto',
                'cuenta_mora_prestamo_defecto',
                'cuenta_otras_deducciones_defecto'
            ]:
                if not getattr(configuracion, field, None):
                    faltan_config.append(field)

        required_codes = ['1105', '5105', '1365', '421005', '4175', '237005', '238030', '2370']
        faltan_cuentas = []
        for codigo in required_codes:
            cuenta = PlanCuentas.objects.filter(codigo=codigo, organization=org).first()
            if not cuenta:
                cuenta = PlanCuentas.objects.filter(codigo=codigo, organization__isnull=True).first()
            if not cuenta:
                faltan_cuentas.append(codigo)

        return Response({
            'faltan_config': faltan_config,
            'faltan_cuentas': faltan_cuentas
        })

    @action(detail=False, methods=['get'])
    def exportar_balance_prueba(self, request):
        """Exporta balance de prueba en Excel o PDF."""
        MAX_EXPORT_ROWS = 5000
        formato = request.query_params.get('formato', 'excel').lower()
        data = self.balance_prueba(request).data[:MAX_EXPORT_ROWS]
        headers = ['Código', 'Cuenta', 'Naturaleza', 'Débito', 'Crédito', 'Saldo']
        rows = [
            [
                item['codigo'],
                item['nombre'],
                item['naturaleza'],
                float(item['total_debito'] or 0),
                float(item['total_credito'] or 0),
                float(item['saldo'] or 0)
            ]
            for item in data
        ]
        total_debito = sum((item['total_debito'] or 0) for item in data)
        total_credito = sum((item['total_credito'] or 0) for item in data)
        total_saldo = sum((item['saldo'] or 0) for item in data)
        totals = ['TOTAL', '', '', float(total_debito), float(total_credito), float(total_saldo)]

        if formato in ['excel', 'xlsx']:
            return self._export_excel(headers, rows, 'balance_prueba', totals=totals)
        if formato == 'pdf':
            return self._export_pdf('Balance de Prueba', headers, rows, 'balance_prueba', organization=request.user.organization, totals=totals)
        return Response({'error': 'Formato inválido'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def exportar_libro_mayor(self, request):
        """Exporta libro mayor en Excel o PDF."""
        MAX_EXPORT_ROWS = 5000
        formato = request.query_params.get('formato', 'excel').lower()
        response = self.libro_mayor(request)
        if response.status_code != status.HTTP_200_OK:
            return response

        data = response.data.get('movimientos', [])[:MAX_EXPORT_ROWS]
        headers = ['Fecha', 'Comprobante', 'Descripción', 'Débito', 'Crédito', 'Tercero']
        rows = [
            [
                item.get('comprobante_fecha'),
                item.get('comprobante_numero'),
                item.get('descripcion'),
                float(item.get('valor_debito') or 0),
                float(item.get('valor_credito') or 0),
                item.get('tercero') or ''
            ]
            for item in data
        ]
        total_debito = sum((item.get('valor_debito') or 0) for item in data)
        total_credito = sum((item.get('valor_credito') or 0) for item in data)
        totals = ['TOTAL', '', '', float(total_debito), float(total_credito), '']

        if formato in ['excel', 'xlsx']:
            return self._export_excel(headers, rows, 'libro_mayor', totals=totals)
        if formato == 'pdf':
            return self._export_pdf('Libro Mayor', headers, rows, 'libro_mayor', organization=request.user.organization, totals=totals)
        return Response({'error': 'Formato inválido'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def exportar_auxiliares(self, request):
        """Exporta auxiliares en Excel o PDF."""
        MAX_EXPORT_ROWS = 5000
        formato = request.query_params.get('formato', 'excel').lower()
        data = self.auxiliares(request).data[:MAX_EXPORT_ROWS]
        headers = ['Cuenta', 'Tercero', 'Débito', 'Crédito']
        rows = [
            [
                f"{item['cuenta__codigo']} - {item['cuenta__nombre']}",
                item.get('tercero') or 'Sin tercero',
                float(item.get('total_debito') or 0),
                float(item.get('total_credito') or 0),
            ]
            for item in data
        ]
        total_debito = sum((item.get('total_debito') or 0) for item in data)
        total_credito = sum((item.get('total_credito') or 0) for item in data)
        totals = ['TOTAL', '', float(total_debito), float(total_credito)]

        if formato in ['excel', 'xlsx']:
            return self._export_excel(headers, rows, 'auxiliares', totals=totals)
        if formato == 'pdf':
            return self._export_pdf('Auxiliares', headers, rows, 'auxiliares', organization=request.user.organization, totals=totals)
        return Response({'error': 'Formato inválido'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def por_cuenta(self, request):
        """Obtener movimientos agrupados por cuenta"""
        cuenta_id = request.query_params.get('cuenta_id')
        if not cuenta_id:
            return Response(
                {'error': 'Se requiere el parámetro cuenta_id'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        movimientos = self.get_queryset().filter(cuenta_id=cuenta_id)
        serializer = self.get_serializer(movimientos, many=True)
        
        # Calcular totales
        totales = movimientos.aggregate(
            total_debitos=Sum('valor_debito'),
            total_creditos=Sum('valor_credito')
        )
        
        return Response({
            'movimientos': serializer.data,
            'totales': totales,
            'saldo': (totales['total_debitos'] or 0) - (totales['total_creditos'] or 0)
        })

    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Obtener estadísticas de movimientos"""
        queryset = self.get_queryset()
        
        # Filtros de fecha
        hoy = timezone.now().date()
        inicio_mes = hoy.replace(day=1)
        
        stats = {
            'total_movimientos': queryset.count(),
            'movimientos_mes': queryset.filter(comprobante__fecha__gte=inicio_mes).count(),
            'total_debitos': queryset.aggregate(total=Sum('valor_debito'))['total'] or 0,
            'total_creditos': queryset.aggregate(total=Sum('valor_credito'))['total'] or 0,
            'movimientos_por_cuenta': queryset.values('cuenta__nombre').annotate(
                total=Count('id')
            ).order_by('-total')[:10]
        }
        return Response(stats)


class FlujoCajaViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet para gestionar Flujo de Caja"""
    queryset = FlujoCaja.objects.select_related('comprobante').all()
    serializer_class = FlujoCajaSerializer
    permission_classes = [ContabilidadAccessPolicy]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['tipo_movimiento', 'fecha']
    search_fields = ['concepto', 'observaciones']
    ordering_fields = ['fecha', 'valor']
    ordering = ['-fecha']

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filtrar por proyecto activo
        project = _get_active_project_for_request(self.request)
        if project:
            queryset = queryset.filter(proyecto=project)
        
        # Filtrar por rango de fechas
        fecha_desde = self.request.query_params.get('fecha_desde')
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        
        if fecha_desde:
            queryset = queryset.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(fecha__lte=fecha_hasta)
            
        return queryset

    def perform_create(self, serializer):
        serializer.save(organization=self.request.user.organization)

    @action(detail=False, methods=['get'])
    def resumen(self, request):
        """Obtener resumen del flujo de caja"""
        queryset = self.get_queryset()
        
        ingresos = queryset.filter(tipo_movimiento='ingreso').aggregate(
            total=Sum('valor')
        )['total'] or 0
        
        egresos = queryset.filter(tipo_movimiento='egreso').aggregate(
            total=Sum('valor')
        )['total'] or 0
        
        return Response({
            'total_ingresos': ingresos,
            'total_egresos': egresos,
            'flujo_neto': ingresos - egresos,
            'movimientos_count': queryset.count()
        })

    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Obtener estadísticas del flujo de caja"""
        queryset = self.get_queryset()
        
        # Estadísticas por período
        hoy = timezone.now().date()
        inicio_mes = hoy.replace(day=1)
        
        stats = {
            'total_movimientos': queryset.count(),
            'movimientos_mes': queryset.filter(fecha__gte=inicio_mes).count(),
            'ingresos_por_categoria': queryset.filter(
                tipo_movimiento='ingreso'
            ).values('concepto').annotate(
                total=Sum('valor')
            ).order_by('-total'),
            'egresos_por_categoria': queryset.filter(
                tipo_movimiento='egreso'
            ).values('concepto').annotate(
                total=Sum('valor')
            ).order_by('-total')
        }
        return Response(stats)


class CentroCostoViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet para gestionar Centros de Costo"""
    queryset = CentroCosto.objects.select_related('centro_padre', 'responsable').all()
    serializer_class = CentroCostoSerializer
    permission_classes = [ContabilidadAccessPolicy]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['activo', 'nivel']
    search_fields = ['codigo', 'nombre', 'descripcion']
    ordering_fields = ['codigo', 'nombre', 'nivel', 'fecha_creacion']
    ordering = ['codigo']


class BalanceComprobacionViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet para gestionar Balances de Comprobación"""
    queryset = BalanceComprobacion.objects.select_related('cuenta')
    serializer_class = BalanceComprobacionSerializer
    permission_classes = [ContabilidadAccessPolicy]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['periodo', 'cuenta']
    search_fields = ['periodo', 'cuenta__codigo', 'cuenta__nombre']
    ordering_fields = ['periodo', 'cuenta__codigo']
    ordering = ['-periodo', 'cuenta__codigo']
