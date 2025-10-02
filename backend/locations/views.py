from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.db.models import Q, Prefetch, Count
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.shortcuts import redirect
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import Departamento, Municipio
from .forms import DepartamentoForm, MunicipioForm
from .serializers import (
    DepartamentoSerializer, DepartamentoSimpleSerializer,
    MunicipioSerializer, MunicipioSimpleSerializer, MunicipioConDepartamentoSerializer,
    BusquedaUbicacionSerializer, UbicacionHierarcaSerializer
)


# Vistas tradicionales de Django (para templates)
class DepartamentoListView(ListView):
    """Vista de lista de departamentos"""
    model = Departamento
    template_name = "locations/departamento_lista.html"
    context_object_name = "departamentos"
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Departamento.objects.all()
        
        # Filtros de búsqueda
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(nombre__icontains=search) |
                Q(codigo__icontains=search)
            )
        
        return queryset.order_by('nombre')


class DepartamentoCreateView(CreateView):
    """Vista para crear departamentos"""
    model = Departamento
    form_class = DepartamentoForm
    template_name = "locations/departamento_formulario.html"
    success_url = reverse_lazy("locations:departamento_lista")


class DepartamentoUpdateView(UpdateView):
    """Vista para editar departamentos"""
    model = Departamento
    form_class = DepartamentoForm
    template_name = "locations/departamento_formulario.html"
    success_url = reverse_lazy("locations:departamento_lista")


class DepartamentoDeleteView(DeleteView):
    """Vista para eliminar departamentos"""
    model = Departamento
    template_name = "locations/departamento_confirmar_eliminar.html"
    success_url = reverse_lazy("locations:departamento_lista")


class DepartamentoDetailView(DetailView):
    """Vista de detalle de departamento"""
    model = Departamento
    template_name = "locations/departamento_detalle.html"
    context_object_name = "departamento"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['municipios'] = self.object.municipios.all().order_by('nombre')
        return context


# Vistas para municipios
class MunicipioListView(ListView):
    """Vista de lista de municipios"""
    model = Municipio
    template_name = "locations/municipio_lista.html"
    context_object_name = "municipios"
    paginate_by = 50
    
    def get_queryset(self):
        queryset = Municipio.objects.select_related('departamento')
        
        # Filtros
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(nombre__icontains=search) |
                Q(departamento__nombre__icontains=search) |
                Q(codigo__icontains=search)
            )
        
        departamento_id = self.request.GET.get('departamento')
        if departamento_id:
            queryset = queryset.filter(departamento_id=departamento_id)
        
        return queryset.order_by('departamento__nombre', 'nombre')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['departamentos'] = Departamento.objects.all().order_by('nombre')
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_departamento'] = self.request.GET.get('departamento', '')
        return context


class MunicipioCreateView(CreateView):
    """Vista para crear municipios"""
    model = Municipio
    form_class = MunicipioForm
    template_name = "locations/municipio_formulario.html"
    success_url = reverse_lazy("locations:municipio_lista")


class MunicipioUpdateView(UpdateView):
    """Vista para editar municipios"""
    model = Municipio
    form_class = MunicipioForm
    template_name = "locations/municipio_formulario.html"
    success_url = reverse_lazy("locations:municipio_lista")


class MunicipioDeleteView(DeleteView):
    """Vista para eliminar municipios"""
    model = Municipio
    template_name = "locations/municipio_confirmar_eliminar.html"
    success_url = reverse_lazy("locations:municipio_lista")


class MunicipioDetailView(DetailView):
    """Vista de detalle de municipio"""
    model = Municipio
    template_name = "locations/municipio_detalle.html"
    context_object_name = "municipio"


# ViewSets para API REST
class DepartamentoViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de departamentos
    
    Proporciona operaciones CRUD y funcionalidades adicionales:
    - Búsqueda por nombre o código
    - Lista de municipios por departamento
    - Estadísticas
    """
    queryset = Departamento.objects.all()
    serializer_class = DepartamentoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nombre', 'codigo']
    ordering_fields = ['nombre', 'created_at']
    ordering = ['nombre']
    
    def get_queryset(self):
        """Filtrar por organización si es necesario"""
        if hasattr(self.request.user, 'organization'):
            return Departamento.objects.filter(organizacion=self.request.user.organizacion)
        return Departamento.objects.all()
    
    def get_serializer_class(self):
        """Usar diferentes serializers según la acción"""
        if self.action == 'list':
            return DepartamentoSimpleSerializer
        return DepartamentoSerializer
    
    def perform_create(self, serializer):
        """Asignar organización al crear"""
        if hasattr(self.request.user, 'organization'):
            serializer.save(organizacion=self.request.user.organizacion)
    
    @action(detail=True, methods=['get'])
    def municipios(self, request, pk=None):
        """Obtener municipios de un departamento"""
        departamento = self.get_object()
        municipios = departamento.municipios.all().order_by('nombre')
        serializer = MunicipioSimpleSerializer(municipios, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def importar_excel(self, request):
        """Importar departamentos desde archivo Excel"""
        if 'archivo' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó archivo'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        archivo = request.FILES['archivo']
        
        # Validar formato
        if not archivo.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'El archivo debe ser Excel (.xlsx o .xls)'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Leer archivo Excel
            df = pd.read_excel(archivo)
            
            # Validar columnas requeridas
            columnas_requeridas = ['nombre']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            
            if columnas_faltantes:
                return Response(
                    {'error': f'Faltan columnas: {", ".join(columnas_faltantes)}'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Procesar datos
            departamentos_creados = 0
            errores = []
            
            for index, row in df.iterrows():
                try:
                    # Validar datos obligatorios
                    if pd.isna(row['nombre']) or not str(row['nombre']).strip():
                        errores.append(f"Fila {index + 2}: Nombre es obligatorio")
                        continue
                    
                    # Crear departamento
                    departamento_data = {
                        'nombre': str(row['nombre']).strip(),
                        'codigo': str(row['codigo']).strip() if 'codigo' in row and not pd.isna(row['codigo']) else '',
                        'capital': str(row['capital']).strip() if 'capital' in row and not pd.isna(row['capital']) else '',
                        'region': str(row['region']).strip() if 'region' in row and not pd.isna(row['region']) else '',
                    }
                    
                    # Asignar organización
                    if hasattr(request.user, 'organization'):
                        departamento_data['organization'] = request.user.organizacion
                    
                    # Verificar si ya existe
                    existing = Departamento.objects.filter(
                        nombre=departamento_data['nombre']
                    )
                    if hasattr(request.user, 'organization'):
                        existing = existing.filter(organizacion=request.user.organizacion)
                    
                    if existing.exists():
                        errores.append(f"Fila {index + 2}: Departamento '{departamento_data['nombre']}' ya existe")
                        continue
                    
                    # Crear departamento
                    Departamento.objects.create(**departamento_data)
                    departamentos_creados += 1
                    
                except Exception as e:
                    errores.append(f"Fila {index + 2}: Error - {str(e)}")
            
            return Response({
                'message': f'Importación completada. {departamentos_creados} departamentos creados.',
                'departamentos_creados': departamentos_creados,
                'errores': errores
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error al procesar archivo: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        """Exportar departamentos a Excel"""
        try:
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Departamentos"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Headers
            headers = ['ID', 'Nombre', 'Código', 'Capital', 'Región', 'Municipios', 'Fecha Creación']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Datos
            queryset = self.get_queryset().annotate(municipios_count=Count('municipios'))
            
            for row, departamento in enumerate(queryset, 2):
                ws.cell(row=row, column=1, value=str(departamento.id))
                ws.cell(row=row, column=2, value=departamento.nombre)
                ws.cell(row=row, column=3, value=departamento.codigo)
                ws.cell(row=row, column=4, value=departamento.capital)
                ws.cell(row=row, column=5, value=departamento.region)
                ws.cell(row=row, column=6, value=departamento.municipios_count)
                ws.cell(row=row, column=7, value=departamento.created_at.strftime('%Y-%m-%d'))
            
            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            
            # Crear respuesta HTTP
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="departamentos.xlsx"'
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Error al exportar: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def template_importacion(self, request):
        """Descargar template para importación de departamentos"""
        try:
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Template Departamentos"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # Headers
            headers = ['nombre', 'codigo', 'capital', 'region']
            help_text = ['Nombre del departamento (OBLIGATORIO)', 'Código DANE', 'Ciudad capital', 'Región geográfica']
            
            for col, (header, help) in enumerate(zip(headers, help_text), 1):
                # Header
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                
                # Help text
                help_cell = ws.cell(row=2, column=col, value=help)
                help_cell.fill = example_fill
            
            # Ejemplos
            ejemplos = [
                ['Cundinamarca', '25', 'Bogotá', 'Andina'],
                ['Antioquia', '05', 'Medellín', 'Andina'],
                ['Valle del Cauca', '76', 'Cali', 'Pacífica'],
            ]
            
            for row, ejemplo in enumerate(ejemplos, 3):
                for col, valor in enumerate(ejemplo, 1):
                    ws.cell(row=row, column=col, value=valor)
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            
            # Crear respuesta HTTP
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_departamentos.xlsx"'
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Error al generar template: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class MunicipioViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de municipios
    
    Proporciona operaciones CRUD y funcionalidades adicionales:
    - Búsqueda por nombre, departamento o código
    - Filtrado por departamento
    - Información jerárquica
    """
    queryset = Municipio.objects.select_related('departamento')
    serializer_class = MunicipioSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['departamento']
    search_fields = ['nombre', 'departamento__nombre', 'codigo']
    ordering_fields = ['nombre', 'departamento__nombre', 'created_at']
    ordering = ['departamento__nombre', 'nombre']
    
    def get_queryset(self):
        """Filtrar por organización si es necesario"""
        if hasattr(self.request.user, 'organization'):
            return Municipio.objects.select_related('departamento').filter(
                organizacion=self.request.user.organizacion
            )
        return Municipio.objects.select_related('departamento')
    
    def get_serializer_class(self):
        """Usar diferentes serializers según la acción"""
        if self.action == 'list':
            return MunicipioSimpleSerializer
        elif self.action in ['con_departamento', 'jerarquia']:
            return MunicipioConDepartamentoSerializer
        return MunicipioSerializer
    
    def perform_create(self, serializer):
        """Asignar organización al crear"""
        if hasattr(self.request.user, 'organization'):
            serializer.save(organizacion=self.request.user.organizacion)
    
    @action(detail=False, methods=['post'])
    def buscar(self, request):
        """Búsqueda avanzada de ubicaciones"""
        serializer = BusquedaUbicacionSerializer(data=request.data)
        if serializer.is_valid():
            query = serializer.validated_data.get('query')
            departamento_id = serializer.validated_data.get('departamento')
            
            queryset = self.get_queryset()
            
            if query:
                queryset = queryset.filter(
                    Q(nombre__icontains=query) |
                    Q(departamento__nombre__icontains=query) |
                    Q(codigo__icontains=query)
                )
            
            if departamento_id:
                queryset = queryset.filter(departamento_id=departamento_id)
            
            # Serializar resultados
            municipios_serializer = MunicipioSimpleSerializer(queryset[:50], many=True)
            return Response(municipios_serializer.data)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def importar_excel(self, request):
        """Importar municipios desde archivo Excel"""
        if 'archivo' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó archivo'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        archivo = request.FILES['archivo']
        
        # Validar formato
        if not archivo.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'El archivo debe ser Excel (.xlsx o .xls)'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Leer archivo Excel
            df = pd.read_excel(archivo)
            
            # Validar columnas requeridas
            columnas_requeridas = ['departamento', 'nombre']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            
            if columnas_faltantes:
                return Response(
                    {'error': f'Faltan columnas: {", ".join(columnas_faltantes)}'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Procesar datos
            municipios_creados = 0
            errores = []
            
            for index, row in df.iterrows():
                try:
                    # Validar datos obligatorios
                    if pd.isna(row['departamento']) or not str(row['departamento']).strip():
                        errores.append(f"Fila {index + 2}: Departamento es obligatorio")
                        continue
                    
                    if pd.isna(row['nombre']) or not str(row['nombre']).strip():
                        errores.append(f"Fila {index + 2}: Nombre es obligatorio")
                        continue
                    
                    # Buscar departamento
                    departamento_nombre = str(row['departamento']).strip()
                    try:
                        departamento = Departamento.objects.get(nombre=departamento_nombre)
                    except Departamento.DoesNotExist:
                        errores.append(f"Fila {index + 2}: Departamento '{departamento_nombre}' no encontrado")
                        continue
                    
                    # Crear municipio
                    municipio_data = {
                        'departamento': departamento,
                        'nombre': str(row['nombre']).strip(),
                        'codigo': str(row['codigo']).strip() if not pd.isna(row['codigo']) else '',
                    }
                    
                    # Asignar organización
                    if hasattr(request.user, 'organization'):
                        municipio_data['organization'] = request.user.organizacion
                    
                    # Verificar si ya existe
                    existing = Municipio.objects.filter(
                        departamento=departamento,
                        nombre=municipio_data['nombre']
                    )
                    if hasattr(request.user, 'organization'):
                        existing = existing.filter(organizacion=request.user.organizacion)
                    
                    if existing.exists():
                        errores.append(f"Fila {index + 2}: Municipio '{municipio_data['nombre']}' ya existe en {departamento_nombre}")
                        continue
                    
                    # Crear municipio
                    Municipio.objects.create(**municipio_data)
                    municipios_creados += 1
                    
                except Exception as e:
                    errores.append(f"Fila {index + 2}: Error - {str(e)}")
            
            return Response({
                'message': f'Importación completada. {municipios_creados} municipios creados.',
                'municipios_creados': municipios_creados,
                'errores': errores
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error al procesar archivo: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        """Exportar municipios a Excel"""
        try:
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Municipios"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Headers
            headers = ['ID', 'Departamento', 'Nombre', 'Código', 'Fecha Creación']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Datos
            queryset = self.get_queryset()
            
            for row, municipio in enumerate(queryset, 2):
                ws.cell(row=row, column=1, value=str(municipio.id))
                ws.cell(row=row, column=2, value=municipio.departamento.nombre)
                ws.cell(row=row, column=3, value=municipio.nombre)
                ws.cell(row=row, column=4, value=municipio.codigo)
                ws.cell(row=row, column=5, value=municipio.created_at.strftime('%Y-%m-%d'))
            
            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            
            # Crear respuesta HTTP
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="municipios.xlsx"'
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Error al exportar: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def template_importacion(self, request):
        """Descargar template para importación de municipios"""
        try:
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Template Municipios"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # Headers
            headers = ['departamento', 'nombre', 'codigo']
            help_text = ['Nombre del departamento (OBLIGATORIO)', 'Nombre del municipio (OBLIGATORIO)', 'Código DANE']
            
            for col, (header, help) in enumerate(zip(headers, help_text), 1):
                # Header
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                
                # Help text
                help_cell = ws.cell(row=2, column=col, value=help)
                help_cell.fill = example_fill
            
            # Ejemplos
            ejemplos = [
                ['Cundinamarca', 'Bogotá', '25001'],
                ['Cundinamarca', 'Soacha', '25754'],
                ['Antioquia', 'Medellín', '05001'],
                ['Antioquia', 'Bello', '05088'],
            ]
            
            for row, ejemplo in enumerate(ejemplos, 3):
                for col, valor in enumerate(ejemplo, 1):
                    ws.cell(row=row, column=col, value=valor)
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            
            # Crear respuesta HTTP
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_municipios.xlsx"'
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Error al generar template: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def jerarquia(self, request):
        """Obtener estructura jerárquica de departamentos y municipios"""
        departamentos = Departamento.objects.prefetch_related(
            Prefetch('municipios', queryset=Municipio.objects.order_by('nombre'))
        ).order_by('nombre')
        
        data = []
        for departamento in departamentos:
            municipios = departamento.municipios.all()
            data.append({
                'departamento': DepartamentoSimpleSerializer(departamento).data,
                'municipios': MunicipioSimpleSerializer(municipios, many=True).data
            })
        
        return Response(data)
    
    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Estadísticas de ubicaciones"""
        total_departamentos = Departamento.objects.count()
        total_municipios = Municipio.objects.count()
        
        # Departamentos con más municipios
        departamentos_stats = Departamento.objects.annotate(
            municipios_count=Count('municipios')
        ).order_by('-municipios_count')[:5]
        
        stats_data = []
        for dept in departamentos_stats:
            stats_data.append({
                'departamento': dept.nombre,
                'municipios_count': dept.municipios_count
            })
        
        return Response({
            'total_departamentos': total_departamentos,
            'total_municipios': total_municipios,
            'departamentos_con_mas_municipios': stats_data
        })


# API AJAX para formularios
def get_municipios_by_departamento(request, departamento_id):
    """API AJAX para obtener municipios por departamento"""
    try:
        municipios = Municipio.objects.filter(
            departamento_id=departamento_id
        ).order_by('nombre').values('id', 'nombre')
        
        return JsonResponse({
            'municipios': list(municipios)
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
