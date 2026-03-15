"""
API Views para Auditoría del Sistema
====================================

ViewSets y endpoints para gestión de logs de auditoría,
estadísticas, anomalías y reportes.

Autor: Sistema CorteSec
Versión: 1.0.0
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from core.policies.auditoria_access import AuditoriaAccessPolicy
from rest_framework.response import Response
from django.db.models import Count, Q, F
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import HttpResponse
import csv
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from .models import LogAuditoria
from .serializers import LogAuditoriaSerializer

logger = logging.getLogger(__name__)


class AuditoriaViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet para consulta de logs de auditoría
    Solo lectura - los logs se crean automáticamente por el sistema
    """
    serializer_class = LogAuditoriaSerializer
    permission_classes = [AuditoriaAccessPolicy]
    filterset_fields = ['accion', 'modelo', 'usuario']
    search_fields = ['accion', 'modelo', 'usuario__username', 'ip_address']
    ordering_fields = ['created_at', 'accion', 'modelo']
    ordering = ['-created_at']  # Por defecto ordenar por más reciente
    
    def list(self, request, *args, **kwargs):
        """Override list para agregar logs detallados"""
        logger.info(f"=== AUDITORIA LIST REQUEST ===")
        logger.info(f"User: {request.user}")
        logger.info(f"User authenticated: {request.user.is_authenticated}")
        logger.info(f"Auth header: {request.headers.get('Authorization', 'NO AUTH HEADER')}")
        
        if not request.user.is_authenticated:
            logger.error("❌ Usuario NO autenticado en list()")
            return Response({'error': 'Authentication required'}, status=401)
        
        logger.info(f"✅ Usuario autenticado: {request.user.username}")
        return super().list(request, *args, **kwargs)
    
    def get_queryset(self):
        """
        Filtra logs por organización del usuario (multi-tenant).
        Soporta filtros por fecha, acción, modelo, usuario.
        """
        queryset = LogAuditoria.objects.all().select_related('usuario')
        
        # ── Filtro multi-tenant ──
        # Superusers ven todo; staff normal solo ve logs de usuarios de su org
        user = self.request.user
        if user.is_authenticated and not user.is_superuser:
            # Obtener la organización del usuario
            org = getattr(user, 'organizacion', None)
            if org:
                from django.contrib.auth import get_user_model
                User = get_user_model()
                org_user_ids = User.objects.filter(organizacion=org).values_list('id', flat=True)
                queryset = queryset.filter(
                    Q(usuario__id__in=org_user_ids) | Q(usuario__isnull=True)
                )
        
        # Filtros opcionales
        fecha_inicio = self.request.query_params.get('fecha_inicio')
        fecha_fin = self.request.query_params.get('fecha_fin')
        accion = self.request.query_params.get('accion')
        modelo = self.request.query_params.get('modelo')
        usuario_id = self.request.query_params.get('usuario')
        
        if fecha_inicio:
            try:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                queryset = queryset.filter(created_at__gte=fecha_inicio_dt)
            except ValueError:
                pass
                
        if fecha_fin:
            try:
                fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
                # Incluir todo el día
                fecha_fin_dt = fecha_fin_dt.replace(hour=23, minute=59, second=59)
                queryset = queryset.filter(created_at__lte=fecha_fin_dt)
            except ValueError:
                pass
        
        if accion:
            queryset = queryset.filter(accion__icontains=accion)
            
        if modelo:
            queryset = queryset.filter(modelo__icontains=modelo)
            
        if usuario_id:
            queryset = queryset.filter(usuario_id=usuario_id)
        
        return queryset.order_by('-created_at')
    
    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """
        Estadísticas generales de auditoría
        """
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        
        queryset = self.get_queryset()
        
        # Estadísticas básicas
        total_eventos = queryset.count()
        usuarios_activos = queryset.values('usuario').distinct().count()
        modulos_activos = queryset.values('modelo').distinct().count()
        
        # Acciones más frecuentes
        acciones_frecuentes = list(
            queryset.values('accion')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )
        
        # Actividad por día (últimos 7 días)
        hoy = timezone.now()
        hace_7_dias = hoy - timedelta(days=7)
        
        actividad_diaria = []
        for i in range(7):
            fecha = hace_7_dias + timedelta(days=i)
            eventos = queryset.filter(
                created_at__date=fecha.date()
            ).count()
            actividad_diaria.append({
                'fecha': fecha.date().isoformat(),
                'eventos': eventos
            })
        
        return Response({
            'total_eventos': total_eventos,
            'usuarios_activos': usuarios_activos,
            'modulos_activos': modulos_activos,
            'acciones_frecuentes': acciones_frecuentes,
            'actividad_diaria': actividad_diaria
        })
    
    @action(detail=False, methods=['POST'], permission_classes=[AuditoriaAccessPolicy], url_path='log-frontend')
    def log_frontend(self, request):
        """
        Endpoint para recibir logs del frontend.
        Crea múltiples logs en una sola petición (batch).
        Valida inputs para evitar inyección de datos falsos.
        """
        logs_data = request.data.get('logs', [])
        
        if not logs_data or not isinstance(logs_data, list):
            return Response(
                {'error': 'Se requiere un array de logs'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Limitar batch size
        if len(logs_data) > 50:
            return Response(
                {'error': 'Máximo 50 logs por batch'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Acciones y modelos permitidos desde el frontend
        ALLOWED_ACTIONS = {
            'navegar_pagina', 'ver_pagina', 'click_boton', 'abrir_modal',
            'cerrar_modal', 'buscar', 'cambiar_tab', 'exportar', 'filtrar',
            'accion_frontend', 'inicio_sesion_frontend', 'cerrar_sesion_frontend',
            'ver_detalle', 'seleccionar', 'cambiar_filtro', 'scroll',
        }
        ALLOWED_MODELS = {
            'Frontend', 'Navegacion', 'Dashboard', 'Configuracion',
            'Auditoria', 'Empleados', 'Nomina', 'Prestamos', 'Usuarios',
            'Roles', 'Permisos', 'Organizaciones', 'Planes', 'Reportes',
            'Conceptos', 'Municipios', 'Parametros', 'Seguridad', 'Email',
        }
        
        created_logs = []
        ip_address = self._get_client_ip(request)
        
        for log_item in logs_data:
            try:
                accion = str(log_item.get('accion', 'accion_frontend'))[:100]
                modelo = str(log_item.get('modelo', 'Frontend'))[:100]
                
                # Sanitizar: solo permitir acciones/modelos conocidos
                if accion not in ALLOWED_ACTIONS:
                    accion = 'accion_frontend'
                if modelo not in ALLOWED_MODELS:
                    modelo = 'Frontend'
                
                log = LogAuditoria.objects.create(
                    usuario=request.user,
                    accion=accion,
                    modelo=modelo,
                    objeto_id=str(log_item.get('objeto_id', ''))[:100] if log_item.get('objeto_id') else None,
                    ip_address=ip_address,
                    user_agent=str(log_item.get('user_agent', request.META.get('HTTP_USER_AGENT', '')))[:255],
                    metadata=log_item.get('metadata', {}) if isinstance(log_item.get('metadata'), dict) else {}
                )
                created_logs.append(str(log.id))
            except Exception as e:
                logger.error(f"Error creando log frontend: {e}")
                continue
        
        return Response({
            'success': True,
            'created': len(created_logs),
            'message': f'Se crearon {len(created_logs)} logs correctamente'
        }, status=status.HTTP_201_CREATED)
    
    def _get_client_ip(self, request):
        """Obtiene la IP real del cliente"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0].strip()
        else:
            ip = request.META.get('REMOTE_ADDR', '')
        return ip[:45]
    
    @action(detail=False, methods=['get'])
    def actividad_usuarios(self, request):
        """
        Actividad agrupada por usuario
        """
        queryset = self.get_queryset()
        
        actividad = list(
            queryset.values('usuario__username', 'usuario__id')
            .annotate(total_acciones=Count('id'))
            .order_by('-total_acciones')[:20]
        )
        
        resultado = [
            {
                'usuario': item['usuario__username'] or 'Sistema',
                'usuario_id': item['usuario__id'],
                'total_acciones': item['total_acciones']
            }
            for item in actividad
        ]
        
        return Response(resultado)
    
    @action(detail=False, methods=['get'])
    def actividad_modulos(self, request):
        """
        Actividad agrupada por módulo
        """
        queryset = self.get_queryset()
        
        actividad = list(
            queryset.values('modelo')
            .annotate(total_eventos=Count('id'))
            .order_by('-total_eventos')[:20]
        )
        
        return Response(actividad)
    
    @action(detail=False, methods=['get'])
    def linea_tiempo(self, request):
        """
        Línea de tiempo de eventos
        """
        agrupacion = request.query_params.get('agrupacion', 'hora')
        queryset = self.get_queryset()
        
        if agrupacion == 'hora':
            # Últimas 24 horas
            hace_24h = timezone.now() - timedelta(hours=24)
            datos = []
            for i in range(24):
                hora = hace_24h + timedelta(hours=i)
                eventos = queryset.filter(
                    created_at__hour=hora.hour,
                    created_at__date=hora.date()
                ).count()
                datos.append({
                    'timestamp': hora.isoformat(),
                    'eventos': eventos
                })
        elif agrupacion == 'dia':
            # Últimos 30 días
            hace_30d = timezone.now() - timedelta(days=30)
            datos = []
            for i in range(30):
                dia = hace_30d + timedelta(days=i)
                eventos = queryset.filter(created_at__date=dia.date()).count()
                datos.append({
                    'timestamp': dia.date().isoformat(),
                    'eventos': eventos
                })
        else:
            datos = []
        
        return Response(datos)
    
    @action(detail=False, methods=['get'])
    def anomalias(self, request):
        """
        Detecta anomalías en el sistema
        - Múltiples intentos fallidos
        - Actividad inusual por horario
        - IPs sospechosas
        """
        queryset = self.get_queryset()
        anomalias = []
        
        # Detectar múltiples acciones del mismo usuario en poco tiempo
        hace_1h = timezone.now() - timedelta(hours=1)
        usuarios_activos = queryset.filter(created_at__gte=hace_1h).values('usuario__username').annotate(
            total=Count('id')
        ).filter(total__gt=50)
        
        for usuario in usuarios_activos:
            anomalias.append({
                'tipo': 'Actividad excesiva',
                'nivel': 'alto',
                'descripcion': f"Usuario {usuario['usuario__username']} con {usuario['total']} acciones en la última hora",
                'timestamp': timezone.now().isoformat(),
                'usuario': usuario['usuario__username']
            })
        
        # Detectar accesos desde IPs diferentes
        hace_24h = timezone.now() - timedelta(hours=24)
        usuarios_multiples_ips = queryset.filter(created_at__gte=hace_24h).values('usuario__username').annotate(
            ips_count=Count('ip_address', distinct=True)
        ).filter(ips_count__gt=3)
        
        for usuario in usuarios_multiples_ips:
            anomalias.append({
                'tipo': 'Múltiples IPs',
                'nivel': 'medio',
                'descripcion': f"Usuario {usuario['usuario__username']} accedió desde {usuario['ips_count']} IPs diferentes en 24h",
                'timestamp': timezone.now().isoformat(),
                'usuario': usuario['usuario__username']
            })
        
        return Response(anomalias)
    
    @action(detail=False, methods=['get'])
    def accesos_fallidos(self, request):
        """
        Obtiene intentos de acceso fallidos
        """
        queryset = self.get_queryset()
        
        # Filtrar logs de login fallido
        fallidos = queryset.filter(
            Q(accion__icontains='login_fallido') | 
            Q(accion__icontains='failed') |
            Q(accion__icontains='error')
        ).order_by('-created_at')[:50]
        
        # Agrupar por usuario e IP
        agrupados = {}
        for log in fallidos:
            key = f"{log.usuario.username if log.usuario else 'Anónimo'}_{log.ip_address}"
            if key not in agrupados:
                agrupados[key] = {
                    'usuario': log.usuario.username if log.usuario else 'Anónimo',
                    'ip_address': log.ip_address,
                    'timestamp': log.created_at.isoformat(),
                    'intentos': 0
                }
            agrupados[key]['intentos'] += 1
        
        return Response(list(agrupados.values()))
    
    @action(detail=False, methods=['get'])
    def exportar_csv(self, request):
        """
        Exporta logs a CSV
        """
        queryset = self.get_queryset()[:1000]  # Limitar a 1000 registros
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="auditoria_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Fecha/Hora', 'Usuario', 'Acción', 'Módulo', 'IP', 'User Agent'])
        
        for log in queryset:
            writer.writerow([
                log.id,
                log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                log.usuario.username if log.usuario else 'Sistema',
                log.accion,
                log.modelo,
                log.ip_address or '',
                log.user_agent or ''
            ])
        
        return response
    
    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        """
        Exporta logs a Excel
        """
        queryset = self.get_queryset()[:1000]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Auditoría"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ['ID', 'Fecha/Hora', 'Usuario', 'Acción', 'Módulo', 'IP', 'User Agent']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row, log in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=log.id)
            ws.cell(row=row, column=2, value=log.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=3, value=log.usuario.username if log.usuario else 'Sistema')
            ws.cell(row=row, column=4, value=log.accion)
            ws.cell(row=row, column=5, value=log.modelo)
            ws.cell(row=row, column=6, value=log.ip_address or '')
            ws.cell(row=row, column=7, value=log.user_agent or '')

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 24

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="auditoria_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        )
        wb.save(response)
        return response
    
    @action(detail=False, methods=['post'])
    def busqueda_avanzada(self, request):
        """
        Búsqueda avanzada con múltiples criterios
        """
        criterios = request.data
        queryset = self.get_queryset()
        
        if 'texto' in criterios and criterios['texto']:
            texto = criterios['texto']
            queryset = queryset.filter(
                Q(accion__icontains=texto) |
                Q(modelo__icontains=texto) |
                Q(usuario__username__icontains=texto)
            )
        
        if 'fecha_desde' in criterios and criterios['fecha_desde']:
            queryset = queryset.filter(created_at__gte=criterios['fecha_desde'])
            
        if 'fecha_hasta' in criterios and criterios['fecha_hasta']:
            queryset = queryset.filter(created_at__lte=criterios['fecha_hasta'])
        
        serializer = self.get_serializer(queryset[:100], many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='filtros-disponibles')
    def filtros_disponibles(self, request):
        """
        Retorna las acciones y módulos únicos existentes en los logs.
        Usado por el frontend para llenar dropdowns dinámicamente.
        """
        queryset = self.get_queryset()

        acciones = list(
            queryset.values_list('accion', flat=True)
            .distinct()
            .order_by('accion')
        )
        modelos = list(
            queryset.values_list('modelo', flat=True)
            .distinct()
            .order_by('modelo')
        )

        return Response({
            'acciones': acciones,
            'modelos': modelos
        })
