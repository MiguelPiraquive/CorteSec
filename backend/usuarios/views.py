"""
ASOGAN - Vistas de Usuarios
API REST para gestion de usuarios completa y profesional
"""

import logging
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth import get_user_model
from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.conf import settings
from core.email_service import send_system_email
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter

from .models import GrupoUsuario, HistorialUsuario
from .serializers import (
    UserListSerializer, UserDetailSerializer, UserCreateSerializer,
    UserUpdateSerializer, UserPasswordChangeSerializer, UserProfileSerializer,
    UserStatsSerializer, UserBulkActionSerializer, UserBulkRoleAssignSerializer
)
from .filters import UserFilter
from .permissions import (
    CanViewUsers, CanCreateUsers, CanEditUsers, CanDeleteUsers,
    CanActivateUsers, CanChangeRoles, CanViewStats, CanExportUsers
)
from roles.models import Rol, AsignacionRol, EstadoAsignacion
from auditing.decorators import audit_user_operation

User = get_user_model()
logger = logging.getLogger(__name__)


class UserViewSet(viewsets.ModelViewSet):
    """
    ViewSet completo para gestión de usuarios
    
    Acciones disponibles:
    - list: Listar usuarios con paginación y filtros
    - retrieve: Ver detalle de un usuario
    - create: Crear nuevo usuario
    - update/partial_update: Actualizar usuario
    - destroy: Eliminar usuario (soft delete)
    - activate/deactivate: Activar/desactivar usuario
    - reset_password: Resetear contraseña
    - change_password: Cambiar contraseña (admin)
    - change_role: Cambiar rol del usuario
    - user_activity: Ver actividad del usuario
    """
    
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = UserFilter
    search_fields = ['username', 'first_name', 'last_name', 'email']
    ordering_fields = ['date_joined', 'first_name', 'last_name', 'email', 'last_login']
    ordering = ['-date_joined']
    
    def get_queryset(self):
        """Filtrar usuarios por organización del usuario actual"""
        user = self.request.user
        if user.organization:
            return User.objects.filter(
                organization=user.organization
            ).select_related('organization')
        # Sin organización, solo ve su propio usuario
        return User.objects.filter(id=user.id)
    
    def get_serializer_class(self):
        """Seleccionar el serializer según la acción"""
        if self.action == 'list':
            return UserListSerializer
        elif self.action == 'retrieve':
            return UserDetailSerializer
        elif self.action == 'create':
            return UserCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return UserUpdateSerializer
        elif self.action == 'change_password':
            return UserPasswordChangeSerializer
        return UserDetailSerializer
    
    def get_permissions(self):
        """Permisos según la acción"""
        if self.action in ['list', 'retrieve']:
            permission_classes = [IsAuthenticated, CanViewUsers]
        elif self.action == 'create':
            permission_classes = [IsAuthenticated, CanCreateUsers]
        elif self.action in ['update', 'partial_update']:
            permission_classes = [IsAuthenticated, CanEditUsers]
        elif self.action == 'destroy':
            permission_classes = [IsAuthenticated, CanDeleteUsers]
        elif self.action in ['activate', 'deactivate']:
            permission_classes = [IsAuthenticated, CanActivateUsers]
        elif self.action in ['change_role', 'change_password']:
            permission_classes = [IsAuthenticated, CanChangeRoles]
        elif self.action in ['reset_password', 'resend_verification_email', 'resetear_contrasena']:
            permission_classes = [IsAuthenticated, CanEditUsers]
        elif self.action in ['activity', 'historial', 'get_permisos', 'estadisticas']:
            permission_classes = [IsAuthenticated, CanViewUsers]
        elif self.action == 'exportar':
            permission_classes = [IsAuthenticated, CanExportUsers]
        elif self.action in ['verificar_username', 'verificar_email']:
            permission_classes = [IsAuthenticated, CanCreateUsers]
        else:
            permission_classes = [IsAuthenticated, CanViewUsers]

        return [permission() for permission in permission_classes]
    
    @audit_user_operation('listar_usuarios')
    def list(self, request, *args, **kwargs):
        """Listar usuarios con filtros y paginación"""
        return super().list(request, *args, **kwargs)
    
    @audit_user_operation('ver_usuario')
    def retrieve(self, request, *args, **kwargs):
        """Ver detalle de un usuario"""
        return super().retrieve(request, *args, **kwargs)
    
    @audit_user_operation('crear_usuario')
    def create(self, request, *args, **kwargs):
        """Crear nuevo usuario"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        enviar_email = request.data.get('enviar_email_verificacion', True)
        user = serializer.save()
        
        # Asignar la organización del usuario que crea
        if not user.organization and request.user.organization:
            user.organization = request.user.organization
            user.save(update_fields=['organization'])
        
        # Enviar email de verificación si está activado
        if enviar_email:
            try:
                token = default_token_generator.make_token(user)
                uid = urlsafe_base64_encode(force_bytes(user.pk))
                verification_url = f"{settings.FRONTEND_URL}/verificar-email/{uid}/{token}/"

                send_system_email(
                    subject='Verificación de Email - CorteSec',
                    message=(
                    f"Hola {user.get_full_name() or user.email},\n\n"
                    "Bienvenido a CorteSec. Por favor verifica tu email en el siguiente enlace:\n"
                    f"{verification_url}\n\n"
                    "Saludos,\nEquipo CorteSec"),
                    recipient_list=[user.email],
                    fail_silently=False
                )
            except Exception as e:
                logger.warning(f"Error enviando email de verificación al crear usuario {user.email}: {e}")
        
        headers = self.get_success_headers(serializer.data)
        return Response(
            UserDetailSerializer(user).data,
            status=status.HTTP_201_CREATED,
            headers=headers
        )
    
    @audit_user_operation('actualizar_usuario')
    def update(self, request, *args, **kwargs):
        """Actualizar usuario completo"""
        return super().update(request, *args, **kwargs)
    
    @audit_user_operation('actualizar_usuario_parcial')
    def partial_update(self, request, *args, **kwargs):
        """Actualización parcial de usuario"""
        return super().partial_update(request, *args, **kwargs)
    
    @audit_user_operation('eliminar_usuario')
    def destroy(self, request, *args, **kwargs):
        """Eliminar usuario (soft delete)"""
        instance = self.get_object()
        
        # No permitir auto-eliminación
        if instance.id == request.user.id:
            return Response(
                {'error': 'No puedes eliminarte a ti mismo.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Soft delete
        instance.deleted = True
        instance.deleted_at = timezone.now()
        instance.deleted_by = request.user
        instance.save()
        
        return Response(
            {'message': 'Usuario eliminado exitosamente.'},
            status=status.HTTP_204_NO_CONTENT
        )
    
    @action(detail=True, methods=['post'])
    @audit_user_operation('activar_usuario')
    def activate(self, request, pk=None):
        """Activar usuario"""
        user = self.get_object()
        
        if user.id == request.user.id:
            return Response(
                {'error': 'No puedes modificar tu propio estado.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user.is_active = True
        user.save()
        
        return Response({
            'message': f'Usuario {user.email} activado exitosamente.',
            'user': UserDetailSerializer(user).data
        })
    
    @action(detail=True, methods=['post'])
    @audit_user_operation('desactivar_usuario')
    def deactivate(self, request, pk=None):
        """Desactivar usuario"""
        user = self.get_object()
        
        if user.id == request.user.id:
            return Response(
                {'error': 'No puedes desactivarte a ti mismo.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user.is_active = False
        user.save()
        
        return Response({
            'message': f'Usuario {user.email} desactivado exitosamente.',
            'user': UserDetailSerializer(user).data
        })
    
    @action(detail=True, methods=['post'], url_path='cambiar-contrasena')
    @audit_user_operation('cambiar_password_usuario')
    def change_password(self, request, pk=None):
        """Cambiar contraseña de un usuario (admin)"""
        user = self.get_object()
        serializer = UserPasswordChangeSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Cambiar contraseña
        user.set_password(serializer.validated_data['nueva_password'])
        user.save()
        
        # Enviar email si está activado
        if serializer.validated_data.get('enviar_email', True):
            try:
                send_system_email(
                    subject='Cambio de contraseña - CorteSec',
                    message=(
                    f"Hola {user.get_full_name() or user.email},\n\n"
                    "Tu contraseña fue cambiada por un administrador.\n"
                    "Si no reconoces este cambio, contacta al soporte.\n\n"
                    "Saludos,\nEquipo CorteSec"),
                    recipient_list=[user.email],
                    fail_silently=False
                )
            except Exception as e:
                logger.error(f"Error enviando email de cambio de contraseña: {e}")
        
        return Response({
            'message': f'Contraseña cambiada exitosamente para {user.email}.'
        })
    
    @action(detail=True, methods=['post'])
    @audit_user_operation('resetear_password_usuario')
    def reset_password(self, request, pk=None):
        """Generar y enviar token de reseteo de contraseña"""
        user = self.get_object()
        try:
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            reset_url = f"{settings.FRONTEND_URL}/reset-password/{uid}/{token}"

            send_system_email(
                subject='Recuperación de contraseña - CorteSec',
                message=(
                f"Hola {user.get_full_name() or user.email},\n\n"
                "Se solicitó un restablecimiento de contraseña para tu cuenta.\n\n"
                f"Enlace de recuperación (válido por 24 horas):\n{reset_url}\n\n"
                "Si no solicitaste este cambio, ignora este correo.\n\n"
                "Saludos,\nEquipo CorteSec"),
                recipient_list=[user.email],
                fail_silently=False
            )

            return Response({
                'message': f'Email de reseteo enviado a {user.email}.'
            })
        except Exception as e:
            logger.error(f"Error enviando email de reseteo: {e}")
            return Response(
                {'error': 'No se pudo enviar el correo de recuperación.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    @audit_user_operation('reenviar_email_verificacion')
    def resend_verification_email(self, request, pk=None):
        """Reenviar email de verificación"""
        user = self.get_object()
        
        # Verificar si el email ya está verificado
        if user.email_verified:
            return Response(
                {'error': 'El email ya está verificado.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            verification_url = f"{settings.FRONTEND_URL}/verificar-email/{uid}/{token}/"

            send_system_email(
                subject='Verificación de Email - CorteSec',
                message=(
                f"Hola {user.get_full_name() or user.email},\n\n"
                "Por favor verifica tu email haciendo clic en el siguiente enlace:\n"
                f"{verification_url}\n\n"
                "Si no solicitaste este correo, ignóralo.\n\n"
                "Saludos,\nEquipo CorteSec"),
                recipient_list=[user.email],
                fail_silently=False
            )

            return Response({
                'message': 'Email de verificación reenviado exitosamente.',
                'email': user.email
            })
        except Exception as e:
            logger.error(f"Error reenviando email de verificación: {e}")
            return Response(
                {'error': 'No se pudo reenviar el email de verificación.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'], url_path='asignar_roles')
    @audit_user_operation('cambiar_rol_usuario')
    def change_role(self, request, pk=None):
        """Asignar uno o varios roles al usuario"""
        user = self.get_object()
        roles_ids = request.data.get('roles_ids') or []
        rol_id = request.data.get('rol_id')

        if rol_id and not roles_ids:
            roles_ids = [rol_id]

        if not roles_ids:
            return Response(
                {'error': 'Debe proporcionar roles_ids o rol_id.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        roles = list(Rol.objects.filter(id__in=roles_ids))
        if not roles:
            return Response({'error': 'Roles no encontrados.'}, status=status.HTTP_404_NOT_FOUND)

        # Desactivar asignaciones previas
        estado_activo, _ = EstadoAsignacion.objects.get_or_create(
            nombre='Activa',
            defaults={'activo': True, 'color': '#22c55e'}
        )
        AsignacionRol.objects.filter(usuario=user, activa=True).update(activa=False)

        for rol in roles:
            AsignacionRol.objects.get_or_create(
                usuario=user,
                rol=rol,
                defaults={
                    'activa': True,
                    'estado': estado_activo,
                    'asignado_por': request.user,
                    'organization': getattr(request.user, 'organization', None),
                }
            )

        return Response({
            'message': 'Roles asignados exitosamente.',
            'roles': [{'id': r.id, 'nombre': r.nombre, 'codigo': r.codigo} for r in roles]
        })
    
    @action(detail=True, methods=['get'])
    def activity(self, request, pk=None):
        """Ver actividad del usuario"""
        user = self.get_object()

        # Obtener historial de auditoría
        try:
            from auditing.models import AuditLog
            logs = AuditLog.objects.filter(user=user).order_by('-timestamp')[:50]

            activity_data = [{
                'accion': log.action,
                'descripcion': log.description,
                'fecha': log.timestamp,
                'ip': log.ip_address
            } for log in logs]
        except Exception:
            activity_data = []

        return Response({
            'usuario': UserDetailSerializer(user).data,
            'actividad_reciente': activity_data
        })

    @action(detail=True, methods=['get'], url_path='historial')
    def historial(self, request, pk=None):
        """Ver historial de actividad del usuario"""
        user = self.get_object()

        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))

        historial = HistorialUsuario.objects.filter(usuario=user).order_by('-fecha')
        total = historial.count()

        start = (page - 1) * page_size
        end = start + page_size
        historial_pagina = historial[start:end]

        results = [{
            'id': h.id,
            'accion': h.accion,
            'descripcion': h.descripcion,
            'ip_address': h.ip_address,
            'fecha': h.fecha
        } for h in historial_pagina]

        return Response({
            'results': results,
            'count': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size
        })

    @action(detail=True, methods=['get'], url_path='permisos')
    def get_permisos(self, request, pk=None):
        """Ver permisos del usuario"""
        user = self.get_object()
        serializer = UserDetailSerializer(user)
        return Response({
            'usuario': user.username,
            'permisos': serializer.data.get('permisos', [])
        })

    @action(detail=False, methods=['get'], url_path='estadisticas')
    def estadisticas(self, request):
        """Estadisticas generales de usuarios (filtradas por organización)"""
        qs = self.get_queryset()
        total_usuarios = qs.count()
        usuarios_activos = qs.filter(is_active=True).count()
        usuarios_inactivos = qs.filter(is_active=False).count()
        administradores = qs.filter(Q(is_staff=True) | Q(is_superuser=True)).count()

        return Response({
            'total_usuarios': total_usuarios,
            'usuarios_activos': usuarios_activos,
            'usuarios_inactivos': usuarios_inactivos,
            'administradores': administradores
        })

    @action(detail=False, methods=['get', 'post'], url_path='verificar-username')
    def verificar_username(self, request):
        """Verificar disponibilidad de username"""
        username = request.data.get('username') or request.query_params.get('username')

        if not username:
            return Response(
                {'error': 'Username es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        disponible = not self.get_queryset().filter(username=username).exists()

        return Response({
            'disponible': disponible,
            'mensaje': 'Username disponible' if disponible else 'Username ya esta en uso'
        })

    @action(detail=False, methods=['get', 'post'], url_path='verificar-email')
    def verificar_email(self, request):
        """Verificar disponibilidad de email"""
        email = request.data.get('email') or request.query_params.get('email')

        if not email:
            return Response(
                {'error': 'Email es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        disponible = not self.get_queryset().filter(email=email).exists()

        return Response({
            'disponible': disponible,
            'mensaje': 'Email disponible' if disponible else 'Email ya esta registrado'
        })

    @action(detail=False, methods=['post'], url_path='resetear-contrasena')
    def resetear_contrasena(self, request):
        """Enviar email de reseteo de contrasena"""
        email = request.data.get('email')

        if not email:
            return Response(
                {'error': 'Email es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user = User.objects.filter(email=email).first()

            # Por seguridad, no revelar si existe
            if not user:
                return Response({
                    'message': 'Si el email existe, recibiras instrucciones para resetear tu contrasena'
                })

            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            reset_url = f"{settings.FRONTEND_URL}/reset-password/{uid}/{token}"

            send_system_email(
                subject='Recuperacion de contrasena - CorteSec',
                message=(
                f"Hola {user.get_full_name() or user.email},\n\n"
                "Se solicito un restablecimiento de contrasena para tu cuenta.\n\n"
                f"Enlace de recuperacion (valido por 24 horas):\n{reset_url}\n\n"
                "Si no solicitaste este cambio, ignora este correo.\n\n"
                "Saludos,\nEquipo CorteSec"),
                recipient_list=[user.email],
                fail_silently=False
            )

            return Response({
                'message': f'Email de reseteo enviado a {email}'
            })
        except Exception as e:
            logger.error(f"Error enviando email de reseteo: {e}")
            return Response(
                {'error': 'No se pudo enviar el correo de recuperacion.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='exportar')
    def exportar(self, request):
        """Exportar usuarios a Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse

        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ['ID', 'Usuario', 'Email', 'Nombre', 'Apellido', 'Activo', 'Staff', 'Superusuario', 'Fecha Registro']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row, user in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=user.id)
            ws.cell(row=row, column=2, value=user.username)
            ws.cell(row=row, column=3, value=user.email)
            ws.cell(row=row, column=4, value=user.first_name)
            ws.cell(row=row, column=5, value=user.last_name)
            ws.cell(row=row, column=6, value="Si" if user.is_active else "No")
            ws.cell(row=row, column=7, value="Si" if user.is_staff else "No")
            ws.cell(row=row, column=8, value="Si" if user.is_superuser else "No")
            ws.cell(row=row, column=9, value=user.date_joined.strftime('%Y-%m-%d %H:%M'))

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'
        wb.save(response)

        return response


class UserStatsViewSet(viewsets.ViewSet):
    """ViewSet para estadísticas de usuarios"""
    
    permission_classes = [IsAuthenticated, CanViewStats]
    
    def _get_org_queryset(self, request):
        """Obtener queryset filtrado por organización"""
        if request.user.organization:
            return User.objects.filter(organization=request.user.organization)
        return User.objects.filter(id=request.user.id)
    
    @action(detail=False, methods=['get'])
    @audit_user_operation('ver_estadisticas_usuarios')
    def general(self, request):
        """Estadísticas generales de usuarios (filtradas por organización)"""
        
        # Contar usuarios
        qs = self._get_org_queryset(request)
        total_usuarios = qs.count()
        usuarios_activos = qs.filter(is_active=True).count()
        usuarios_inactivos = qs.filter(is_active=False).count()
        usuarios_verificados = qs.filter(email_verified=True).count()
        usuarios_sin_verificar = qs.filter(email_verified=False).count()
        usuarios_staff = qs.filter(is_staff=True).count()
        
        # Usuarios por rol - COMENTADO porque CustomUser no tiene campo 'rol'
        usuarios_por_rol = {}
        # roles_stats = User.objects.values('rol__nombre').annotate(
        #     total=Count('id')
        # ).order_by('-total')
        # 
        # for stat in roles_stats:
        #     rol_nombre = stat['rol__nombre'] or 'Sin rol'
        #     usuarios_por_rol[rol_nombre] = stat['total']
        
        # Usuarios recientes (últimos 7 días)
        hace_7_dias = timezone.now() - timedelta(days=7)
        usuarios_recientes = qs.filter(
            date_joined__gte=hace_7_dias
        ).order_by('-date_joined')[:10]
        
        usuarios_recientes_data = [{
            'id': u.id,
            'nombre_completo': f"{u.nombre} {u.apellido}",
            'email': u.email,
            'fecha_registro': u.date_joined,
            # 'rol': u.rol.nombre if u.rol else None  # COMENTADO
        } for u in usuarios_recientes]
        
        # Serializar
        stats = {
            'total_usuarios': total_usuarios,
            'usuarios_activos': usuarios_activos,
            'usuarios_inactivos': usuarios_inactivos,
            'usuarios_verificados': usuarios_verificados,
            'usuarios_sin_verificar': usuarios_sin_verificar,
            'usuarios_staff': usuarios_staff,
            'usuarios_por_rol': usuarios_por_rol,
            'usuarios_recientes': usuarios_recientes_data
        }
        
        serializer = UserStatsSerializer(stats)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def active_users(self, request):
        """Usuarios activos (filtrados por organización)"""
        qs = self._get_org_queryset(request)
        usuarios = qs.filter(is_active=True)
        serializer = UserListSerializer(usuarios, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def users_by_role(self, request):
        """Distribución de usuarios por rol (filtrada por organización)"""
        org_filter = {}
        if request.user.organization:
            org_filter['organization'] = request.user.organization
        stats = AsignacionRol.objects.filter(activa=True, **org_filter).values(
            'rol__id', 'rol__nombre'
        ).annotate(total=Count('id')).order_by('-total')

        return Response({
            'distribución': [{
                'rol_id': s['rol__id'],
                'rol_nombre': s['rol__nombre'],
                'total_usuarios': s['total']
            } for s in stats]
        })
    
    @action(detail=False, methods=['get'])
    def recent_registrations(self, request):
        """Registros recientes (filtrados por organización)"""
        dias = int(request.query_params.get('dias', 30))
        fecha_desde = timezone.now() - timedelta(days=dias)
        
        qs = self._get_org_queryset(request)
        usuarios = qs.filter(
            date_joined__gte=fecha_desde
        ).order_by('-date_joined')
        
        serializer = UserListSerializer(usuarios, many=True)
        return Response({
            'dias': dias,
            'total': usuarios.count(),
            'usuarios': serializer.data
        })


class UserBulkActionsViewSet(viewsets.ViewSet):
    """ViewSet para acciones masivas sobre usuarios"""
    
    permission_classes = [IsAuthenticated, CanEditUsers]
    
    def _get_org_queryset(self, request):
        """Obtener queryset filtrado por organización"""
        if request.user.organization:
            return User.objects.filter(organization=request.user.organization)
        return User.objects.filter(id=request.user.id)
    
    @action(detail=False, methods=['post'])
    @audit_user_operation('activar_usuarios_masivo')
    def bulk_activate(self, request):
        """Activar múltiples usuarios"""
        serializer = UserBulkActionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        user_ids = serializer.validated_data['user_ids']
        
        # No permitir modificar el usuario actual
        if request.user.id in user_ids:
            user_ids.remove(request.user.id)
        
        usuarios = self._get_org_queryset(request).filter(id__in=user_ids)
        count = usuarios.update(is_active=True)
        
        return Response({
            'message': f'{count} usuario(s) activado(s) exitosamente.',
            'usuarios_modificados': count
        })
    
    @action(detail=False, methods=['post'])
    @audit_user_operation('desactivar_usuarios_masivo')
    def bulk_deactivate(self, request):
        """Desactivar múltiples usuarios"""
        serializer = UserBulkActionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        user_ids = serializer.validated_data['user_ids']
        
        # No permitir modificar el usuario actual
        if request.user.id in user_ids:
            user_ids.remove(request.user.id)
        
        usuarios = self._get_org_queryset(request).filter(id__in=user_ids)
        count = usuarios.update(is_active=False)
        
        return Response({
            'message': f'{count} usuario(s) desactivado(s) exitosamente.',
            'usuarios_modificados': count
        })
    
    @action(detail=False, methods=['post'])
    @audit_user_operation('asignar_rol_masivo')
    def bulk_assign_role(self, request):
        """Asignar rol a múltiples usuarios"""
        serializer = UserBulkRoleAssignSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        user_ids = serializer.validated_data['user_ids']
        rol_id = serializer.validated_data['rol_id']

        try:
            rol = Rol.objects.get(id=rol_id)
            usuarios = self._get_org_queryset(request).filter(id__in=user_ids)

            count = 0
            estado_activo, _ = EstadoAsignacion.objects.get_or_create(
                nombre='Activa',
                defaults={'activo': True, 'color': '#22c55e'}
            )
            for usuario in usuarios:
                AsignacionRol.objects.update_or_create(
                    usuario=usuario,
                    rol=rol,
                    defaults={
                        'activa': True,
                        'estado': estado_activo,
                        'asignado_por': request.user,
                        'organization': getattr(usuario, 'organization', None),
                    }
                )
                count += 1

            return Response({
                'message': f'{count} usuario(s) asignado(s) al rol {rol.nombre}.',
                'usuarios_modificados': count,
                'rol_asignado': rol.nombre
            })
        except Rol.DoesNotExist:
            return Response(
                {'error': 'Rol no encontrado.'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=False, methods=['post'])
    @audit_user_operation('eliminar_usuarios_masivo')
    def bulk_delete(self, request):
        """Eliminar múltiples usuarios (soft delete)"""
        serializer = UserBulkActionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        user_ids = serializer.validated_data['user_ids']
        
        # No permitir eliminar el usuario actual
        if request.user.id in user_ids:
            user_ids.remove(request.user.id)
        
        usuarios = self._get_org_queryset(request).filter(id__in=user_ids)
        count = usuarios.update(
            deleted=True,
            deleted_at=timezone.now(),
            deleted_by=request.user
        )
        
        return Response({
            'message': f'{count} usuario(s) eliminado(s) exitosamente.',
            'usuarios_modificados': count
        })


# ViewSets existentes (mantener compatibilidad)
class GrupoUsuarioViewSet(viewsets.ModelViewSet):
    """ViewSet para grupos de usuarios"""
    
    queryset = GrupoUsuario.objects.all()
    permission_classes = [IsAuthenticated]


class HistorialUsuarioViewSet(viewsets.ModelViewSet):
    """ViewSet para historial de usuarios"""
    
    queryset = HistorialUsuario.objects.all()
    permission_classes = [IsAuthenticated]


