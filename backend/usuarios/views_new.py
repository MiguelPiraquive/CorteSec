"""
Vistas API del Módulo de Usuarios - CorteSec
==========================================

API REST para gestión completa de usuarios con Multi-Tenancy y Roles Personalizados.

Autor: Sistema CorteSec
Versión: 2.0.0
Fecha: 2026-01-01
"""

import logging
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth import get_user_model
from django.db.models import Q, Count
from django.utils import timezone
from datetime import timedelta
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .models import HistorialUsuario
from .serializers_new import (
    UserListSerializer, UserDetailSerializer, UserCreateSerializer,
    UserUpdateSerializer, CambiarContrasenaSerializer, AsignarRolesSerializer,
    UserStatsSerializer, HistorialUsuarioSerializer, VerificarDisponibilidadSerializer
)
from .filters import UserFilter
from roles.models import Rol, AsignacionRol

User = get_user_model()
logger = logging.getLogger(__name__)


class UsuarioViewSet(viewsets.ModelViewSet):
    """
    ViewSet completo para gestión de usuarios con Multi-Tenancy
    
    Funcionalidades:
    - Filtrado automático por organización
    - CRUD completo de usuarios
    - Asignación de roles personalizados
    - Estadísticas y reportes
    - Exportación a Excel
    - Historial de actividad
    """
    
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = UserFilter
    search_fields = ['username', 'email', 'full_name', 'first_name', 'last_name']
    ordering_fields = ['date_joined', 'username', 'email', 'last_login']
    ordering = ['-date_joined']
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Filtrar usuarios por organización del usuario actual"""
        user = self.request.user
        
        # Todos los usuarios (incluidos superusuarios) solo ven usuarios de su organización
        if user.organization:
            return User.objects.filter(
                organization=user.organization
            ).select_related('organization').prefetch_related('historial_actividad')
        
        # Si no tiene organización asignada, solo ve su propio usuario
        return User.objects.filter(id=user.id)
    
    def get_serializer_class(self):
        """Seleccionar el serializer según la acción"""
        if self.action == 'list':
            return UserListSerializer
        elif self.action == 'retrieve':
            return UserDetailSerializer
        elif self.action == 'create':
            return UserCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return UserUpdateSerializer
        elif self.action == 'cambiar_contrasena':
            return CambiarContrasenaSerializer
        elif self.action == 'asignar_roles':
            return AsignarRolesSerializer
        return UserDetailSerializer
    
    def get_serializer_context(self):
        """Agregar contexto adicional al serializer"""
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    def perform_create(self, serializer):
        """Asignar organización al crear usuario"""
        serializer.save()
    
    def create(self, request, *args, **kwargs):
        """Crear nuevo usuario"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        
        # Registrar en historial
        HistorialUsuario.objects.create(
            usuario=user,
            accion='usuario_creado',
            descripcion=f'Usuario {user.username} creado por {request.user.username}',
            ip_address=self.get_client_ip(request),
            organization=user.organization
        )
        
        headers = self.get_success_headers(serializer.data)
        return Response(
            UserDetailSerializer(user, context={'request': request}).data,
            status=status.HTTP_201_CREATED,
            headers=headers
        )
    
    def update(self, request, *args, **kwargs):
        """Actualizar usuario completo"""
        response = super().update(request, *args, **kwargs)
        
        # Registrar en historial
        usuario = self.get_object()
        HistorialUsuario.objects.create(
            usuario=usuario,
            accion='usuario_actualizado',
            descripcion=f'Usuario {usuario.username} actualizado por {request.user.username}',
            ip_address=self.get_client_ip(request),
            organization=usuario.organization
        )
        
        return response
    
    def destroy(self, request, *args, **kwargs):
        """Eliminar usuario"""
        instance = self.get_object()
        
        # No permitir auto-eliminación
        if instance.id == request.user.id:
            return Response(
                {'error': 'No puedes eliminarte a ti mismo.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # No permitir eliminar usuarios de otra organización
        if not request.user.is_superuser and instance.organization != request.user.organization:
            return Response(
                {'error': 'No tienes permiso para eliminar este usuario.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Registrar antes de eliminar
        HistorialUsuario.objects.create(
            usuario=instance,
            accion='usuario_eliminado',
            descripcion=f'Usuario {instance.username} eliminado por {request.user.username}',
            ip_address=self.get_client_ip(request),
            organization=instance.organization
        )
        
        return super().destroy(request, *args, **kwargs)
    
    @action(detail=True, methods=['post'])
    def toggle_activo(self, request, pk=None):
        """Activar/desactivar usuario"""
        user = self.get_object()
        
        if user.id == request.user.id:
            return Response(
                {'error': 'No puedes modificar tu propio estado.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Obtener el valor del body o invertir el actual
        activo = request.data.get('activo')
        if activo is None:
            user.is_active = not user.is_active
        else:
            user.is_active = activo
        
        user.save()
        
        # Registrar en historial
        HistorialUsuario.objects.create(
            usuario=user,
            accion='estado_cambiado',
            descripcion=f'Usuario {user.username} {"activado" if user.is_active else "desactivado"} por {request.user.username}',
            ip_address=self.get_client_ip(request),
            organization=user.organization
        )
        
        return Response({
            'message': f'Usuario {user.username} {"activado" if user.is_active else "desactivado"} exitosamente.',
            'is_active': user.is_active
        })
    
    @action(detail=True, methods=['post'])
    def cambiar_contrasena(self, request, pk=None):
        """Cambiar contraseña de un usuario (admin)"""
        user = self.get_object()
        serializer = CambiarContrasenaSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Cambiar contraseña
        user.set_password(serializer.validated_data['new_password'])
        user.password_changed_at = timezone.now()
        user.save()
        
        # Registrar en historial
        HistorialUsuario.objects.create(
            usuario=user,
            accion='contrasena_cambiada',
            descripcion=f'Contraseña de {user.username} cambiada por admin {request.user.username}',
            ip_address=self.get_client_ip(request),
            organization=user.organization
        )
        
        return Response({
            'message': f'Contraseña cambiada exitosamente para {user.username}.'
        })
    
    @action(detail=True, methods=['post'])
    def asignar_roles(self, request, pk=None):
        """Asignar roles personalizados al usuario"""
        user = self.get_object()
        serializer = AsignarRolesSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        
        roles_ids = serializer.validated_data['roles_ids']
        
        # Desactivar asignaciones actuales
        AsignacionRol.objects.filter(usuario=user, activa=True).update(activa=False)
        
        # Crear nuevas asignaciones
        roles_asignados = []
        for rol_id in roles_ids:
            try:
                rol = Rol.objects.get(id=rol_id, organization=user.organization)
                AsignacionRol.objects.create(
                    usuario=user,
                    rol=rol,
                    asignado_por=request.user,
                    organization=user.organization
                )
                roles_asignados.append(rol.nombre)
            except Rol.DoesNotExist:
                pass
        
        # Registrar en historial
        roles_nombres = ', '.join(roles_asignados)
        HistorialUsuario.objects.create(
            usuario=user,
            accion='roles_asignados',
            descripcion=f'Roles asignados a {user.username}: {roles_nombres}',
            ip_address=self.get_client_ip(request),
            organization=user.organization
        )
        
        return Response({
            'message': 'Roles asignados exitosamente.',
            'roles': roles_asignados
        })
    
    @action(detail=True, methods=['get'])
    def get_permisos(self, request, pk=None):
        """Ver permisos del usuario"""
        user = self.get_object()
        serializer = UserDetailSerializer(user, context={'request': request})
        return Response({
            'usuario': user.username,
            'permisos': serializer.data['permisos']
        })
    
    @action(detail=True, methods=['get'])
    def get_historial_actividad(self, request, pk=None):
        """Ver historial de actividad del usuario"""
        user = self.get_object()
        
        # Parámetros de paginación
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        
        # Obtener historial
        historial = HistorialUsuario.objects.filter(
            usuario=user,
            organization=user.organization
        ).order_by('-fecha')
        total = historial.count()
        
        # Paginar
        start = (page - 1) * page_size
        end = start + page_size
        historial_pagina = historial[start:end]
        
        serializer = HistorialUsuarioSerializer(historial_pagina, many=True)
        
        return Response({
            'results': serializer.data,
            'count': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size
        })
    
    @action(detail=False, methods=['get'], url_path='estadisticas')
    def estadisticas(self, request):
        """Estadísticas generales de usuarios de la organización"""
        queryset = self.get_queryset()
        
        total_usuarios = queryset.count()
        usuarios_activos = queryset.filter(is_active=True).count()
        usuarios_inactivos = queryset.filter(is_active=False).count()
        administradores = queryset.filter(Q(is_staff=True) | Q(is_superuser=True)).count()
        
        stats = {
            'total_usuarios': total_usuarios,
            'usuarios_activos': usuarios_activos,
            'usuarios_inactivos': usuarios_inactivos,
            'administradores': administradores
        }
        
        serializer = UserStatsSerializer(stats)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def exportar(self, request):
        """Exportar usuarios a Excel"""
        # Aplicar filtros
        queryset = self.filter_queryset(self.get_queryset())
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Encabezados
        headers = ['ID', 'Usuario', 'Email', 'Nombre Completo', 'Teléfono', 'Organización', 'Activo', 'Staff', 'Fecha Registro']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for row, user in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=user.id)
            ws.cell(row=row, column=2, value=user.username)
            ws.cell(row=row, column=3, value=user.email)
            ws.cell(row=row, column=4, value=user.full_name or f"{user.first_name} {user.last_name}".strip())
            ws.cell(row=row, column=5, value=user.phone)
            ws.cell(row=row, column=6, value=user.organization.nombre if user.organization else 'N/A')
            ws.cell(row=row, column=7, value="Sí" if user.is_active else "No")
            ws.cell(row=row, column=8, value="Sí" if user.is_staff else "No")
            ws.cell(row=row, column=9, value=user.date_joined.strftime('%Y-%m-%d %H:%M'))
        
        # Ajustar anchos de columna
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        
        # Generar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'
        wb.save(response)
        
        return response
    
    @action(detail=False, methods=['post'])
    def verificar_username(self, request):
        """Verificar disponibilidad de username"""
        username = request.data.get('username')
        
        if not username:
            return Response(
                {'error': 'Username es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        disponible = not User.objects.filter(username=username).exists()
        
        serializer = VerificarDisponibilidadSerializer({
            'disponible': disponible,
            'mensaje': 'Username disponible' if disponible else 'Username ya está en uso'
        })
        
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def verificar_email(self, request):
        """Verificar disponibilidad de email"""
        email = request.data.get('email')
        
        if not email:
            return Response(
                {'error': 'Email es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        disponible = not User.objects.filter(email=email).exists()
        
        serializer = VerificarDisponibilidadSerializer({
            'disponible': disponible,
            'mensaje': 'Email disponible' if disponible else 'Email ya está registrado'
        })
        
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def resetear_contrasena(self, request):
        """Enviar email de reseteo de contraseña (placeholder)"""
        email = request.data.get('email')
        
        if not email:
            return Response(
                {'error': 'Email es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            user = User.objects.get(email=email)
            
            # TODO: Implementar envío de email de reseteo
            # Por ahora solo retornamos éxito
            
            return Response({
                'message': f'Email de reseteo enviado a {email}'
            })
        except User.DoesNotExist:
            # Por seguridad, no revelar que el email no existe
            return Response({
                'message': f'Si el email existe, recibirás instrucciones para resetear tu contraseña'
            })
    
    def get_client_ip(self, request):
        """Obtener IP del cliente"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip
