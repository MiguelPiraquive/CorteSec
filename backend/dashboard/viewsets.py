import io
import json
import math
from datetime import timedelta, date

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.utils.encoding import escape_uri_path
from .policies import DashboardAccessPolicy
from django.db.models import Q, Sum, Count, Avg, F, Value, CharField
from django.db.models.functions import TruncMonth, Coalesce
from django.utils import timezone
from decimal import Decimal
from core.pagination import StandardResultsSetPagination
from .models import Project, AsignacionProyecto, ActiveProject
from .serializers import (
    ProjectSerializer, ProjectSummarySerializer,
    AsignacionProyectoSerializer, ActiveProjectSerializer,
)


def _get_request_org(request):
    return (
        getattr(request, 'tenant', None)
        or getattr(request.user, 'organization', None)
        or getattr(request.user, 'organizacion', None)
    )


def _filter_by_org(queryset, org):
    if not org:
        return queryset.none()
    model = queryset.model
    if hasattr(model, 'organization'):
        return queryset.filter(organization=org)
    if hasattr(model, 'organizacion'):
        return queryset.filter(organizacion=org)
    return queryset


class ProjectViewSet(viewsets.ModelViewSet):
    """
    ViewSet completo de Proyectos con filtros avanzados, estadísticas,
    asignaciones de empleados y gestión de proyecto activo.
    """
    serializer_class = ProjectSerializer
    permission_classes = [DashboardAccessPolicy]
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        org = _get_request_org(self.request)
        queryset = _filter_by_org(Project.objects.all(), org)

        # Filtros
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search)
                | Q(description__icontains=search)
                | Q(codigo_proyecto__icontains=search)
                | Q(cliente__icontains=search)
            )

        estado = self.request.query_params.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)

        prioridad = self.request.query_params.get('prioridad')
        if prioridad:
            queryset = queryset.filter(prioridad=prioridad)

        responsable = self.request.query_params.get('responsable')
        if responsable:
            queryset = queryset.filter(responsable_id=responsable)

        return queryset.order_by('-start_date')

    def perform_create(self, serializer):
        org = _get_request_org(self.request)
        user = self.request.user
        # Si no se especifica responsable, asignar al usuario que crea
        responsable = serializer.validated_data.get('responsable')
        serializer.save(
            organization=org,
            created_by=user,
            responsable=responsable or user,
        )

    # ── Estadísticas globales ───────────────────────────────────────
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """KPIs globales de todos los proyectos."""
        qs = self.get_queryset()
        total = qs.count()
        by_estado = {}
        for estado, label in Project.ESTADO_CHOICES:
            by_estado[estado] = qs.filter(estado=estado).count()

        agg = qs.aggregate(
            presupuesto_total=Sum('presupuesto_aprobado'),
            progreso_promedio=Avg('progreso'),
        )

        return Response({
            'total': total,
            'by_estado': by_estado,
            'presupuesto_total': str(agg['presupuesto_total'] or Decimal('0.00')),
            'progreso_promedio': round(agg['progreso_promedio'] or 0, 1),
        })

    # ── Datos agrupados para Kanban ─────────────────────────────────
    @action(detail=False, methods=['get'])
    def kanban(self, request):
        """Proyectos agrupados por estado para vista kanban."""
        qs = self.get_queryset()
        data = {}
        for estado, label in Project.ESTADO_CHOICES:
            items = qs.filter(estado=estado).order_by('prioridad', '-start_date')
            data[estado] = {
                'label': label,
                'count': items.count(),
                'projects': ProjectSummarySerializer(items, many=True).data,
            }
        return Response(data)

    # ── Cambiar estado rápido (para kanban drag & drop) ─────────────
    @action(detail=True, methods=['patch'], url_path='cambiar-estado')
    def cambiar_estado(self, request, pk=None):
        """Cambiar solo el estado del proyecto."""
        project = self.get_object()
        nuevo_estado = request.data.get('estado')
        if nuevo_estado not in dict(Project.ESTADO_CHOICES):
            return Response(
                {'error': f'Estado inválido: {nuevo_estado}'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        project.estado = nuevo_estado
        project.save()
        return Response(ProjectSerializer(project).data)

    # ── KPIs de un proyecto específico ──────────────────────────────
    @action(detail=True, methods=['get'])
    def kpis(self, request, pk=None):
        """KPIs detallados de un proyecto."""
        project = self.get_object()
        from nomina.models import NominaSimple
        from prestamos.models import Prestamo

        nominas_pagadas = NominaSimple.objects.filter(
            proyecto=project, estado='pagada'
        ).count()
        total_nomina = NominaSimple.objects.filter(
            proyecto=project, estado='pagada'
        ).aggregate(t=Sum('total_pagar'))['t'] or Decimal('0.00')

        prestamos_activos = Prestamo.objects.filter(
            proyecto=project, estado='activo'
        ).count()

        return Response({
            'gasto_acumulado': str(project.gasto_acumulado),
            'presupuesto_restante': str(project.presupuesto_restante),
            'porcentaje_ejecucion': project.porcentaje_ejecucion,
            'empleados_asignados': project.empleados_count,
            'nominas_pagadas': nominas_pagadas,
            'total_nomina': str(total_nomina),
            'prestamos_activos': prestamos_activos,
        })

    # ── Asignaciones de empleados al proyecto ───────────────────────
    @action(detail=True, methods=['get', 'post'], url_path='asignaciones')
    def asignaciones(self, request, pk=None):
        """Listar o crear asignaciones de empleados al proyecto."""
        project = self.get_object()

        if request.method == 'GET':
            asignaciones = project.asignaciones.select_related('empleado').all()
            serializer = AsignacionProyectoSerializer(asignaciones, many=True)
            return Response(serializer.data)

        # POST — crear asignación
        data = request.data.copy()
        data['proyecto'] = project.id
        org = _get_request_org(request)
        serializer = AsignacionProyectoSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save(organization=org)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['delete'], url_path='asignaciones/(?P<asignacion_id>[^/.]+)')
    def desasignar(self, request, pk=None, asignacion_id=None):
        """Eliminar una asignación de empleado."""
        project = self.get_object()
        try:
            asignacion = project.asignaciones.get(id=asignacion_id)
            asignacion.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except AsignacionProyecto.DoesNotExist:
            return Response({'error': 'Asignación no encontrada'}, status=status.HTTP_404_NOT_FOUND)

    # ── Listado resumido (para dropdowns/selectores) ────────────────
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Lista resumida de proyectos para selectores."""
        qs = self.get_queryset().exclude(estado='cancelado')
        serializer = ProjectSummarySerializer(qs, many=True)
        return Response(serializer.data)

    # ── Timeline — datos para vista Gantt ───────────────────────────
    @action(detail=False, methods=['get'])
    def timeline(self, request):
        """Proyectos con fechas para vista de línea temporal."""
        qs = self.get_queryset().order_by('start_date')
        data = []
        for p in qs:
            data.append({
                'id': p.id,
                'name': p.name,
                'codigo_proyecto': p.codigo_proyecto,
                'estado': p.estado,
                'prioridad': p.prioridad,
                'progreso': p.progreso,
                'color': p.color or '#6366f1',
                'start_date': str(p.start_date) if p.start_date else None,
                'end_date': str(p.end_date) if p.end_date else None,
                'fecha_real_fin': str(p.fecha_real_fin) if p.fecha_real_fin else None,
                'responsable_nombre': (
                    p.responsable.get_full_name() or p.responsable.username
                ) if p.responsable else None,
                'empleados_count': p.empleados_count,
                'presupuesto_aprobado': str(p.presupuesto_aprobado or 0),
                'cliente': p.cliente,
            })
        return Response(data)

    # ── Comparativa multi-proyecto ──────────────────────────────────
    @action(detail=False, methods=['get'])
    def comparativa(self, request):
        """Datos comparativos de todos los proyectos para reportes."""
        from nomina.models import NominaSimple
        from prestamos.models import Prestamo
        from contabilidad.models import FlujoCaja

        qs = self.get_queryset().exclude(estado='cancelado')
        data = []
        for p in qs:
            nominas_total = NominaSimple.objects.filter(
                proyecto=p, estado='pagada'
            ).aggregate(t=Sum('total_pagar'))['t'] or Decimal('0.00')

            nominas_count = NominaSimple.objects.filter(
                proyecto=p, estado='pagada'
            ).count()

            prestamos_count = Prestamo.objects.filter(
                proyecto=p, estado='activo'
            ).count()

            prestamos_total = Prestamo.objects.filter(
                proyecto=p
            ).aggregate(t=Sum('monto_aprobado'))['t'] or Decimal('0.00')

            flujo_ingresos = FlujoCaja.objects.filter(
                proyecto=p, tipo_movimiento='ingreso'
            ).aggregate(t=Sum('valor'))['t'] or Decimal('0.00')

            flujo_egresos = FlujoCaja.objects.filter(
                proyecto=p, tipo_movimiento='egreso'
            ).aggregate(t=Sum('valor'))['t'] or Decimal('0.00')

            data.append({
                'id': p.id,
                'name': p.name,
                'codigo_proyecto': p.codigo_proyecto,
                'estado': p.estado,
                'prioridad': p.prioridad,
                'progreso': p.progreso,
                'color': p.color or '#6366f1',
                'empleados_count': p.empleados_count,
                'presupuesto_estimado': str(p.presupuesto_estimado or 0),
                'presupuesto_aprobado': str(p.presupuesto_aprobado or 0),
                'gasto_acumulado': str(p.gasto_acumulado),
                'porcentaje_ejecucion': p.porcentaje_ejecucion,
                'nominas_pagadas': nominas_count,
                'total_nomina': str(nominas_total),
                'prestamos_activos': prestamos_count,
                'total_prestamos': str(prestamos_total),
                'flujo_ingresos': str(flujo_ingresos),
                'flujo_egresos': str(flujo_egresos),
                'flujo_neto': str(flujo_ingresos - flujo_egresos),
                'start_date': str(p.start_date) if p.start_date else None,
                'end_date': str(p.end_date) if p.end_date else None,
                'duration_days': p.duration_days,
                'responsable_nombre': (
                    p.responsable.get_full_name() or p.responsable.username
                ) if p.responsable else None,
            })
        return Response(data)

    # ════════════════════════════════════════════════════════════════
    #  FASE 3 — PLANTILLAS DE PROYECTO
    # ════════════════════════════════════════════════════════════════

    PLANTILLAS = [
        {
            'id': 'consultoria',
            'name': 'Consultoría Empresarial',
            'description': 'Proyecto de consultoría con fases de diagnóstico, análisis y entrega de resultados.',
            'icon': '🏢',
            'color': '#6366f1',
            'estado': 'planificacion',
            'prioridad': 'alta',
            'tags': ['consultoría', 'empresarial'],
            'duracion_dias': 90,
            'presupuesto_sugerido': 15000000,
            'moneda': 'COP',
            'equipo_sugerido': 4,
            'fases': ['Diagnóstico', 'Análisis', 'Implementación', 'Entrega'],
        },
        {
            'id': 'construccion',
            'name': 'Obra Civil / Construcción',
            'description': 'Proyecto de construcción o remodelación con control de presupuesto y equipos.',
            'icon': '🏗️',
            'color': '#f59e0b',
            'estado': 'planificacion',
            'prioridad': 'alta',
            'tags': ['construcción', 'obra'],
            'duracion_dias': 180,
            'presupuesto_sugerido': 50000000,
            'moneda': 'COP',
            'equipo_sugerido': 5,
            'fases': ['Diseño', 'Permisos', 'Cimentación', 'Estructura', 'Acabados', 'Entrega'],
        },
        {
            'id': 'desarrollo_software',
            'name': 'Desarrollo de Software',
            'description': 'Proyecto de desarrollo con sprints, QA y despliegue.',
            'icon': '💻',
            'color': '#10b981',
            'estado': 'planificacion',
            'prioridad': 'media',
            'tags': ['software', 'desarrollo', 'tech'],
            'duracion_dias': 120,
            'presupuesto_sugerido': 25000000,
            'moneda': 'COP',
            'equipo_sugerido': 5,
            'fases': ['Planificación', 'Sprint 1', 'Sprint 2', 'Sprint 3', 'QA', 'Deploy'],
        },
        {
            'id': 'evento',
            'name': 'Evento / Producción',
            'description': 'Organización de evento corporativo, conferencia o producción.',
            'icon': '🎪',
            'color': '#ec4899',
            'estado': 'planificacion',
            'prioridad': 'media',
            'tags': ['evento', 'producción'],
            'duracion_dias': 60,
            'presupuesto_sugerido': 8000000,
            'moneda': 'COP',
            'equipo_sugerido': 3,
            'fases': ['Pre-producción', 'Logística', 'Montaje', 'Evento', 'Desmontaje'],
        },
        {
            'id': 'nomina_rrhh',
            'name': 'Gestión de Nómina / RRHH',
            'description': 'Proyecto de gestión de recursos humanos, nómina y bienestar laboral.',
            'icon': '👥',
            'color': '#8b5cf6',
            'estado': 'activo',
            'prioridad': 'alta',
            'tags': ['nómina', 'rrhh', 'permanente'],
            'duracion_dias': 365,
            'presupuesto_sugerido': 100000000,
            'moneda': 'COP',
            'equipo_sugerido': 2,
            'fases': ['Inicio', 'Operación continua', 'Cierre fiscal'],
        },
        {
            'id': 'marketing',
            'name': 'Campaña de Marketing',
            'description': 'Campaña de marketing digital o tradicional con métricas de ROI.',
            'icon': '📢',
            'color': '#f97316',
            'estado': 'planificacion',
            'prioridad': 'media',
            'tags': ['marketing', 'campaña', 'digital'],
            'duracion_dias': 45,
            'presupuesto_sugerido': 5000000,
            'moneda': 'COP',
            'equipo_sugerido': 3,
            'fases': ['Estrategia', 'Creativo', 'Lanzamiento', 'Medición'],
        },
        {
            'id': 'investigacion',
            'name': 'Investigación / Estudio',
            'description': 'Proyecto de investigación, estudios de mercado o análisis técnico.',
            'icon': '🔬',
            'color': '#06b6d4',
            'estado': 'planificacion',
            'prioridad': 'media',
            'tags': ['investigación', 'estudio', 'análisis'],
            'duracion_dias': 90,
            'presupuesto_sugerido': 10000000,
            'moneda': 'COP',
            'equipo_sugerido': 3,
            'fases': ['Revisión literaria', 'Recolección', 'Análisis', 'Informe'],
        },
        {
            'id': 'en_blanco',
            'name': 'Proyecto en Blanco',
            'description': 'Empieza desde cero con un proyecto completamente personalizable.',
            'icon': '📋',
            'color': '#64748b',
            'estado': 'planificacion',
            'prioridad': 'media',
            'tags': [],
            'duracion_dias': 90,
            'presupuesto_sugerido': 0,
            'moneda': 'COP',
            'equipo_sugerido': 1,
            'fases': [],
        },
    ]

    @action(detail=False, methods=['get'])
    def plantillas(self, request):
        """Devuelve las plantillas de proyecto disponibles."""
        return Response(self.PLANTILLAS)

    @action(detail=False, methods=['post'], url_path='crear-desde-plantilla')
    def crear_desde_plantilla(self, request):
        """Crear un proyecto a partir de una plantilla."""
        plantilla_id = request.data.get('plantilla_id')
        nombre = request.data.get('name')
        descripcion = request.data.get('description')

        plantilla = next((p for p in self.PLANTILLAS if p['id'] == plantilla_id), None)
        if not plantilla:
            return Response(
                {'error': f'Plantilla "{plantilla_id}" no encontrada'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        org = _get_request_org(request)
        user = request.user
        today = timezone.now().date()

        project = Project.objects.create(
            organization=org,
            name=nombre or plantilla['name'],
            description=descripcion or plantilla['description'],
            estado=plantilla['estado'],
            prioridad=plantilla['prioridad'],
            color=plantilla['color'],
            icono=plantilla['icon'],
            tags=plantilla['tags'],
            moneda=plantilla['moneda'],
            presupuesto_estimado=Decimal(str(plantilla['presupuesto_sugerido'])),
            start_date=today,
            end_date=today + timedelta(days=plantilla['duracion_dias']),
            responsable=user,
            created_by=user,
            notas_internas=f"Creado desde plantilla: {plantilla['name']}\nFases sugeridas: {', '.join(plantilla.get('fases', []))}",
        )

        return Response(ProjectSerializer(project).data, status=status.HTTP_201_CREATED)

    # ════════════════════════════════════════════════════════════════
    #  FASE 3 — EXPORTAR A EXCEL / PDF
    # ════════════════════════════════════════════════════════════════

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """Exportar todos los proyectos a Excel (.xlsx)."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        MAX_EXPORT_ROWS = 5000
        qs = self.get_queryset()[:MAX_EXPORT_ROWS]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Proyectos'

        # Estilos
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        headers = [
            'Código', 'Nombre', 'Estado', 'Prioridad', 'Progreso (%)',
            'Presupuesto Estimado', 'Presupuesto Aprobado', 'Gasto Acumulado',
            '% Ejecución', 'Moneda', 'Cliente', 'Responsable',
            'Empleados', 'Fecha Inicio', 'Fecha Fin', 'Centro Costo',
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Datos
        for row_idx, p in enumerate(qs, 2):
            responsable_name = ''
            if p.responsable:
                responsable_name = p.responsable.get_full_name() or p.responsable.username

            values = [
                p.codigo_proyecto,
                p.name,
                p.get_estado_display(),
                p.get_prioridad_display(),
                p.progreso,
                float(p.presupuesto_estimado or 0),
                float(p.presupuesto_aprobado or 0),
                float(p.gasto_acumulado),
                p.porcentaje_ejecucion,
                p.moneda,
                p.cliente,
                responsable_name,
                p.empleados_count,
                str(p.start_date) if p.start_date else '',
                str(p.end_date) if p.end_date else '',
                p.centro_costo,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border

        # Auto-width
        for col_idx in range(1, len(headers) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ''))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)

        # Hoja 2: Resumen financiero
        ws2 = wb.create_sheet('Resumen Financiero')
        summary_headers = ['Métrica', 'Valor']
        for col, h in enumerate(summary_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        total_pres = sum(float(p.presupuesto_aprobado or p.presupuesto_estimado or 0) for p in qs)
        total_gasto = sum(float(p.gasto_acumulado) for p in qs)
        total_empleados = sum(p.empleados_count for p in qs)
        promedio_progreso = sum(p.progreso for p in qs) / max(qs.count(), 1)

        metrics = [
            ('Total Proyectos', qs.count()),
            ('Activos', qs.filter(estado='activo').count()),
            ('Completados', qs.filter(estado='completado').count()),
            ('Presupuesto Total', total_pres),
            ('Gasto Total', total_gasto),
            ('Diferencia', total_pres - total_gasto),
            ('Total Empleados Asignados', total_empleados),
            ('Progreso Promedio (%)', round(promedio_progreso, 1)),
        ]
        for r, (label, val) in enumerate(metrics, 2):
            ws2.cell(row=r, column=1, value=label)
            ws2.cell(row=r, column=2, value=val)

        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        safe_name = f"proyectos_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{escape_uri_path(safe_name)}"
        return response

    @action(detail=True, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request, pk=None):
        """Exportar reporte PDF de un proyecto específico."""
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        project = self.get_object()
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5 * inch)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Title'],
            fontSize=20, spaceAfter=20, textColor=colors.HexColor('#4F46E5'),
        )
        heading_style = ParagraphStyle(
            'CustomHeading', parent=styles['Heading2'],
            fontSize=14, spaceAfter=10, textColor=colors.HexColor('#1e1b4b'),
        )

        elements = []

        # Título
        elements.append(Paragraph(f'{project.codigo_proyecto} — {project.name}', title_style))
        elements.append(Spacer(1, 10))

        # Info general
        elements.append(Paragraph('Información General', heading_style))
        responsable_name = ''
        if project.responsable:
            responsable_name = project.responsable.get_full_name() or project.responsable.username

        info_data = [
            ['Estado', project.get_estado_display()],
            ['Prioridad', project.get_prioridad_display()],
            ['Progreso', f'{project.progreso}%'],
            ['Responsable', responsable_name or 'Sin asignar'],
            ['Cliente', project.cliente or '—'],
            ['Centro Costo', project.centro_costo or '—'],
            ['Fecha Inicio', str(project.start_date)],
            ['Fecha Fin', str(project.end_date) if project.end_date else '—'],
            ['Duración', f'{project.duration_days} días'],
        ]
        info_table = Table(info_data, colWidths=[2.5 * inch, 4 * inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#EEF2FF')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#C7D2FE')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 20))

        # Finanzas
        elements.append(Paragraph('Información Financiera', heading_style))
        finance_data = [
            ['Presupuesto Estimado', f'${float(project.presupuesto_estimado):,.0f} {project.moneda}'],
            ['Presupuesto Aprobado', f'${float(project.presupuesto_aprobado):,.0f} {project.moneda}'],
            ['Gasto Acumulado', f'${float(project.gasto_acumulado):,.0f} {project.moneda}'],
            ['Presupuesto Restante', f'${float(project.presupuesto_restante):,.0f} {project.moneda}'],
            ['% Ejecución', f'{project.porcentaje_ejecucion}%'],
        ]
        fin_table = Table(finance_data, colWidths=[2.5 * inch, 4 * inch])
        fin_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#FEF3C7')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#FDE68A')),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(fin_table)
        elements.append(Spacer(1, 20))

        # Equipo
        elements.append(Paragraph('Equipo Asignado', heading_style))
        asignaciones = project.asignaciones.filter(activo=True).select_related('empleado')
        if asignaciones.exists():
            team_data = [['Empleado', 'Desde']]
            for a in asignaciones:
                team_data.append([
                    getattr(a.empleado, 'nombre_completo', str(a.empleado)),
                    str(a.fecha_asignacion),
                ])
            team_table = Table(team_data, colWidths=[4 * inch, 2.5 * inch])
            team_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#C7D2FE')),
                ('PADDING', (0, 0), (-1, -1), 6),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F3FF')]),
            ]))
            elements.append(team_table)
        else:
            elements.append(Paragraph('Sin empleados asignados.', styles['Normal']))

        elements.append(Spacer(1, 20))

        # Footer
        elements.append(Paragraph(
            f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")} — CorteSec',
            ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray),
        ))

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.read(), content_type='application/pdf')
        safe_name = f"proyecto_{project.codigo_proyecto}_{timezone.now().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{escape_uri_path(safe_name)}"
        return response

    # ════════════════════════════════════════════════════════════════
    #  FASE 3 — IA PREDICTIVA POR PROYECTO
    # ════════════════════════════════════════════════════════════════

    @action(detail=True, methods=['get'])
    def predicciones(self, request, pk=None):
        """
        Análisis predictivo del proyecto basado en datos históricos.
        Calcula: fecha estimada real de finalización, riesgo de retraso,
        proyección de gasto, velocidad de progreso, salud general.
        """
        project = self.get_object()
        from nomina.models import NominaSimple
        from contabilidad.models import FlujoCaja

        today = timezone.now().date()
        start = project.start_date
        end_planned = project.end_date
        progreso = project.progreso or 0

        # ── Velocidad de progreso ────────────────────────────────
        dias_transcurridos = max((today - start).days, 1)
        velocidad_diaria = progreso / dias_transcurridos  # % por día

        # ── Fecha estimada real de finalización ──────────────────
        if progreso >= 100:
            fecha_estimada_fin = project.fecha_real_fin or today
            dias_restantes_estimados = 0
        elif velocidad_diaria > 0:
            dias_restantes_estimados = math.ceil((100 - progreso) / velocidad_diaria)
            fecha_estimada_fin = today + timedelta(days=dias_restantes_estimados)
        else:
            dias_restantes_estimados = None
            fecha_estimada_fin = None

        # ── Riesgo de retraso ────────────────────────────────────
        if end_planned:
            dias_plan_total = max((end_planned - start).days, 1)
            progreso_esperado = min(100, (dias_transcurridos / dias_plan_total) * 100)
            desviacion_progreso = progreso_esperado - progreso

            if fecha_estimada_fin and fecha_estimada_fin > end_planned:
                dias_retraso = (fecha_estimada_fin - end_planned).days
            else:
                dias_retraso = 0

            # Score de riesgo 0-100
            riesgo_score = min(100, max(0, int(
                (desviacion_progreso * 0.6) + (dias_retraso * 0.4)
            )))
        else:
            progreso_esperado = 0
            desviacion_progreso = 0
            dias_retraso = 0
            riesgo_score = 20  # Sin fecha fin = riesgo bajo por defecto

        if riesgo_score >= 70:
            riesgo_nivel = 'critico'
        elif riesgo_score >= 40:
            riesgo_nivel = 'alto'
        elif riesgo_score >= 20:
            riesgo_nivel = 'moderado'
        else:
            riesgo_nivel = 'bajo'

        # ── Proyección de gasto ──────────────────────────────────
        gasto_actual = float(project.gasto_acumulado)
        presupuesto = float(project.presupuesto_aprobado or project.presupuesto_estimado or 0)

        if progreso > 0:
            gasto_por_porcentaje = gasto_actual / progreso
            gasto_proyectado_100 = gasto_por_porcentaje * 100
        else:
            gasto_proyectado_100 = 0

        if presupuesto > 0:
            desviacion_presupuesto = ((gasto_proyectado_100 - presupuesto) / presupuesto) * 100
        else:
            desviacion_presupuesto = 0

        # ── Tendencia de gasto mensual ───────────────────────────
        egresos_mensuales = FlujoCaja.objects.filter(
            proyecto=project, tipo_movimiento='egreso',
        ).values('fecha__month', 'fecha__year').annotate(
            total=Sum('valor'),
        ).order_by('fecha__year', 'fecha__month')

        nominas_mensuales = NominaSimple.objects.filter(
            proyecto=project, estado='pagada',
        ).values('fecha_pago__month', 'fecha_pago__year').annotate(
            total=Sum('total_pagar'),
        ).order_by('fecha_pago__year', 'fecha_pago__month')

        tendencia_gastos = []
        for e in egresos_mensuales:
            tendencia_gastos.append({
                'mes': f"{e['fecha__year']}-{e['fecha__month']:02d}",
                'monto': float(e['total']),
                'tipo': 'egreso',
            })
        for n in nominas_mensuales:
            tendencia_gastos.append({
                'mes': f"{n['fecha_pago__year']}-{n['fecha_pago__month']:02d}",
                'monto': float(n['total']),
                'tipo': 'nomina',
            })

        # ── Salud general del proyecto ───────────────────────────
        # Score 0-100 basado en: progreso vs plan, presupuesto, velocidad
        score_components = []

        # Componente progreso (40%)
        if end_planned:
            if desviacion_progreso <= 0:
                score_components.append(40)
            elif desviacion_progreso < 20:
                score_components.append(30)
            elif desviacion_progreso < 40:
                score_components.append(15)
            else:
                score_components.append(5)
        else:
            score_components.append(30)

        # Componente presupuesto (35%)
        if presupuesto > 0:
            ejecucion_pct = (gasto_actual / presupuesto) * 100
            if ejecucion_pct <= 80:
                score_components.append(35)
            elif ejecucion_pct <= 100:
                score_components.append(25)
            elif ejecucion_pct <= 120:
                score_components.append(10)
            else:
                score_components.append(0)
        else:
            score_components.append(25)

        # Componente velocidad (25%)
        if velocidad_diaria > 0:
            if velocidad_diaria >= 1:
                score_components.append(25)
            elif velocidad_diaria >= 0.5:
                score_components.append(20)
            elif velocidad_diaria >= 0.2:
                score_components.append(10)
            else:
                score_components.append(5)
        else:
            score_components.append(0)

        salud_score = sum(score_components)
        if salud_score >= 80:
            salud_nivel = 'excelente'
        elif salud_score >= 60:
            salud_nivel = 'bueno'
        elif salud_score >= 40:
            salud_nivel = 'regular'
        else:
            salud_nivel = 'critico'

        # ── Recomendaciones automáticas ──────────────────────────
        recomendaciones = []
        if riesgo_nivel in ('critico', 'alto'):
            recomendaciones.append({
                'tipo': 'alerta',
                'icono': '⚠️',
                'titulo': 'Riesgo de retraso',
                'detalle': f'El proyecto lleva {desviacion_progreso:.0f}% de retraso vs. lo planificado. Considere reasignar recursos.',
            })
        if desviacion_presupuesto > 20:
            recomendaciones.append({
                'tipo': 'alerta',
                'icono': '💰',
                'titulo': 'Sobrecosto proyectado',
                'detalle': f'El gasto proyectado supera el presupuesto en {desviacion_presupuesto:.0f}%. Revise la ejecución.',
            })
        if velocidad_diaria < 0.1 and progreso < 100:
            recomendaciones.append({
                'tipo': 'warning',
                'icono': '🐌',
                'titulo': 'Velocidad muy baja',
                'detalle': f'Solo {velocidad_diaria:.2f}%/día. A este ritmo, el proyecto tardaría {dias_restantes_estimados or "∞"} días más.',
            })
        if project.empleados_count == 0:
            recomendaciones.append({
                'tipo': 'info',
                'icono': '👥',
                'titulo': 'Sin equipo asignado',
                'detalle': 'Asigne empleados al proyecto para mejor seguimiento.',
            })
        if not project.end_date:
            recomendaciones.append({
                'tipo': 'info',
                'icono': '📅',
                'titulo': 'Sin fecha fin definida',
                'detalle': 'Defina una fecha de fin para habilitar análisis de riesgo completo.',
            })
        if salud_nivel == 'excelente' and progreso > 0:
            recomendaciones.append({
                'tipo': 'success',
                'icono': '🎉',
                'titulo': 'Proyecto saludable',
                'detalle': '¡El proyecto va muy bien! Mantenga el ritmo actual.',
            })

        return Response({
            'progreso_actual': progreso,
            'dias_transcurridos': dias_transcurridos,
            'velocidad_diaria': round(velocidad_diaria, 3),
            'velocidad_semanal': round(velocidad_diaria * 7, 1),

            'fecha_estimada_fin': str(fecha_estimada_fin) if fecha_estimada_fin else None,
            'dias_restantes_estimados': dias_restantes_estimados,

            'progreso_esperado': round(progreso_esperado, 1),
            'desviacion_progreso': round(desviacion_progreso, 1),
            'dias_retraso': dias_retraso,
            'riesgo_score': riesgo_score,
            'riesgo_nivel': riesgo_nivel,

            'gasto_actual': gasto_actual,
            'presupuesto': presupuesto,
            'gasto_proyectado_100': round(gasto_proyectado_100, 0),
            'desviacion_presupuesto': round(desviacion_presupuesto, 1),

            'salud_score': salud_score,
            'salud_nivel': salud_nivel,

            'tendencia_gastos': tendencia_gastos,
            'recomendaciones': recomendaciones,
        })

    # ════════════════════════════════════════════════════════════════
    #  FASE 3 — GAMIFICACIÓN / LOGROS
    # ════════════════════════════════════════════════════════════════

    @action(detail=False, methods=['get'])
    def logros(self, request):
        """
        Calcula logros/achievements del usuario y la organización
        basado en métricas de proyectos, equipo y finanzas.
        """
        qs = self.get_queryset()
        user = request.user

        total_proyectos = qs.count()
        completados = qs.filter(estado='completado').count()
        activos = qs.filter(estado='activo').count()
        total_empleados = sum(p.empleados_count for p in qs)
        presupuesto_total = sum(float(p.presupuesto_aprobado or p.presupuesto_estimado or 0) for p in qs)
        progreso_promedio = (sum(p.progreso for p in qs) / max(total_proyectos, 1))

        # Para logros de presupuesto perfecto
        proyectos_bajo_presupuesto = 0
        for p in qs.filter(estado='completado'):
            aprobado = float(p.presupuesto_aprobado or p.presupuesto_estimado or 0)
            if aprobado > 0 and float(p.gasto_acumulado) <= aprobado:
                proyectos_bajo_presupuesto += 1

        # Para logros de velocidad
        proyectos_a_tiempo = 0
        for p in qs.filter(estado='completado', fecha_real_fin__isnull=False, end_date__isnull=False):
            if p.fecha_real_fin <= p.end_date:
                proyectos_a_tiempo += 1

        # ── Definición de logros ─────────────────────────────────
        logros_def = [
            # ── Progreso ──
            {
                'id': 'primer_proyecto',
                'nombre': 'Primer Paso',
                'descripcion': 'Crear tu primer proyecto',
                'icono': '🚀',
                'categoria': 'inicio',
                'condicion': total_proyectos >= 1,
                'progreso': min(total_proyectos, 1),
                'meta': 1,
                'xp': 50,
            },
            {
                'id': 'cinco_proyectos',
                'nombre': 'Gestor Activo',
                'descripcion': 'Tener 5 proyectos creados',
                'icono': '📊',
                'categoria': 'produccion',
                'condicion': total_proyectos >= 5,
                'progreso': min(total_proyectos, 5),
                'meta': 5,
                'xp': 150,
            },
            {
                'id': 'diez_proyectos',
                'nombre': 'Director de Portafolio',
                'descripcion': 'Gestionar 10+ proyectos',
                'icono': '👔',
                'categoria': 'produccion',
                'condicion': total_proyectos >= 10,
                'progreso': min(total_proyectos, 10),
                'meta': 10,
                'xp': 300,
            },
            # ── Completados ──
            {
                'id': 'primer_completado',
                'nombre': 'Misión Cumplida',
                'descripcion': 'Completar tu primer proyecto',
                'icono': '✅',
                'categoria': 'logro',
                'condicion': completados >= 1,
                'progreso': min(completados, 1),
                'meta': 1,
                'xp': 100,
            },
            {
                'id': 'cinco_completados',
                'nombre': 'Ejecutor Experto',
                'descripcion': 'Completar 5 proyectos',
                'icono': '🏆',
                'categoria': 'logro',
                'condicion': completados >= 5,
                'progreso': min(completados, 5),
                'meta': 5,
                'xp': 500,
            },
            # ── Equipo ──
            {
                'id': 'primer_equipo',
                'nombre': 'Constructor de Equipos',
                'descripcion': 'Asignar al menos 3 empleados a proyectos',
                'icono': '👥',
                'categoria': 'equipo',
                'condicion': total_empleados >= 3,
                'progreso': min(total_empleados, 3),
                'meta': 3,
                'xp': 100,
            },
            {
                'id': 'equipo_grande',
                'nombre': 'Líder de Equipo',
                'descripcion': 'Tener 10+ empleados asignados',
                'icono': '🫂',
                'categoria': 'equipo',
                'condicion': total_empleados >= 10,
                'progreso': min(total_empleados, 10),
                'meta': 10,
                'xp': 250,
            },
            # ── Finanzas ──
            {
                'id': 'presupuesto_millon',
                'nombre': 'Primer Millón',
                'descripcion': 'Gestionar $1M+ en presupuestos',
                'icono': '💰',
                'categoria': 'finanzas',
                'condicion': presupuesto_total >= 1_000_000,
                'progreso': min(presupuesto_total, 1_000_000),
                'meta': 1_000_000,
                'xp': 200,
            },
            {
                'id': 'bajo_presupuesto',
                'nombre': 'Presupuesto Perfecto',
                'descripcion': 'Completar un proyecto sin exceder el presupuesto',
                'icono': '🎯',
                'categoria': 'finanzas',
                'condicion': proyectos_bajo_presupuesto >= 1,
                'progreso': min(proyectos_bajo_presupuesto, 1),
                'meta': 1,
                'xp': 300,
            },
            {
                'id': 'tres_bajo_presupuesto',
                'nombre': 'Control Total',
                'descripcion': 'Completar 3 proyectos sin exceder presupuesto',
                'icono': '🏅',
                'categoria': 'finanzas',
                'condicion': proyectos_bajo_presupuesto >= 3,
                'progreso': min(proyectos_bajo_presupuesto, 3),
                'meta': 3,
                'xp': 750,
            },
            # ── Velocidad ──
            {
                'id': 'a_tiempo',
                'nombre': 'Puntualidad',
                'descripcion': 'Completar un proyecto antes de la fecha fin',
                'icono': '⏱️',
                'categoria': 'velocidad',
                'condicion': proyectos_a_tiempo >= 1,
                'progreso': min(proyectos_a_tiempo, 1),
                'meta': 1,
                'xp': 200,
            },
            {
                'id': 'tres_a_tiempo',
                'nombre': 'Maestro del Tiempo',
                'descripcion': 'Completar 3 proyectos a tiempo',
                'icono': '⌛',
                'categoria': 'velocidad',
                'condicion': proyectos_a_tiempo >= 3,
                'progreso': min(proyectos_a_tiempo, 3),
                'meta': 3,
                'xp': 600,
            },
            # ── Especiales ──
            {
                'id': 'progreso_promedio_alto',
                'nombre': 'Alta Performance',
                'descripcion': 'Mantener progreso promedio >= 70% en todos los proyectos',
                'icono': '🔥',
                'categoria': 'especial',
                'condicion': progreso_promedio >= 70 and total_proyectos >= 3,
                'progreso': min(round(progreso_promedio), 70),
                'meta': 70,
                'xp': 400,
            },
            {
                'id': 'diversificado',
                'nombre': 'Diversificado',
                'descripcion': 'Tener proyectos en 3+ estados diferentes',
                'icono': '🌈',
                'categoria': 'especial',
                'condicion': len(set(p.estado for p in qs)) >= 3,
                'progreso': len(set(p.estado for p in qs)),
                'meta': 3,
                'xp': 150,
            },
        ]

        # Calcular totales
        desbloqueados = [l for l in logros_def if l['condicion']]
        xp_total = sum(l['xp'] for l in desbloqueados)

        # Nivel basado en XP
        niveles = [
            (0, 'Principiante', '🌱'),
            (200, 'Aprendiz', '📗'),
            (500, 'Profesional', '⭐'),
            (1000, 'Experto', '💎'),
            (2000, 'Maestro', '👑'),
            (3500, 'Leyenda', '🏛️'),
        ]
        nivel_actual = niveles[0]
        nivel_siguiente = niveles[1] if len(niveles) > 1 else None
        for i, (xp_req, nombre, icono) in enumerate(niveles):
            if xp_total >= xp_req:
                nivel_actual = (xp_req, nombre, icono)
                nivel_siguiente = niveles[i + 1] if i + 1 < len(niveles) else None

        return Response({
            'xp_total': xp_total,
            'nivel': {
                'nombre': nivel_actual[1],
                'icono': nivel_actual[2],
                'xp_minimo': nivel_actual[0],
            },
            'nivel_siguiente': {
                'nombre': nivel_siguiente[1],
                'icono': nivel_siguiente[2],
                'xp_requerido': nivel_siguiente[0],
                'xp_faltante': nivel_siguiente[0] - xp_total,
            } if nivel_siguiente else None,
            'logros': [
                {
                    'id': l['id'],
                    'nombre': l['nombre'],
                    'descripcion': l['descripcion'],
                    'icono': l['icono'],
                    'categoria': l['categoria'],
                    'desbloqueado': l['condicion'],
                    'progreso': l['progreso'],
                    'meta': l['meta'],
                    'porcentaje': min(100, round((l['progreso'] / max(l['meta'], 1)) * 100)),
                    'xp': l['xp'],
                }
                for l in logros_def
            ],
            'resumen': {
                'total_logros': len(logros_def),
                'desbloqueados': len(desbloqueados),
                'porcentaje': round((len(desbloqueados) / max(len(logros_def), 1)) * 100),
            },
        })


class ActiveProjectViewSet(viewsets.ViewSet):
    """
    Gestión del proyecto activo del usuario.
    
    GET  /api/dashboard/active-project/  → devuelve proyecto activo actual
    POST /api/dashboard/active-project/  → establece proyecto activo
    DELETE /api/dashboard/active-project/ → limpia selección (modo "todos")
    """
    permission_classes = [DashboardAccessPolicy]

    def list(self, request):
        """Obtener proyecto activo del usuario."""
        try:
            ap = ActiveProject.objects.select_related('project').get(user=request.user)
            return Response(ActiveProjectSerializer(ap).data)
        except ActiveProject.DoesNotExist:
            # Sin selección = modo "todos"
            return Response({
                'mode': 'all',
                'project': None,
                'project_detail': None,
            })

    def create(self, request):
        """Establecer proyecto activo."""
        mode = request.data.get('mode', 'single')
        project_id = request.data.get('project_id')

        if mode == 'all':
            # Modo "todos" — limpiar selección
            ActiveProject.objects.update_or_create(
                user=request.user,
                defaults={'mode': 'all', 'project': None},
            )
            return Response({'mode': 'all', 'project': None})

        if not project_id:
            return Response(
                {'error': 'project_id es requerido para modo single'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validar que el proyecto pertenece a la org del usuario
        org = _get_request_org(request)
        try:
            project = _filter_by_org(Project.objects.all(), org).get(pk=project_id)
        except Project.DoesNotExist:
            return Response(
                {'error': 'Proyecto no encontrado o no pertenece a tu organización'},
                status=status.HTTP_404_NOT_FOUND,
            )

        ap, _ = ActiveProject.objects.update_or_create(
            user=request.user,
            defaults={'mode': 'single', 'project': project},
        )
        return Response(ActiveProjectSerializer(ap).data)

    def destroy(self, request, pk=None):
        """Limpiar proyecto activo (volver a modo todos)."""
        ActiveProject.objects.filter(user=request.user).delete()
        return Response({'mode': 'all', 'project': None})

    @action(detail=False, methods=['delete'])
    def clear(self, request):
        """Limpiar proyecto activo (volver a modo todos) — endpoint sin PK."""
        ActiveProject.objects.filter(user=request.user).delete()
        return Response({'mode': 'all', 'project': None})
