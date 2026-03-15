"""
╔══════════════════════════════════════════════════════════════════════════════╗
║                    VIEWS DE NÓMINA - CORTESEC                                 ║
║                Sistema de Nómina para Construcción                            ║
╚══════════════════════════════════════════════════════════════════════════════╝

ViewSets DRF con CRUD completo para todos los modelos de nómina.
Incluye endpoint de cálculo automático.

Autor: Sistema CorteSec
Versión: 1.0.0
Fecha: Enero 2026
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from drf_spectacular.utils import extend_schema, OpenApiParameter
from django.db.models.deletion import ProtectedError

from core.mixins import MultiTenantViewSetMixin

from .models import (
    Empleado,
    TipoContrato,
    Contrato,
    ParametroLegal,
    ConceptoLaboral,
    NominaSimple,
    NominaItem,
    NominaConcepto,
)
from .serializers import (
    EmpleadoListSerializer,
    EmpleadoDetailSerializer,
    EmpleadoCreateSerializer,
    TipoContratoSerializer,
    ContratoSerializer,
    ContratoCreateSerializer,
    ParametroLegalSerializer,
    ConceptoLaboralSerializer,
    NominaSimpleListSerializer,
    NominaSimpleDetailSerializer,
    NominaSimpleCreateSerializer,
    NominaItemSerializer,
    NominaItemCreateSerializer,
    NominaConceptoSerializer,
    CalculoNominaSerializer,
)
from .services import NominaValidationError
from .services import calcular_nomina
from .policies import (
    EmpleadoAccessPolicy,
    ContratoAccessPolicy,
    TipoContratoAccessPolicy,
    ParametrosLegalesAccessPolicy,
    ConceptoLaboralAccessPolicy,
    NominaAccessPolicy,
)


from core.utils import get_active_project_for_request as _get_active_project_for_request


# ══════════════════════════════════════════════════════════════════════════════
# VIEWSET: EMPLEADO
# ══════════════════════════════════════════════════════════════════════════════

@extend_schema(tags=['Nómina - Empleados'])
class EmpleadoViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para gestión de empleados.
    
    Endpoints:
    - GET /empleados/ - Listar empleados
    - POST /empleados/ - Crear empleado
    - GET /empleados/{id}/ - Detalle de empleado
    - PUT /empleados/{id}/ - Actualizar empleado
    - DELETE /empleados/{id}/ - Eliminar empleado
    """
    
    queryset = Empleado.objects.all()
    permission_classes = [EmpleadoAccessPolicy]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['estado', 'tipo_documento']
    search_fields = ['primer_nombre', 'primer_apellido', 'numero_documento', 'email']
    ordering_fields = ['primer_apellido', 'fecha_ingreso', 'created_at']
    ordering = ['primer_apellido', 'primer_nombre']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return EmpleadoListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return EmpleadoCreateSerializer
        return EmpleadoDetailSerializer
    
    @extend_schema(
        summary="Listar empleados activos",
        description="Retorna solo empleados en estado activo"
    )
    def get_queryset(self):
        queryset = super().get_queryset()
        project = _get_active_project_for_request(self.request)
        if project:
            from dashboard.models import AsignacionProyecto
            empleado_ids = AsignacionProyecto.objects.filter(
                proyecto=project, activo=True
            ).values_list('empleado_id', flat=True)
            queryset = queryset.filter(id__in=empleado_ids)
        return queryset

    @action(detail=False, methods=['get'])
    def activos(self, request):
        """Lista solo empleados activos"""
        queryset = self.get_queryset().filter(estado='activo')
        serializer = EmpleadoListSerializer(queryset, many=True)
        return Response(serializer.data)


# ══════════════════════════════════════════════════════════════════════════════
# VIEWSET: TIPO DE CONTRATO
# ══════════════════════════════════════════════════════════════════════════════

@extend_schema(tags=['Nómina - Configuración'])
class TipoContratoViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para tipos de contrato.
    
    Define las reglas de aplicación de aportes según el tipo de contrato.
    """
    
    queryset = TipoContrato.objects.all()
    serializer_class = TipoContratoSerializer
    permission_classes = [TipoContratoAccessPolicy]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['activo', 'aplica_salud', 'aplica_pension']
    search_fields = ['nombre', 'codigo']
    ordering = ['nombre']


# ══════════════════════════════════════════════════════════════════════════════
# VIEWSET: CONTRATO
# ══════════════════════════════════════════════════════════════════════════════

@extend_schema(tags=['Nómina - Contratos'])
class ContratoViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para contratos laborales.
    
    Define la relación empleado-empresa con salario y configuración.
    """
    
    queryset = Contrato.objects.select_related('empleado', 'tipo_contrato', 'cargo').all()
    permission_classes = [ContratoAccessPolicy]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['empleado', 'tipo_contrato', 'activo', 'nivel_arl', 'proyecto']
    search_fields = ['empleado__primer_nombre', 'empleado__numero_documento', 'cargo__nombre']
    ordering_fields = ['fecha_inicio', 'salario']
    ordering = ['-fecha_inicio']
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ContratoCreateSerializer
        return ContratoSerializer
    
    @extend_schema(
        summary="Contratos activos",
        description="Lista solo contratos vigentes"
    )
    def get_queryset(self):
        queryset = super().get_queryset()
        project = _get_active_project_for_request(self.request)
        if project:
            queryset = queryset.filter(proyecto=project)
        return queryset

    @action(detail=False, methods=['get'])
    def activos(self, request):
        """Lista contratos activos"""
        queryset = self.get_queryset().filter(activo=True)
        serializer = ContratoSerializer(queryset, many=True)
        return Response(serializer.data)
    
    @extend_schema(
        summary="Contratos de un empleado",
        parameters=[
            OpenApiParameter(name='empleado_id', description='ID del empleado', required=True, type=str)
        ]
    )
    @action(detail=False, methods=['get'])
    def por_empleado(self, request):
        """Lista contratos de un empleado específico"""
        empleado_id = request.query_params.get('empleado_id')
        if not empleado_id:
            return Response(
                {'error': 'Se requiere empleado_id'},
                status=status.HTTP_400_BAD_REQUEST
            )
        queryset = self.get_queryset().filter(empleado_id=empleado_id)
        serializer = ContratoSerializer(queryset, many=True)
        return Response(serializer.data)


# ══════════════════════════════════════════════════════════════════════════════
# VIEWSET: PARÁMETRO LEGAL
# ══════════════════════════════════════════════════════════════════════════════

@extend_schema(tags=['Nómina - Configuración'])
class ParametroLegalViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para parámetros legales.
    
    Tabla de configuración para todos los porcentajes legales.
    No requiere valores quemados en código.
    """
    
    queryset = ParametroLegal.objects.all()
    serializer_class = ParametroLegalSerializer
    permission_classes = [ParametrosLegalesAccessPolicy]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['concepto', 'activo']
    search_fields = ['concepto', 'descripcion']
    ordering_fields = ['concepto', 'vigente_desde']
    ordering = ['concepto', '-vigente_desde']

    # Parámetros críticos que el CalculadorNomina necesita para funcionar
    CONCEPTOS_CRITICOS = {
        'SMMLV', 'AUXILIO_TRANSPORTE', 'TOPE_AUXILIO_TRANSPORTE',
        'SALUD', 'PENSION',
        'ARL_NIVEL_I', 'ARL_NIVEL_II', 'ARL_NIVEL_III', 'ARL_NIVEL_IV', 'ARL_NIVEL_V',
        'CAJA_COMPENSACION', 'SENA', 'ICBF',
        'TOPE_FSP', 'TOPE_SUBSISTENCIA', 'FSP', 'SUBSISTENCIA',
    }

    def perform_destroy(self, instance):
        """Impide eliminar parámetros críticos del sistema."""
        if instance.concepto in self.CONCEPTOS_CRITICOS:
            from rest_framework.exceptions import ValidationError
            raise ValidationError({
                'detail': f'No se puede eliminar "{instance.get_concepto_display()}" '
                          f'porque es un parámetro crítico del sistema de nómina. '
                          f'Puedes desactivarlo si no lo necesitas actualmente.'
            })
        instance.delete()
    
    @extend_schema(
        summary="Parámetros vigentes",
        description="Lista solo parámetros activos y vigentes"
    )
    @action(detail=False, methods=['get'])
    def vigentes(self, request):
        """Lista parámetros vigentes"""
        from django.utils import timezone
        from django.db.models import Q
        
        hoy = timezone.now().date()
        queryset = self.get_queryset().filter(
            activo=True,
            vigente_desde__lte=hoy
        ).filter(
            Q(vigente_hasta__isnull=True) | Q(vigente_hasta__gte=hoy)
        )
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


# ══════════════════════════════════════════════════════════════════════════════
# VIEWSET: CONCEPTO LABORAL
# ══════════════════════════════════════════════════════════════════════════════

@extend_schema(tags=['Nómina - Configuración'])
class ConceptoLaboralViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para conceptos laborales.
    
    Define devengados y deducciones que pueden aplicarse en la nómina.
    """
    
    queryset = ConceptoLaboral.objects.all()
    serializer_class = ConceptoLaboralSerializer
    permission_classes = [ConceptoLaboralAccessPolicy]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['tipo', 'activo', 'es_legal', 'aplica_porcentaje']
    search_fields = ['codigo', 'nombre', 'descripcion']
    ordering_fields = ['tipo', 'orden', 'nombre']
    ordering = ['tipo', 'orden']

    # Códigos de conceptos que el sistema necesita para funcionar
    CODIGOS_PROTEGIDOS = {'SALUD_EMPLEADO', 'PENSION_EMPLEADO', 'FSP', 'SUBSISTENCIA', 'RESTAURANTE'}

    def perform_destroy(self, instance):
        """
        Impide eliminar conceptos del sistema o conceptos en uso.
        Los conceptos legales/del sistema solo se pueden desactivar.
        """
        if instance.codigo in self.CODIGOS_PROTEGIDOS:
            from rest_framework.exceptions import ValidationError
            raise ValidationError({
                'detail': f'"{instance.nombre}" es un concepto del sistema y no se puede eliminar. '
                          f'Puedes desactivarlo si no lo necesitas.'
            })
        try:
            instance.delete()
        except ProtectedError:
            from rest_framework.exceptions import ValidationError
            raise ValidationError({
                'detail': f'No se puede eliminar "{instance.nombre}" porque ya fue usado en nóminas calculadas. '
                          f'Puedes desactivarlo en su lugar.'
            })

    def perform_update(self, serializer):
        """
        Impide cambiar el código o tipo de conceptos protegidos del sistema.
        """
        instance = serializer.instance
        if instance.codigo in self.CODIGOS_PROTEGIDOS:
            new_codigo = serializer.validated_data.get('codigo', instance.codigo)
            new_tipo = serializer.validated_data.get('tipo', instance.tipo)
            if new_codigo != instance.codigo or new_tipo != instance.tipo:
                from rest_framework.exceptions import ValidationError
                raise ValidationError({
                    'detail': f'No se puede cambiar el código ni el tipo de "{instance.nombre}" '
                              f'porque es un concepto del sistema.'
                })
        serializer.save()

    @extend_schema(
        summary="Devengados activos",
        description="Lista solo conceptos de tipo devengado activos"
    )
    @action(detail=False, methods=['get'])
    def devengados(self, request):
        """Lista conceptos de tipo DEVENGADO"""
        queryset = self.get_queryset().filter(tipo='DEVENGADO', activo=True)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @extend_schema(
        summary="Deducciones activas",
        description="Lista solo conceptos de tipo deducción activos"
    )
    @action(detail=False, methods=['get'])
    def deducciones(self, request):
        """Lista conceptos de tipo DEDUCCION"""
        queryset = self.get_queryset().filter(tipo='DEDUCCION', activo=True)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


# ══════════════════════════════════════════════════════════════════════════════
# VIEWSET: NÓMINA SIMPLE
# ══════════════════════════════════════════════════════════════════════════════

@extend_schema(tags=['Nómina - Principal'])
class NominaSimpleViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para nóminas.
    
    Modelo principal de nómina con cálculo automático de:
    - IBC según tipo de contrato
    - Aportes de seguridad social
    - Parafiscales
    - Conceptos laborales
    - Descuentos de préstamos
    """
    
    queryset = NominaSimple.objects.select_related(
        'contrato', 'contrato__empleado', 'contrato__tipo_contrato'
    ).prefetch_related('items', 'conceptos', 'prestamos').all()
    permission_classes = [NominaAccessPolicy]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['contrato', 'contrato__empleado', 'estado', 'proyecto']
    search_fields = ['numero', 'contrato__empleado__primer_nombre', 'contrato__empleado__numero_documento']
    ordering_fields = ['periodo_fin', 'created_at', 'total_pagar']
    ordering = ['-periodo_fin', '-created_at']

    def get_queryset(self):
        queryset = super().get_queryset()
        project = _get_active_project_for_request(self.request)
        if project:
            queryset = queryset.filter(proyecto=project)
        return queryset
    
    def get_serializer_class(self):
        if self.action == 'list':
            return NominaSimpleListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return NominaSimpleCreateSerializer
        return NominaSimpleDetailSerializer
    
    @extend_schema(
        summary="Calcular nómina",
        description="Calcula automáticamente todos los valores de la nómina",
        request=CalculoNominaSerializer,
        responses={200: NominaSimpleDetailSerializer}
    )
    @action(detail=True, methods=['post'])
    def calcular(self, request, pk=None):
        """
        Calcula la nómina automáticamente.
        
        Incluye:
        - IBC según tipo de contrato
        - Salud y pensión (empleado y empleador)
        - ARL según nivel de riesgo
        - Parafiscales si aplica
        - Conceptos laborales configurados
        - Descuentos de préstamos
        - Totales
        """
        nomina = self.get_object()
        
        if nomina.estado not in ['borrador', 'calculada']:
            return Response(
                {'error': f'No se puede calcular una nómina en estado {nomina.estado}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Refrescar desde BD para obtener datos más recientes
            nomina.refresh_from_db()
            resumen = calcular_nomina(nomina)
            
            response_data = {
                'mensaje': 'Nómina calculada exitosamente',
                'resumen': resumen,
                'nomina': NominaSimpleDetailSerializer(nomina).data
            }
            
            # Incluir advertencias si existen
            if resumen.get('advertencias'):
                response_data['advertencias'] = resumen['advertencias']
            
            return Response(response_data)
        except NominaValidationError as e:
            return Response(
                {'error': str(e), 'tipo': 'validacion'},
                status=status.HTTP_422_UNPROCESSABLE_ENTITY
            )
        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception(f'Error calculando nómina {nomina.id}')
            return Response(
                {'error': 'Error interno al calcular la nómina. Contacte al administrador.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @extend_schema(
        summary="Aprobar nómina",
        description="Cambia el estado de la nómina a aprobada"
    )
    @action(detail=True, methods=['post'])
    def aprobar(self, request, pk=None):
        """Aprueba la nómina"""
        nomina = self.get_object()
        
        if nomina.estado != 'calculada':
            return Response(
                {'error': 'Solo se pueden aprobar nóminas calculadas'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        nomina.estado = 'aprobada'
        nomina.save()
        
        return Response({
            'mensaje': 'Nómina aprobada exitosamente',
            'nomina': NominaSimpleDetailSerializer(nomina).data
        })
    
    @extend_schema(
        summary="Marcar como pagada",
        description="Cambia el estado de la nómina a pagada"
    )
    @action(detail=True, methods=['post'])
    def pagar(self, request, pk=None):
        """Marca la nómina como pagada"""
        nomina = self.get_object()

        if nomina.estado != 'aprobada':
            return Response(
                {'error': 'Solo se pueden pagar nóminas aprobadas'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que las deducciones no excedan el total devengado
        from decimal import Decimal
        total_deducciones = (nomina.total_deducciones or Decimal('0')) + (nomina.total_prestamos or Decimal('0'))
        if total_deducciones > nomina.total_devengado:
            return Response(
                {'error': 'Las deducciones totales exceden el total devengado. Revise los montos.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from django.utils import timezone
        nomina.estado = 'pagada'
        nomina.fecha_pago = timezone.now().date()
        nomina.save()

        # Actualizar cuotas de préstamos y crear pagos
        from prestamos.models import Prestamo, PagoPrestamo

        for nomina_prestamo in nomina.prestamos.all():
            prestamo = nomina_prestamo.prestamo
            if not prestamo:
                continue

            numero_pago = f'PAG-{prestamo.numero_prestamo}-{nomina_prestamo.numero_cuota:03d}'
            pago_existe = PagoPrestamo.objects.filter(
                prestamo=prestamo,
                numero_pago=numero_pago
            ).exists()

            if pago_existe:
                continue

            saldo_anterior = prestamo.saldo_pendiente or Decimal('0.00')
            saldo_nuevo = max(saldo_anterior - nomina_prestamo.valor_cuota, Decimal('0.00'))

            PagoPrestamo.objects.create(
                organization=nomina.organization,
                prestamo=prestamo,
                numero_pago=numero_pago,
                fecha_pago=nomina.fecha_pago or timezone.now().date(),
                tipo_pago='cuota',
                metodo_pago='descuento_nomina',
                monto_pago=nomina_prestamo.valor_cuota,
                monto_capital=nomina_prestamo.valor_cuota,
                monto_interes=Decimal('0.00'),
                monto_mora=Decimal('0.00'),
                saldo_anterior=saldo_anterior,
                saldo_nuevo=saldo_nuevo,
                observaciones=f'Pago automático vía nómina {nomina.numero}',
                registrado_por=request.user
            )

            Prestamo.objects.filter(pk=prestamo.pk).update(
                saldo_pendiente=saldo_nuevo,
                total_pagado=(prestamo.total_pagado or Decimal('0.00')) + nomina_prestamo.valor_cuota,
                estado='liquidado' if saldo_nuevo == Decimal('0.00') else 'activo'
            )
        
        # Avanzar la fecha de primer pago de los préstamos descontados
        # (esto se hace SOLO al pagar, no al calcular, para evitar desplazar fechas al recalcular)
        from dateutil.relativedelta import relativedelta
        prestamos_procesados = set()
        for nomina_prestamo in nomina.prestamos.all():
            prestamo = nomina_prestamo.prestamo
            if not prestamo or prestamo.id in prestamos_procesados:
                continue
            prestamos_procesados.add(prestamo.id)
            
            # Contar cuántas cuotas se descontaron de este préstamo en esta nómina
            cuotas_descontadas = nomina.prestamos.filter(prestamo=prestamo).count()
            if prestamo.fecha_primer_pago and cuotas_descontadas > 0:
                nueva_fecha = prestamo.fecha_primer_pago + relativedelta(months=cuotas_descontadas)
                Prestamo.objects.filter(pk=prestamo.pk).update(
                    fecha_primer_pago=nueva_fecha
                )
        
        return Response({
            'mensaje': 'Nómina marcada como pagada',
            'nomina': NominaSimpleDetailSerializer(nomina).data
        })
    
    @extend_schema(
        summary="Anular nómina",
        description="Anula la nómina (no se puede deshacer)"
    )
    @action(detail=True, methods=['post'])
    def anular(self, request, pk=None):
        """Anula la nómina"""
        nomina = self.get_object()
        
        if nomina.estado == 'anulada':
            return Response(
                {'error': 'La nómina ya está anulada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if nomina.estado == 'pagada':
            return Response(
                {'error': 'No se puede anular una nómina ya pagada. Debe crear una nómina de ajuste.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        nomina.estado = 'anulada'
        nomina.save()
        
        return Response({
            'mensaje': 'Nómina anulada exitosamente',
            'nomina': NominaSimpleDetailSerializer(nomina).data
        })

    @extend_schema(
        summary="Descargar desprendible",
        description="Genera y descarga el desprendible en PDF"
    )
    @action(detail=True, methods=['get'])
    def desprendible(self, request, pk=None):
        """Genera PDF del desprendible de nómina"""
        from io import BytesIO
        from decimal import Decimal
        from django.http import HttpResponse
        from django.utils.encoding import escape_uri_path
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas

        nomina = self.get_object()

        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        page_w, page_h = letter
        margin = 40
        left = margin
        right = page_w - margin
        top = page_h - margin
        y = top

        def fmt(valor):
            if valor is None:
                return "0"
            if isinstance(valor, Decimal):
                return f"{valor:,.2f}"
            return f"{valor}"

        def draw_title(text):
            nonlocal y
            c.setFont("Helvetica-Bold", 14)
            c.drawString(left, y, text)
            y -= 18

        def draw_section_title(text):
            nonlocal y
            c.setFont("Helvetica-Bold", 11)
            c.drawString(left, y, text)
            y -= 12
            c.setLineWidth(0.5)
            c.line(left, y, right, y)
            y -= 8

        def draw_kv(x, y_pos, label, value):
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x, y_pos, f"{label}:")
            c.setFont("Helvetica", 9)
            c.drawString(x + 90, y_pos, str(value))

        def draw_table(x, y_pos, width, title, rows):
            row_h = 14
            header_h = 16
            c.setFont("Helvetica-Bold", 10)
            c.drawString(x, y_pos, title)
            y_pos -= 6
            c.setLineWidth(0.5)
            c.rect(x, y_pos - header_h, width, header_h, stroke=1, fill=0)
            c.setFont("Helvetica-Bold", 8)
            c.drawString(x + 4, y_pos - 12, "Concepto")
            c.drawString(x + width * 0.62, y_pos - 12, "Cantidad")
            c.drawRightString(x + width - 4, y_pos - 12, "Valor")
            y_pos -= header_h

            c.setFont("Helvetica", 8)
            for concepto, cantidad, valor in rows:
                c.rect(x, y_pos - row_h, width, row_h, stroke=1, fill=0)
                c.drawString(x + 4, y_pos - 11, str(concepto))
                c.drawString(x + width * 0.62, y_pos - 11, str(cantidad) if cantidad else "-")
                c.drawRightString(x + width - 4, y_pos - 11, fmt(valor))
                y_pos -= row_h
            return y_pos

        # Encabezado
        draw_title(f"Desprendible de Nómina {nomina.numero}")
        c.setFont("Helvetica", 10)
        c.drawString(left, y, f"Período: {nomina.periodo_inicio} a {nomina.periodo_fin}")
        y -= 12
        c.drawString(left, y, f"Fecha de Pago: {nomina.fecha_pago or ''}")
        y -= 12
        c.drawString(left, y, f"Estado: {nomina.get_estado_display()}")
        y -= 18

        # Información del empleado y contrato en dos columnas
        col_gap = 20
        col_w = (right - left - col_gap) / 2
        col1_x = left
        col2_x = left + col_w + col_gap
        info_y = y

        c.setFont("Helvetica-Bold", 10)
        c.drawString(col1_x, info_y, "Empleado")
        c.drawString(col2_x, info_y, "Contrato")
        info_y -= 12

        draw_kv(col1_x, info_y, "Nombre", nomina.empleado.nombre_completo)
        draw_kv(col2_x, info_y, "Tipo", nomina.contrato.tipo_contrato.nombre)
        info_y -= 12
        draw_kv(col1_x, info_y, "Documento", nomina.empleado.numero_documento)
        draw_kv(col2_x, info_y, "Salario Base", fmt(nomina.salario_base))
        info_y -= 12
        contrato_vigencia = (
            f"{nomina.contrato.fecha_inicio} - {nomina.contrato.fecha_fin}"
            if nomina.contrato.fecha_fin
            else "Indefinido"
        )
        draw_kv(col1_x, info_y, "Contrato", contrato_vigencia)
        draw_kv(col2_x, info_y, "Incluye Salario", "Sí" if nomina.incluir_salario_base else "No")
        info_y -= 18

        y = info_y
        draw_section_title("Ingresos y Deducciones")

        # Construcción de filas
        ingresos_rows = []
        total_items_qty = sum((item.cantidad or 0) for item in nomina.items.all())
        if nomina.total_items and nomina.total_items > 0:
            ingresos_rows.append(("Items de trabajo", total_items_qty, nomina.total_items))
        if nomina.incluir_salario_base or nomina.total_items == 0:
            ingresos_rows.append(("Salario base", "1", nomina.salario_base))

        devengados = nomina.conceptos.filter(tipo='DEVENGADO')
        for cpto in devengados:
            cantidad = f"{cpto.porcentaje_aplicado}%" if cpto.porcentaje_aplicado and cpto.porcentaje_aplicado > 0 else "1"
            ingresos_rows.append((cpto.concepto.nombre, cantidad, cpto.valor))

        deducciones_rows = []
        deducciones = nomina.conceptos.filter(tipo='DEDUCCION')
        for cpto in deducciones:
            cantidad = f"{cpto.porcentaje_aplicado}%" if cpto.porcentaje_aplicado and cpto.porcentaje_aplicado > 0 else "1"
            deducciones_rows.append((cpto.concepto.nombre, cantidad, cpto.valor))
        if nomina.total_prestamos and nomina.total_prestamos > 0:
            deducciones_rows.append(("Préstamos", "1", nomina.total_prestamos))

        table_top = y
        left_table_x = left
        right_table_x = left + col_w + col_gap
        table_width = col_w

        y_left = draw_table(left_table_x, table_top, table_width, "Ingresos", ingresos_rows)
        y_right = draw_table(right_table_x, table_top, table_width, "Deducciones", deducciones_rows)
        y = min(y_left, y_right) - 14

        # Totales
        c.setFont("Helvetica-Bold", 10)
        c.drawString(left, y, f"Total Ingresos: {fmt(nomina.total_devengado)}")
        c.drawRightString(right, y, f"Total Deducciones: {fmt(nomina.total_deducciones)}")
        y -= 18
        c.setFont("Helvetica-Bold", 12)
        c.drawRightString(right, y, f"NETO A PAGAR: {fmt(nomina.total_pagar)}")
        y -= 20

        # Detalle de items
        if y < 120:
            c.showPage()
            y = top
        draw_section_title("Detalle de Items")
        for item in nomina.items.all():
            if y < 80:
                c.showPage()
                y = top
            c.setFont("Helvetica", 9)
            c.drawString(left, y, f"{item.item.nombre}")
            c.drawRightString(right - 120, y, f"Cantidad: {item.cantidad}")
            c.drawRightString(right, y, f"Total: {fmt(item.valor_total)}")
            y -= 12

        c.showPage()
        c.save()

        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        safe_name = f"desprendible_{nomina.numero}.pdf"
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{escape_uri_path(safe_name)}"
        return response
    
    @extend_schema(
        summary="Nóminas por período",
        parameters=[
            OpenApiParameter(name='periodo_inicio', description='Fecha inicio', required=True, type=str),
            OpenApiParameter(name='periodo_fin', description='Fecha fin', required=True, type=str),
        ]
    )
    @action(detail=False, methods=['get'])
    def por_periodo(self, request):
        """Lista nóminas de un período específico"""
        periodo_inicio = request.query_params.get('periodo_inicio')
        periodo_fin = request.query_params.get('periodo_fin')
        
        if not periodo_inicio or not periodo_fin:
            return Response(
                {'error': 'Se requieren periodo_inicio y periodo_fin'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = self.get_queryset().filter(
            periodo_inicio=periodo_inicio,
            periodo_fin=periodo_fin
        ).exclude(estado='anulada')
        
        serializer = NominaSimpleListSerializer(queryset, many=True)
        return Response(serializer.data)

    @extend_schema(
        summary="Estadísticas de nóminas",
        description="Retorna estadísticas generales de nóminas"
    )
    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Retorna estadísticas de nóminas"""
        from django.db.models import Sum, Avg, Count
        from django.db.models.functions import Coalesce
        from decimal import Decimal
        
        queryset = self.get_queryset().exclude(estado='anulada')
        
        # Calcular estadísticas
        stats = queryset.aggregate(
            total_nominas=Count('id'),
            total_pagado=Coalesce(Sum('total_pagar'), Decimal('0')),
            promedio_por_nomina=Coalesce(Avg('total_pagar'), Decimal('0')),
            total_devengado=Coalesce(Sum('total_devengado'), Decimal('0')),
            total_deducciones=Coalesce(Sum('total_deducciones'), Decimal('0')),
        )
        
        # Empleados/contratos con nómina
        empleados_con_nomina = queryset.values('contrato__empleado').distinct().count()
        
        # Estadísticas por estado
        por_estado = queryset.values('estado').annotate(
            cantidad=Count('id'),
            total=Coalesce(Sum('total_pagar'), Decimal('0'))
        )
        
        return Response({
            'total_nominas': stats['total_nominas'],
            'total_pagado': float(stats['total_pagado']),
            'promedio_por_nomina': float(stats['promedio_por_nomina']),
            'total_devengado': float(stats['total_devengado']),
            'total_deducciones': float(stats['total_deducciones']),
            'empleados_con_nomina': empleados_con_nomina,
            'por_estado': list(por_estado)
        })

    @extend_schema(
        summary="Exportar nóminas a Excel",
        description="Exporta el listado de nóminas filtrado a Excel"
    )
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exportar nóminas a Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse

        MAX_EXPORT_ROWS = 5000
        queryset = self.filter_queryset(self.get_queryset())[:MAX_EXPORT_ROWS]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Nominas"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = [
            'Número', 'Empleado', 'Documento', 'Tipo Contrato',
            'Período Inicio', 'Período Fin', 'Estado',
            'Total Devengado', 'Total Deducciones', 'Total Pagar',
            'Fecha Pago'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row, nomina in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=nomina.numero)
            ws.cell(row=row, column=2, value=nomina.contrato.empleado.nombre_completo)
            ws.cell(row=row, column=3, value=nomina.contrato.empleado.numero_documento)
            ws.cell(row=row, column=4, value=nomina.contrato.tipo_contrato.nombre)
            ws.cell(row=row, column=5, value=str(nomina.periodo_inicio))
            ws.cell(row=row, column=6, value=str(nomina.periodo_fin))
            ws.cell(row=row, column=7, value=nomina.get_estado_display())
            ws.cell(row=row, column=8, value=float(nomina.total_devengado or 0))
            ws.cell(row=row, column=9, value=float(nomina.total_deducciones or 0))
            ws.cell(row=row, column=10, value=float(nomina.total_pagar or 0))
            ws.cell(row=row, column=11, value=str(nomina.fecha_pago) if nomina.fecha_pago else '')

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=nominas.xlsx'
        wb.save(response)
        return response


# ══════════════════════════════════════════════════════════════════════════════
# VIEWSET: NÓMINA ITEM
# ══════════════════════════════════════════════════════════════════════════════

@extend_schema(tags=['Nómina - Items'])
class NominaItemViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para items de nómina.
    
    Representa el trabajo realizado (producción) asociado a una nómina.
    """
    
    queryset = NominaItem.objects.select_related('item', 'nomina').all()
    serializer_class = NominaItemSerializer
    permission_classes = [NominaAccessPolicy]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['nomina', 'item']
    search_fields = ['item__nombre']
    ordering = ['item__nombre']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        project = _get_active_project_for_request(self.request)
        if project:
            queryset = queryset.filter(nomina__proyecto=project)
        return queryset
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return NominaItemCreateSerializer
        return NominaItemSerializer
    
    def perform_create(self, serializer):
        """Al crear, copiar el valor unitario del item"""
        item = serializer.validated_data['item']
        valor_unitario = serializer.validated_data.get('valor_unitario', item.precio_unitario)
        serializer.save(
            organization=self.request.user.organization,
            valor_unitario=valor_unitario
        )


# ══════════════════════════════════════════════════════════════════════════════
# VIEWSET: NÓMINA CONCEPTO
# ══════════════════════════════════════════════════════════════════════════════

@extend_schema(tags=['Nómina - Conceptos'])
class NominaConceptoViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para conceptos de nómina.
    
    Almacena los devengados y deducciones calculados para cada nómina.
    """
    
    queryset = NominaConcepto.objects.select_related('concepto', 'nomina').all()
    serializer_class = NominaConceptoSerializer
    permission_classes = [NominaAccessPolicy]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['nomina', 'concepto', 'tipo']
    search_fields = ['concepto__nombre', 'concepto__codigo']
    ordering = ['tipo', 'concepto__orden']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        project = _get_active_project_for_request(self.request)
        if project:
            queryset = queryset.filter(nomina__proyecto=project)
        return queryset
